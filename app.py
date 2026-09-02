import collections
import os
import re
import json
import base64
import urllib.request
import ssl
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import Flask, request, jsonify, render_template, send_file, redirect, url_for, session, flash
import pandas as pd
import openpyxl
from pypdf import PdfReader
import pymupdf
from dotenv import load_dotenv
import io
import psycopg2
import psycopg2.extras
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from master_data import MASTER_BRANCHES, MASTER_VENDORS, get_branch_state, match_master_vendor, normalize_gstin

# Load environment variables
load_dotenv()

app = Flask(__name__, static_folder='static', template_folder='templates')

# DB tables initialized after init_db definition

app.secret_key = os.getenv("FLASK_SECRET_KEY", "949539d0c64bdf34138e6be019a552bf")

# SSL context for API calls
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

# Read API Key
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

# Ultra-fast high-accuracy vision model via OpenRouter for OCR scanning.
# google/gemini-2.5-flash achieves ~1.5s response times with state-of-the-art
# OCR resolution and exact 15-character GSTIN pattern precision.
AI_VISION_MODEL_NAME = "google/gemini-2.5-flash"
AI_MODEL_NAME = "google/gemini-2.5-flash"
AI_MODEL_DISPLAY_NAME = "Gemini 2.5 Flash (Ultra-Fast)"

# High-accuracy vision model for forensic re-scanning of incomplete bills / handwriting / faint stamps.
AI_RESCAN_VISION_MODEL = "google/gemini-2.5-pro"
AI_RESCAN_MODEL_DISPLAY_NAME = "Gemini 2.5 Pro (High Accuracy)"

# Reliable fallback vision models if primary provider is unreachable.
AI_VISION_FALLBACK_MODEL_NAME = "x-ai/grok-4.6"
AI_VISION_SECONDARY_FALLBACK = "claude-3-5-sonnet-20241022"

@app.context_processor
def inject_global_template_vars():
    return {
        'api_key_configured': bool(ANTHROPIC_API_KEY or OPENROUTER_API_KEY),
        'ai_model_name': AI_MODEL_DISPLAY_NAME
    }

def render_pdf_page_to_png_base64(file_bytes, page_index=0, dpi=150):
    """Renders one PDF page to a base64 JPEG at an optimized 150 DPI and 85% quality,
    reducing network payload size by over 90% (~200KB vs 6MB) while retaining
    ultra-crisp clarity for fine-print and handwritten GSTIN / date stamps."""
    doc = pymupdf.open(stream=file_bytes, filetype="pdf")
    page = doc[page_index]
    pix = page.get_pixmap(dpi=dpi)
    jpg_bytes = pix.tobytes("jpg", jpg_quality=85)
    doc.close()
    return base64.b64encode(jpg_bytes).decode("utf-8")

def optimize_image_bytes(file_bytes, ext):
    """Optimizes uploaded image files (PNG/JPG/WEBP) by re-encoding as high-quality
    JPEG at quality 85, reducing huge camera photo payloads from 10MB+ down to ~200KB."""
    try:
        pix = pymupdf.Pixmap(file_bytes)
        jpg_bytes = pix.tobytes("jpg", jpg_quality=85)
        return jpg_bytes, "image/jpeg"
    except Exception:
        mime = f"image/{ext}"
        if ext == 'jpg':
            mime = "image/jpeg"
        return file_bytes, mime

# PostgreSQL Connection Helper
def get_db_connection():
    conn = psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=os.getenv("DB_PORT", "5432"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASSWORD", "Jay@0381"),
        database=os.getenv("DB_NAME", "postgres")
    )
    return conn

import datetime

FY_MONTH_ORDER = ['April', 'May', 'June', 'July', 'August', 'September',
                   'October', 'November', 'December', 'January', 'February', 'March']

def month_sort_key(month_name):
    """Sorts month names in financial-year order (April..March) instead of alphabetically."""
    try:
        return FY_MONTH_ORDER.index(month_name)
    except ValueError:
        return len(FY_MONTH_ORDER)

def _dt_to_fy_and_month(dt):
    if dt.month in (1, 2, 3):
        fy = f"{dt.year - 1}-{str(dt.year)[2:]}"
    else:
        fy = f"{dt.year}-{str(dt.year + 1)[2:]}"
    return fy, dt.strftime('%B')

def parse_date_to_fy_and_month(date_str):
    """
    Parses date strings in a range of formats commonly seen in scanned
    invoices, AI-extracted text, and GSTR-2B exports (e.g. '2026-08-20',
    '20-08-2026', '20 Aug 2026', '2026-08-20T00:00:00Z').
    Returns (financial_year_str, month_str) or (None, None) if invalid.
    Financial Year runs from April 1 to March 31.
    """
    if not date_str or date_str in ('N/A', '-', 'None'):
        return None, None

    raw = date_str.strip()
    # Strip a trailing time-of-day component, whether space- or T-separated
    date_only = raw.split(' ')[0].split('T')[0]

    numeric_formats = ('%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y', '%d.%m.%Y',
                        '%d-%m-%y', '%d/%m/%y', '%d.%m.%y')
    for fmt in numeric_formats:
        try:
            dt = datetime.datetime.strptime(date_only, fmt)
            return _dt_to_fy_and_month(dt)
        except ValueError:
            continue

    # Formats with a textual month need the full string (may contain spaces).
    # Two-digit-year variants (e.g. '06-MAR-26', a common scanned-invoice
    # date stamp) are included alongside the 4-digit-year ones.
    textual_formats = ('%d %B %Y', '%d %b %Y', '%d-%b-%Y', '%d-%B-%Y', '%B %d, %Y', '%b %d, %Y',
                        '%d %b %y', '%d %B %y', '%d-%b-%y', '%d-%B-%y')
    for fmt in textual_formats:
        try:
            dt = datetime.datetime.strptime(raw, fmt)
            return _dt_to_fy_and_month(dt)
        except ValueError:
            continue

    return None, None


CLIENTS_CONFIG = {
    'nutan_nagrik': {
        'id': 'nutan_nagrik',
        'name': 'Nutan Nagrik Sahakari Bank Ltd.',
        'short_name': 'Nutan Nagrik Bank',
        'type': 'banking',
        'itc_rule': 'section_17_4_50',
        'itc_claim_pct': 50,
        'rule_title': 'CGST Act Section 17(4) Compliance',
        'rule_description': 'Banking companies opting for this method must claim exactly 50% of the eligible Input Tax Credit (ITC) on inputs, capital goods, and input services. The remaining 50% is treated as ineligible credit and lapses.',
        'eligible_card_title': 'Eligible ITC (50%)',
        'ineligible_card_title': 'Ineligible ITC (50%)',
        'icon': 'fa-building-columns',
        'tag': 'Active Client'
    },
    'sun_builders': {
        'id': 'sun_builders',
        'name': 'Sun Builders',
        'short_name': 'Sun Builders',
        'type': 'builder_realestate',
        'itc_rule': 'standard_100',
        'itc_claim_pct': 100,
        'rule_title': 'CGST Act Section 16 & 17(5) Compliance (Real Estate)',
        'rule_description': 'Standard Real Estate & Infrastructure ITC: Claim 100% eligible Input Tax Credit on qualifying commercial procurement and services, subject to Section 17(5) blocked credit rules (construction materials, motor vehicles, etc.).',
        'eligible_card_title': 'Eligible ITC (100%)',
        'ineligible_card_title': 'Blocked / Ineligible ITC',
        'icon': 'fa-city',
        'tag': 'Client 2'
    },
    'client_3': {
        'id': 'client_3',
        'name': 'Client 3',
        'short_name': 'Client 3',
        'type': 'commercial',
        'itc_rule': 'standard_100',
        'itc_claim_pct': 100,
        'rule_title': 'Standard Corporate GST ITC Compliance',
        'rule_description': 'Standard 100% ITC claim workspace for commercial trading and corporate operations.',
        'eligible_card_title': 'Eligible ITC',
        'ineligible_card_title': 'Blocked ITC',
        'icon': 'fa-building',
        'tag': 'Coming Soon'
    }
}

def get_current_client_id():
    client_id = request.args.get('client_id') or request.headers.get('X-Client-Id')
    if not client_id and request.is_json:
        try:
            body = request.get_json(silent=True) or {}
            client_id = body.get('client_id')
        except Exception:
            pass
    if not client_id:
        client_id = session.get('active_client_id')
    if not client_id or client_id not in CLIENTS_CONFIG:
        client_id = 'nutan_nagrik'
    session['active_client_id'] = client_id
    return client_id

def get_client_config(client_id=None):
    cid = client_id or get_current_client_id()
    return CLIENTS_CONFIG.get(cid, CLIENTS_CONFIG['nutan_nagrik'])

# Database Tables Initialization
def init_db():
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Create users table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                is_admin BOOLEAN NOT NULL DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')

        # Migrate existing installs that predate the is_admin column
        cur.execute('ALTER TABLE users ADD COLUMN IF NOT EXISTS is_admin BOOLEAN NOT NULL DEFAULT FALSE;')


        # Create invoices table (using numeric for high-precision currency values)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                invoice_number VARCHAR(50),
                invoice_date VARCHAR(20),
                vendor_name VARCHAR(150),
                taxable_value NUMERIC(15,2),
                cgst NUMERIC(15,2),
                sgst NUMERIC(15,2),
                igst NUMERIC(15,2),
                eligible_itc NUMERIC(15,2),
                ineligible_itc NUMERIC(15,2),
                file_data BYTEA,
                file_mime_type VARCHAR(100),
                file_name VARCHAR(255),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')

        # Migrate existing installs that predate the original-file columns
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS file_data BYTEA;')
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS file_mime_type VARCHAR(100);')
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS file_name VARCHAR(255);')

        # Migrate existing installs that predate the branch/GSTIN columns
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS branch VARCHAR(100);')
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS gstin VARCHAR(20);')

        # Migrate existing installs that predate the state (GST registration --
        # Gujarat vs Maharashtra) column. Distinct from branch: the same branch
        # name can exist under both state registrations, and GSTR-2B is filed
        # separately per state.
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS state VARCHAR(50);')
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS financial_year VARCHAR(10);')
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS month VARCHAR(20);')

        # Migrate existing installs that predate ITC-blocked tracking and the
        # stamped/handwritten payment date (distinct from the printed invoice date)
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS itc_blocked BOOLEAN NOT NULL DEFAULT FALSE;')
        cur.execute('ALTER TABLE invoices ADD COLUMN IF NOT EXISTS payment_date VARCHAR(20);')

        # Create GSTR-2B table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS gstr2b_entries (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                financial_year VARCHAR(10) NOT NULL,
                month VARCHAR(20) NOT NULL,
                supplier_gstin VARCHAR(20) NOT NULL,
                supplier_name VARCHAR(150),
                invoice_number VARCHAR(50) NOT NULL,
                invoice_date VARCHAR(20),
                taxable_value NUMERIC(15,2),
                cgst NUMERIC(15,2),
                sgst NUMERIC(15,2),
                igst NUMERIC(15,2),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')

        # Migrate existing installs that predate the state (GST registration) column
        cur.execute('ALTER TABLE gstr2b_entries ADD COLUMN IF NOT EXISTS state VARCHAR(50);')

        # Create Filing History / activity log table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS activity_log (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                action VARCHAR(50) NOT NULL,
                description VARCHAR(255),
                financial_year VARCHAR(10),
                month VARCHAR(20),
                record_count INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')



        # Income Statements & Output GST Module Table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS income_entries (
                id SERIAL PRIMARY KEY,
                user_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                client_id VARCHAR(50) NOT NULL DEFAULT 'nutan_nagrik',
                branch VARCHAR(100) NOT NULL,
                state VARCHAR(50) DEFAULT 'Gujarat',
                financial_year VARCHAR(10) NOT NULL,
                month VARCHAR(20) NOT NULL,
                gl_code VARCHAR(50) NOT NULL,
                particulars VARCHAR(255),
                is_taxable BOOLEAN DEFAULT TRUE,
                income_amount NUMERIC(15,2) DEFAULT 0.0,
                cgst NUMERIC(15,2) DEFAULT 0.0,
                sgst NUMERIC(15,2) DEFAULT 0.0,
                igst NUMERIC(15,2) DEFAULT 0.0,
                refund_without_gst NUMERIC(15,2) DEFAULT 0.0,
                refund_with_gst NUMERIC(15,2) DEFAULT 0.0,
                file_name VARCHAR(255),
                file_data BYTEA,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
        ''')
        # Deduplicate any existing duplicate entries in income_entries
        cur.execute('''
            DELETE FROM income_entries a USING income_entries b
            WHERE a.id < b.id 
              AND a.client_id = b.client_id 
              AND UPPER(a.branch) = UPPER(b.branch)
              AND a.financial_year = b.financial_year 
              AND a.month = b.month 
              AND a.gl_code = b.gl_code;
        ''')
        cur.execute("CREATE INDEX IF NOT EXISTS idx_income_client_fy ON income_entries(client_id, financial_year, month);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_income_branch ON income_entries(client_id, branch);")
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS uq_income_branch_code_month ON income_entries(client_id, branch, financial_year, month, gl_code);")

                # Multi-client data isolation schema
        cur.execute("ALTER TABLE invoices ADD COLUMN IF NOT EXISTS client_id VARCHAR(50) NOT NULL DEFAULT 'nutan_nagrik';")
        cur.execute("ALTER TABLE gstr2b_entries ADD COLUMN IF NOT EXISTS client_id VARCHAR(50) NOT NULL DEFAULT 'nutan_nagrik';")
        cur.execute("ALTER TABLE activity_log ADD COLUMN IF NOT EXISTS client_id VARCHAR(50) NOT NULL DEFAULT 'nutan_nagrik';")
        cur.execute("UPDATE invoices SET client_id = 'nutan_nagrik' WHERE client_id IS NULL OR client_id = '' OR client_id = 'default';")
        cur.execute("UPDATE gstr2b_entries SET client_id = 'nutan_nagrik' WHERE client_id IS NULL OR client_id = '' OR client_id = 'default';")
        cur.execute("UPDATE activity_log SET client_id = 'nutan_nagrik' WHERE client_id IS NULL OR client_id = '' OR client_id = 'default';")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_invoices_client ON invoices(client_id, user_id);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_gstr2b_client ON gstr2b_entries(client_id, user_id);")

        conn.commit()

        # Auto-seed a default user 'admin' if the users table is empty
        cur.execute('SELECT COUNT(*) FROM users;')
        count = cur.fetchone()[0]
        if count == 0:
            default_username = 'admin'
            default_password = 'admin'
            hashed_pw = generate_password_hash(default_password)
            cur.execute('INSERT INTO users (username, password_hash, is_admin) VALUES (%s, %s, TRUE)', (default_username, hashed_pw))
            conn.commit()
            print("Seeded default admin user account: admin / admin")

        # Ensure the 'admin' account is always flagged as admin, even on installs
        # created before the is_admin column existed
        cur.execute("UPDATE users SET is_admin = TRUE WHERE username = 'admin' AND is_admin = FALSE;")
        conn.commit()

        # Backfill NULL fields on invoices saved before extracted-field
        # normalization existed, which otherwise crash the table renderer
        cur.execute("UPDATE invoices SET vendor_name = 'Unknown Vendor' WHERE vendor_name IS NULL;")
        cur.execute("UPDATE invoices SET invoice_number = 'N/A' WHERE invoice_number IS NULL;")
        cur.execute("UPDATE invoices SET invoice_date = 'N/A' WHERE invoice_date IS NULL;")
        cur.execute("UPDATE invoices SET taxable_value = 0 WHERE taxable_value IS NULL;")
        cur.execute("UPDATE invoices SET cgst = 0 WHERE cgst IS NULL;")
        cur.execute("UPDATE invoices SET sgst = 0 WHERE sgst IS NULL;")
        cur.execute("UPDATE invoices SET igst = 0 WHERE igst IS NULL;")
        cur.execute("UPDATE invoices SET eligible_itc = 0 WHERE eligible_itc IS NULL;")
        cur.execute("UPDATE invoices SET ineligible_itc = 0 WHERE ineligible_itc IS NULL;")
        cur.execute("UPDATE invoices SET branch = 'Unassigned' WHERE branch IS NULL OR branch = '';")
        cur.execute("UPDATE invoices SET gstin = 'N/A' WHERE gstin IS NULL OR gstin = '';")
        cur.execute("UPDATE invoices SET state = 'Unassigned' WHERE state IS NULL OR state = '';")
        cur.execute("UPDATE gstr2b_entries SET state = 'Unassigned' WHERE state IS NULL OR state = '';")
        conn.commit()

        # Backfill financial_year and month for old invoices
        cur.execute("SELECT id, invoice_date FROM invoices WHERE financial_year IS NULL OR month IS NULL;")
        old_invoices = cur.fetchall()
        for row in old_invoices:
            inv_id, inv_date = row
            fy, m = parse_date_to_fy_and_month(inv_date)
            if fy and m:
                cur.execute("UPDATE invoices SET financial_year = %s, month = %s WHERE id = %s;", (fy, m, inv_id))
        conn.commit()

        cur.close()
        conn.close()
        print("PostgreSQL Database initialized successfully.")
    except Exception as e:
        print(f"Error initializing PostgreSQL database: {e}")

# Run database initialization & migrations on application startup
try:
    init_db()
except Exception as e:
    print(f"Startup DB init error: {e}")

# Decorator to secure endpoints
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not is_admin_user():
            flash("Access denied: Administrator privileges required.", "error")
            return redirect(url_for('home'))
        return f(*args, **kwargs)
    return decorated_function

def is_admin_user():
    """Checks admin status directly from the database rather than trusting a
    value cached in the session cookie, so role changes take effect immediately
    without requiring the user to log out and back in."""
    if 'user_id' not in session:
        return False
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT is_admin FROM users WHERE id = %s', (session['user_id'],))
        row = cur.fetchone()
        cur.close()
        conn.close()
        return bool(row[0]) if row else False
    except Exception:
        return False

def log_activity(user_id, action, description=None, fy=None, month=None, record_count=None):
    """Best-effort audit-trail entry for Filing History. Never raises -- a
    logging failure shouldn't block the export/upload it's recording."""
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO activity_log (user_id, action, description, financial_year, month, record_count)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (user_id, action, description, fy, month, record_count))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Error logging activity: {e}")

def call_claude_api(payload):
    """Utility to make direct HTTP requests to the Anthropic Claude API."""
    url = "https://api.anthropic.com/v1/messages"
    headers = {
        "x-api-key": ANTHROPIC_API_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json"
    }
    
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST"
    )
    
    try:
        with urllib.request.urlopen(req, context=ctx) as res:
            response = json.loads(res.read().decode("utf-8"))
            # Extended-thinking models (e.g. claude-opus-5) return a leading
            # "thinking" block before the actual "text" block, so the text
            # response isn't reliably at a fixed index -- find it by type.
            for block in response["content"]:
                if block.get("type") == "text":
                    return block["text"]
            raise ValueError(f"No text block in response content: {response['content']}")
    except Exception as e:
        print(f"Error calling Anthropic API: {e}")
        raise e

def call_openrouter_api(payload):
    """Utility to make direct HTTP requests to OpenRouter's OpenAI-compatible
    chat completions endpoint -- used for vision/OCR (see AI_VISION_MODEL_NAME),
    kept separate from call_claude_api() since the request/response shape
    differs (OpenAI chat format vs Anthropic's messages format)."""
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json"
    }

    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers=headers,
        method="POST"
    )

    try:
        with urllib.request.urlopen(req, context=ctx, timeout=90) as res:
            response = json.loads(res.read().decode("utf-8"))
            if "error" in response:
                raise ValueError(f"OpenRouter error: {response['error']}")
            return response["choices"][0]["message"]["content"]
    except Exception as e:
        print(f"Error calling OpenRouter API: {e}")
        raise e

def call_vision_model(system_prompt, user_prompt, base64_data, mime_type):
    """Calls the ultra-fast primary vision model (Gemini 2.5 Flash via OpenRouter)
    and, if that fails for any reason, falls back to Grok 4.6 (OpenRouter), then
    retries Gemini 2.5 Flash once more via OpenRouter as a last resort (covers a
    transient rate-limit/empty-response blip on the primary attempt). Returns
    (text, model_used). Deliberately stays on OpenRouter throughout -- the old
    tertiary fallback called Claude directly via Anthropic's API using a model
    id that no longer exists there, so every page that reached it hard-failed
    with a 404 instead of actually falling back to anything."""
    # 1. Primary: Google Gemini 2.5 Flash via OpenRouter (~1.5s)
    # reasoning:effort=none turns off Gemini's internal "thinking" pass --
    # without it, hidden reasoning tokens can silently eat the whole
    # max_tokens budget before the model ever emits the actual JSON, leaving
    # nothing (or a truncated fragment) for the visible answer. This is both
    # the fix for those empty/truncated responses and the main lever on
    # per-call latency, since a straight field-extraction task like this
    # never needed a reasoning pass to begin with.
    openrouter_payload = {
        "model": AI_VISION_MODEL_NAME,
        "max_tokens": 3000,
        "reasoning": {"effort": "none"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{base64_data}"}},
                    {"type": "text", "text": user_prompt}
                ]
            }
        ]
    }
    try:
        return call_openrouter_api(openrouter_payload), AI_VISION_MODEL_NAME
    except Exception as e:
        print(f"Primary vision model ({AI_VISION_MODEL_NAME}) failed: {e}")

    # 2. Secondary Fallback: Grok 4.6 via OpenRouter
    try:
        fallback_payload = dict(openrouter_payload)
        fallback_payload["model"] = AI_VISION_FALLBACK_MODEL_NAME
        return call_openrouter_api(fallback_payload), AI_VISION_FALLBACK_MODEL_NAME
    except Exception as e:
        print(f"Secondary vision fallback ({AI_VISION_FALLBACK_MODEL_NAME}) failed: {e}")

    # 3. Tertiary Fallback: retry Gemini 2.5 Flash via OpenRouter once more
    return call_openrouter_api(openrouter_payload), AI_VISION_MODEL_NAME

def call_rescan_vision_model(system_prompt, user_prompt, base64_data, mime_type):
    """Uses a higher-capacity reasoning vision model (e.g. Gemini 2.5 Pro or Claude 3.5 Sonnet)
    specifically designed for recovering ambiguous handwritten notes, fine-print GSTINs,
    and payment date stamps with maximum precision."""
    # 1. Primary higher model: Gemini 2.5 Pro via OpenRouter
    # reasoning:effort=none avoids Gemini's hidden "thinking" pass consuming
    # the whole max_tokens budget before it ever writes the actual JSON --
    # confirmed live on invoice #542 ("Expecting property name enclosed in
    # double quotes", "Unterminated string...") on exactly the kind of
    # complex, heavily handwritten/stamped bill this model exists to
    # handle. Also raised max_tokens for the same reason: a genuinely
    # thorough forensic re-read needs room to write out a longer answer.
    openrouter_payload = {
        "model": AI_RESCAN_VISION_MODEL,
        "max_tokens": 3000,
        "reasoning": {"effort": "none"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{base64_data}"}},
                    {"type": "text", "text": user_prompt}
                ]
            }
        ]
    }
    try:
        return call_openrouter_api(openrouter_payload), AI_RESCAN_VISION_MODEL
    except Exception as e:
        print(f"Rescan primary vision model ({AI_RESCAN_VISION_MODEL}) failed: {e}")

    # 2. Fallback to Claude 3.5 Sonnet
    try:
        fallback_payload = dict(openrouter_payload)
        fallback_payload["model"] = "anthropic/claude-3.5-sonnet"
        return call_openrouter_api(fallback_payload), "anthropic/claude-3.5-sonnet"
    except Exception as e:
        print(f"Rescan secondary fallback failed: {e}")

    # 3. Fallback to default vision model pipeline (Gemini 2.5 Flash / Grok / Claude)
    return call_vision_model(system_prompt, user_prompt, base64_data, mime_type)

def extract_from_pdf_binary_rescan(file_bytes, page_index=0):
    """Renders the specified PDF page at high-resolution 250 DPI for forensic clarity
    and runs the high-accuracy vision model to recover missing fields."""
    system_prompt = (
        "You are a senior forensic financial OCR specialist examining Indian tax invoices, "
        "utility bills, bank debit vouchers, vendor payment receipts, and purchase registers. "
        "Your task is to thoroughly analyze the document image with forensic attention to detail, "
        "especially for MISSING, BLANK, HANDWRITTEN, STAMPED, or FAINT data fields.\n"
        "You must respond with ONLY a valid JSON object. Do not include any explanations outside JSON."
    )
    user_prompt = (
        "Perform a meticulous high-accuracy extraction of invoice details, reading BOTH PRINTED and HANDWRITTEN entries:\n"
        "- invoice_number: The bill, invoice, cash memo, or challan number.\n"
        "- invoice_date: Date of invoice issue (standardize to DD-MM-YYYY format).\n"
        "- payment_date: The date the bill was actually PAID / passed. Search diligently for rubber stamps ("
        "'PAID', 'SANCTIONED', 'PASSED FOR PAYMENT', 'CHEQUE NO', 'RTGS/NEFT', 'DEBITED ON'), "
        "handwritten pen notes, cashier signatures with dates, or voucher stamp blocks. Standardize to DD-MM-YYYY. "
        "If truly absent, leave empty string \"\".\n"
        "- vendor_name: Full seller / supplier / service provider entity name.\n"
        "- gstin: The 15-character GSTIN of the SELLER/SUPPLIER (not buyer/bank). Must follow strict Indian GSTIN format: "
        "2 digits (state), 5 uppercase letters (PAN), 4 digits, 1 uppercase letter, 1 alphanumeric, 'Z', 1 checksum alphanumeric. "
        "Disambiguate: letter 'O' vs digit '0', letter 'I' vs digit '1', letter 'S' vs digit '5', letter 'B' vs digit '8'. "
        "If not found, leave empty string \"\".\n"
        "- taxable_value: Base taxable amount before GST (numeric).\n"
        "- cgst: Central GST amount (numeric).\n"
        "- sgst: State/UT GST amount (numeric).\n"
        "- igst: Integrated GST amount (numeric).\n"
        "Ensure exact numeric math: Total GST = CGST + SGST + IGST."
    )
    # Render PDF page at 250 DPI for ultra-crisp resolution of small print and handwriting
    base64_jpg = render_pdf_page_to_png_base64(file_bytes, page_index=page_index, dpi=250)
    result, model_used = call_rescan_vision_model(system_prompt, user_prompt, base64_jpg, "image/jpeg")
    if "```json" in result:
        result = result.split("```json")[1].split("```")[0].strip()
    elif "```" in result:
        result = result.split("```")[1].split("```")[0].strip()
    parsed = json.loads(result)
    parsed["_ai_model"] = model_used
    return parsed

def extract_from_image_rescan(file_bytes, ext):
    """Runs high-accuracy vision model on invoice image files (PNG/JPG/WEBP)."""
    opt_bytes, mime_type = optimize_image_bytes(file_bytes, ext)
    base64_img = base64.b64encode(opt_bytes).decode('utf-8')
    system_prompt = (
        "You are a senior forensic financial OCR specialist examining Indian tax invoices, "
        "utility bills, bank debit vouchers, vendor payment receipts, and purchase registers. "
        "Your task is to thoroughly analyze the document image with forensic attention to detail, "
        "especially for MISSING, BLANK, HANDWRITTEN, STAMPED, or FAINT data fields.\n"
        "You must respond with ONLY a valid JSON object. Do not include any explanations outside JSON."
    )
    user_prompt = (
        "Perform a meticulous high-accuracy extraction of invoice details, reading BOTH PRINTED and HANDWRITTEN entries:\n"
        "- invoice_number: The bill, invoice, cash memo, or challan number.\n"
        "- invoice_date: Date of invoice issue (standardize to DD-MM-YYYY format).\n"
        "- payment_date: The date the bill was actually PAID / passed. Search diligently for rubber stamps ("
        "'PAID', 'SANCTIONED', 'PASSED FOR PAYMENT', 'CHEQUE NO', 'RTGS/NEFT', 'DEBITED ON'), "
        "handwritten pen notes, cashier signatures with dates, or voucher stamp blocks. Standardize to DD-MM-YYYY. "
        "If truly absent, leave empty string \"\".\n"
        "- vendor_name: Full seller / supplier / service provider entity name.\n"
        "- gstin: The 15-character GSTIN of the SELLER/SUPPLIER (not buyer/bank). Must follow strict Indian GSTIN format: "
        "2 digits (state), 5 uppercase letters (PAN), 4 digits, 1 uppercase letter, 1 alphanumeric, 'Z', 1 checksum alphanumeric. "
        "Disambiguate: letter 'O' vs digit '0', letter 'I' vs digit '1', letter 'S' vs digit '5', letter 'B' vs digit '8'. "
        "If not found, leave empty string \"\".\n"
        "- taxable_value: Base taxable amount before GST (numeric).\n"
        "- cgst: Central GST amount (numeric).\n"
        "- sgst: State/UT GST amount (numeric).\n"
        "- igst: Integrated GST amount (numeric).\n"
        "Ensure exact numeric math: Total GST = CGST + SGST + IGST."
    )
    result, model_used = call_rescan_vision_model(system_prompt, user_prompt, base64_img, mime_type)
    if "```json" in result:
        result = result.split("```json")[1].split("```")[0].strip()
    elif "```" in result:
        result = result.split("```")[1].split("```")[0].strip()
    parsed = json.loads(result)
    parsed["_ai_model"] = model_used
    return parsed


_MASTER_REFERENCE_BLOCK = None
def build_master_reference_block():
    """Formats the known-branch and known-vendor master lists into one compact
    text block to inject into every extraction prompt, so the model can match
    OCR'd branch/vendor text against known-correct values directly while
    reading the bill (fixing abbreviations, letterhead OCR noise, etc.)
    instead of relying only on post-hoc fuzzy string matching after the fact.
    Cached module-wide since the master lists never change during a run."""
    global _MASTER_REFERENCE_BLOCK
    if _MASTER_REFERENCE_BLOCK is not None:
        return _MASTER_REFERENCE_BLOCK
    branch_lines = "\n".join(f"- {b['name']} ({b['state']})" for b in MASTER_BRANCHES)
    vendor_lines = "\n".join(f"- {v['name']} -> {v['gstin']}" for v in MASTER_VENDORS)
    _MASTER_REFERENCE_BLOCK = (
        "REFERENCE LISTS (use ONLY to correct OCR/reading errors when the bill's own "
        "printed/handwritten text CLOSELY matches one of these -- never force a match "
        "onto a clearly different, unlisted vendor or branch, and never invent a branch "
        "that isn't actually indicated on the bill):\n\n"
        "Known bank branches (name (state)):\n" + branch_lines + "\n\n"
        "Known vendor master list (name -> GSTIN) -- if the vendor name or GSTIN you read "
        "closely matches one of these, prefer the listed GSTIN over a noisy OCR read:\n"
        + vendor_lines
    )
    return _MASTER_REFERENCE_BLOCK


def _parse_bill_array(result, model_name):
    """Parses a model's JSON response as a LIST of bills -- a single PDF page
    or image can legitimately contain more than one distinct bill (e.g.
    several small vouchers scanned onto one page). Tolerates a bare single
    object if the model ignored the array instruction by wrapping it, and
    tags every bill with which model produced it."""
    if "```json" in result:
        result = result.split("```json")[1].split("```")[0].strip()
    elif "```" in result:
        result = result.split("```")[1].split("```")[0].strip()
    parsed = json.loads(result)
    if isinstance(parsed, dict):
        parsed = [parsed]
    for bill in parsed:
        bill["_ai_model"] = model_name
    return parsed


_BILL_FIELDS_PROMPT = """- Invoice Number (invoice_number)
- Invoice Date (invoice_date) - The printed date the invoice/bill was issued.
- Payment Date (payment_date) - The date the bill was actually PAID, which is usually a
  HANDWRITTEN note or RUBBER-STAMPED annotation added after the invoice was printed
  (commonly near text like "RTGS/P.O. No.", "NEFT", "Cheque No.", or a "Sanctioned" /
  "Please Pay" stamp/signature block). This is DIFFERENT from the printed Invoice Date above.
  Leave blank if no such stamped/handwritten payment date is visible anywhere on the document.
- Vendor Name (vendor_name)
- Vendor GSTIN (gstin) - The SELLER/SUPPLIER's own GST registration number (usually printed
  near the letterhead, or near the signature/footer). Do NOT use the buyer/recipient's GSTIN,
  which is often printed next to the "M/s" or "Bill To" / customer name-and-address block -
  that number belongs to the customer, not the vendor. If only a buyer GSTIN is visible and no
  distinct seller GSTIN appears anywhere on the invoice, leave this blank rather than guessing.
  IMPORTANT: a GSTIN always follows a fixed 15-character pattern: 2 DIGITS (state code),
  5 LETTERS, 4 DIGITS, 1 LETTER, 1 DIGIT, the LETTER 'Z', then 1 final alphanumeric checksum
  character. Use this pattern to resolve ambiguous characters - e.g. if position 3 looks like
  it could be "O" or "0", the pattern says it must be a LETTER, so it is "O". Pay close
  attention to these commonly-confused pairs in both the gstin and invoice_number: letter O vs
  digit 0, letter I/L vs digit 1, letter S vs digit 5, letter B vs digit 8, letter G vs digit 6,
  letter Z vs digit 2.
- Branch (branch) - Only fill this in if the bill itself clearly indicates which bank branch it
  belongs to (a stamp, letterhead, or handwritten note explicitly naming one) -- match it to the
  closest name in the known branches reference list below. Leave blank/null if the bill has no
  such explicit indication; do NOT guess a branch from the vendor's address or general context.
- Taxable Value (taxable_value) - The value before taxes
- CGST Amount (cgst)
- SGST Amount (sgst)
- IGST Amount (igst)"""

_BILL_JSON_SCHEMA = """{
      "invoice_number": "...",
      "invoice_date": "...",
      "payment_date": "...",
      "vendor_name": "...",
      "gstin": "...",
      "branch": "...",
      "taxable_value": 0.0,
      "cgst": 0.0,
      "sgst": 0.0,
      "igst": 0.0
    }"""


def extract_from_text(text):
    """Sends extracted page text to the AI to parse invoice details into a
    JSON array of bills (a page can hold more than one distinct bill)."""
    system_prompt = (
        "You are an expert financial OCR assistant. Analyze the provided invoice text "
        "and extract the key values. You must respond with ONLY a valid JSON array. "
        "Do not include any explanation or markdown formatting outside the JSON."
    )

    user_prompt = f"""
    This text may contain ONE OR MORE separate bills/invoices (for example, several
    distinct vendor bills combined into one purchase-register page). Identify EACH
    distinct bill separately -- do not merge different bills' totals into one entry --
    and return a JSON ARRAY with one object per bill. If there is genuinely only one
    bill, return an array containing exactly one object.

    For each bill, extract:
    {_BILL_FIELDS_PROMPT}

    {build_master_reference_block()}

    Invoice Text:
    ---
    {text}
    ---

    Provide the output as a JSON array, one object per bill, each using this format:
    [
      {_BILL_JSON_SCHEMA}
    ]
    """

    # Stays on OpenRouter/Gemini throughout -- the old fallback called Claude
    # directly via Anthropic's API using a model id that no longer exists
    # there (a hard 404 on every retry), so it never actually provided a
    # working fallback. Retrying the same fast primary model instead covers
    # the transient rate-limit/empty-response blips actually being seen.
    openrouter_payload = {
        "model": AI_VISION_MODEL_NAME,
        "max_tokens": 3000,
        "reasoning": {"effort": "none"},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]
    }
    try:
        result = call_openrouter_api(openrouter_payload)
        return _parse_bill_array(result, AI_VISION_MODEL_NAME)
    except Exception as e:
        print(f"OpenRouter text extraction failed: {e}")

    result = call_openrouter_api(openrouter_payload)
    return _parse_bill_array(result, AI_VISION_MODEL_NAME)

_VISION_SYSTEM_PROMPT = (
    "You are an expert financial OCR assistant specialized in reading Indian tax invoices, "
    "utility bills, handwritten vouchers, bank payment stamps, and purchase registers. "
    "Analyze the invoice image and extract the key values with high precision. "
    "You must respond with ONLY a valid JSON array. Do not include any explanation outside JSON."
)


def _vision_bill_prompt():
    return (
        "This image may show ONE OR MORE separate bills/invoices (for example, several small "
        "vouchers or receipts scanned onto a single page). Identify EACH distinct bill "
        "separately -- do not merge different bills' totals into one entry -- and return a "
        "JSON ARRAY with one object per bill. If there is genuinely only one bill, return an "
        "array containing exactly one object.\n\n"
        "For each bill, extract details (reading BOTH PRINTED and HANDWRITTEN text across all fields):\n"
        "- invoice_number (printed or handwritten invoice/bill/challan number)\n"
        "- invoice_date (printed or handwritten invoice issue date)\n"
        "- payment_date (the date the bill was actually PAID - usually a HANDWRITTEN note, "
        "RUBBER-STAMPED annotation, cheque date, RTGS/NEFT date, or 'Sanctioned'/'Please Pay' stamp; "
        "leave blank if no payment date/stamp is found)\n"
        "- vendor_name (seller/supplier company or person name)\n"
        "- gstin (the SELLER/SUPPLIER's 15-character GSTIN, usually near letterhead or footer; "
        "do NOT extract buyer/bank GSTIN; leave blank if no vendor GSTIN is present)\n"
        "- branch (only if the bill clearly indicates which bank branch it belongs to -- a stamp, "
        "letterhead, or handwritten note explicitly naming one; match it to the closest name in the "
        "known branches reference list below; leave blank/null if there's no explicit indication -- "
        "do NOT guess a branch from the vendor's address or general context)\n"
        "- taxable_value (pre-tax base amount, printed or handwritten)\n"
        "- cgst, sgst, igst (tax amounts, printed or handwritten)\n"
        "IMPORTANT for accuracy:\n"
        "1. Read handwritten entries in blank form fields, rubber stamps, and pen notes accurately.\n"
        "2. A GSTIN always follows 15 chars: 2 digits (state), 5 letters (PAN), 4 digits, 1 letter, "
        "1 alphanumeric entity code, the letter 'Z', and 1 checksum char. Use this pattern to resolve "
        "ambiguous handwritten/printed characters (e.g. position 3 is always a letter 'O' not '0').\n\n"
        + build_master_reference_block() +
        "\n\nProvide the output as a JSON array, one object per bill, each using this format:\n"
        "[\n  " + _BILL_JSON_SCHEMA + "\n]"
    )


def extract_from_image(base64_data, mime_type):
    """Sends a base64 encoded invoice image to the vision model (via
    OpenRouter, see AI_VISION_MODEL_NAME) to parse details into a JSON
    array of bills (one image can hold more than one distinct bill)."""
    result, model_used = call_vision_model(_VISION_SYSTEM_PROMPT, _vision_bill_prompt(), base64_data, mime_type)
    return _parse_bill_array(result, model_used)


def extract_from_pdf_binary(file_bytes, page_index=0):
    """Renders the specified PDF page to an optimized JPEG and sends that
    image to the vision model (via OpenRouter, see AI_VISION_MODEL_NAME),
    rather than the raw PDF -- gives optimal control over resolution and
    fast transfer. Returns a JSON array of bills found on that page."""
    base64_jpg = render_pdf_page_to_png_base64(file_bytes, page_index=page_index)
    result, model_used = call_vision_model(_VISION_SYSTEM_PROMPT, _vision_bill_prompt(), base64_jpg, "image/jpeg")
    return _parse_bill_array(result, model_used)

def parse_excel_register(file_bytes):
    """Parses a purchase register or GSTR-2B Excel file using Pandas."""
    df = pd.read_excel(io.BytesIO(file_bytes))
    
    # Clean column names
    orig_cols = list(df.columns)
    clean_cols = [re.sub(r'[^a-z0-9]', '', str(c).strip().lower()) for c in df.columns]
    
    col_mapping = {}
    for orig, clean in zip(orig_cols, clean_cols):
        col_mapping[clean] = orig
        
    def find_column(options):
        for opt in options:
            if opt in col_mapping:
                return col_mapping[opt]
        return None

    inv_num_cols = ["invoicenumber", "invoiceno", "invno", "invoicenum", "billno", "documentnumber", "docno"]
    inv_date_cols = ["invoicedate", "invdate", "billdate", "documentdate", "docdate", "date", "inovicedate"]
    vendor_cols = ["vendorname", "suppliername", "partyname", "vendor", "supplier", "party", "legalname"]
    taxable_cols = ["taxablevalue", "taxableamt", "taxableamount", "assessablevalue", "taxval", "value"]
    cgst_cols = ["cgst", "cgstamount", "cgstamt", "centraltax"]
    sgst_cols = ["sgst", "sgstamount", "sgstamt", "statetax", "utgst", "unionterritorytax"]
    igst_cols = ["igst", "igstamount", "igstamt", "integratedtax"]
    gstin_cols = ["gstin", "gstno", "gstnumber", "vendorgstin", "suppliergstin", "gstregistrationnumber"]
    branch_cols = ["branch", "branchname", "location", "office", "unit"]
    state_cols = ["state", "section", "gstregistration", "registrationstate"]
    payment_date_cols = ["paymentdate", "paiddate", "datepaid", "paymentdt"]
    itc_blocked_cols = ["itcblocked", "gstblocked", "blocked", "noitc", "itcineligible"]

    col_num = find_column(inv_num_cols)
    col_date = find_column(inv_date_cols)
    col_vendor = find_column(vendor_cols)
    col_taxable = find_column(taxable_cols)
    col_cgst = find_column(cgst_cols)
    col_sgst = find_column(sgst_cols)
    col_igst = find_column(igst_cols)
    col_gstin = find_column(gstin_cols)
    col_branch = find_column(branch_cols)
    col_state = find_column(state_cols)
    col_payment_date = find_column(payment_date_cols)
    col_itc_blocked = find_column(itc_blocked_cols)

    # Positional fallback only makes sense for legacy 3-column registers (vendor, invoice no,
    # date with no headers). It's intentionally NOT applied to invoice number, since manual-bill
    # templates put GST No / Branch in those early columns and would otherwise get misread as
    # invoice numbers.
    if not col_vendor and len(orig_cols) > 0: col_vendor = orig_cols[0]
    if not col_date and len(orig_cols) > 2: col_date = orig_cols[2]

    invoices = []
    for _, row in df.iterrows():
        try:
            vendor = str(row[col_vendor]) if col_vendor and pd.notna(row[col_vendor]) else "Unknown Vendor"
            inv_no = str(row[col_num]) if col_num and pd.notna(row[col_num]) else "N/A"
            inv_date = str(row[col_date]).split(" ")[0] if col_date and pd.notna(row[col_date]) else "N/A"
            payment_date = str(row[col_payment_date]).split(" ")[0].strip() if col_payment_date and pd.notna(row[col_payment_date]) else None
            gstin = str(row[col_gstin]) if col_gstin and pd.notna(row[col_gstin]) else "N/A"
            branch = str(row[col_branch]).strip() if col_branch and pd.notna(row[col_branch]) else None
            state = str(row[col_state]).strip() if col_state and pd.notna(row[col_state]) else None
            itc_blocked = str(row[col_itc_blocked]).strip().lower() in ('yes', 'true', '1', 'y') if col_itc_blocked and pd.notna(row[col_itc_blocked]) else False

            taxable = float(row[col_taxable]) if col_taxable and pd.notna(row[col_taxable]) else 0.0
            cgst = float(row[col_cgst]) if col_cgst and pd.notna(row[col_cgst]) else 0.0
            sgst = float(row[col_sgst]) if col_sgst and pd.notna(row[col_sgst]) else 0.0
            igst = float(row[col_igst]) if col_igst and pd.notna(row[col_igst]) else 0.0

            invoices.append({
                "invoice_number": inv_no,
                "invoice_date": inv_date,
                "payment_date": payment_date,
                "vendor_name": vendor,
                "gstin": gstin,
                "branch": branch,
                "state": state,
                "itc_blocked": itc_blocked,
                "taxable_value": taxable,
                "cgst": cgst,
                "sgst": sgst,
                "igst": igst
            })
        except Exception as ex:
            print(f"Error parsing row: {ex}")
            continue
            
    return invoices

# Web Router Pages
@app.route('/')
@login_required
def home():
    return render_template('index.html', is_admin=is_admin_user(), api_key_configured=bool(ANTHROPIC_API_KEY or OPENROUTER_API_KEY), ai_model_name=AI_MODEL_DISPLAY_NAME)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('home'))
        
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        try:
            conn = get_db_connection()
            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
            cur.execute('SELECT * FROM users WHERE username = %s', (username,))
            user = cur.fetchone()
            cur.close()
            conn.close()
            
            if user and check_password_hash(user['password_hash'], password):
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['is_admin'] = bool(user['is_admin'])
                return redirect(url_for('home'))
            else:
                error = "Invalid username or password"
        except Exception as e:
            error = f"Database connection error: {e}"
            
    return render_template('login.html', error=error)

@app.route('/register', methods=['GET', 'POST'])
def register():
    is_admin = is_admin_user()
    
    # If user is logged in as non-admin, redirect to home
    if 'user_id' in session and not is_admin:
        return redirect(url_for('home'))
        
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        make_admin = request.form.get('is_admin') in ['on', 'true', '1']

        if not username or not password:
            error = "All fields are required"
        else:
            try:
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute('SELECT id FROM users WHERE username = %s', (username,))
                if cur.fetchone():
                    error = f"Username '{username}' already exists"
                else:
                    hashed_pw = generate_password_hash(password)
                    cur.execute('INSERT INTO users (username, password_hash, is_admin) VALUES (%s, %s, %s)', (username, hashed_pw, make_admin))
                    conn.commit()
                    cur.close()
                    conn.close()

                    if 'user_id' in session and is_admin:
                        log_activity(session['user_id'], 'CREATE_USER', f"Admin created user '{username}' (Admin: {make_admin})")
                        flash(f"User '{username}' registered successfully!", "success")
                        return redirect(url_for('admin_users'))
                    else:
                        flash("Registration successful! You can now log in.", "success")
                        return redirect(url_for('login'))
                cur.close()
                conn.close()
            except Exception as e:
                error = f"Database error: {e}"

    return render_template('register.html', error=error, is_admin=is_admin)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    error = None
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not current_password or not new_password or not confirm_password:
            error = "All fields are required"
        elif new_password != confirm_password:
            error = "New password and confirmation do not match"
        else:
            try:
                conn = get_db_connection()
                cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                cur.execute('SELECT * FROM users WHERE id = %s', (session['user_id'],))
                user = cur.fetchone()

                if not user or not check_password_hash(user['password_hash'], current_password):
                    error = "Current password is incorrect"
                else:
                    hashed_pw = generate_password_hash(new_password)
                    cur.execute('UPDATE users SET password_hash = %s WHERE id = %s', (hashed_pw, session['user_id']))
                    conn.commit()

                cur.close()
                conn.close()

                if not error:
                    flash("Password updated successfully.", "success")
                    return redirect(url_for('settings'))
            except Exception as e:
                error = f"Database connection error: {e}"

    return render_template('settings.html', error=error, is_admin=is_admin_user(), api_key_configured=bool(ANTHROPIC_API_KEY or OPENROUTER_API_KEY), ai_model_name=AI_MODEL_DISPLAY_NAME)

# Admin User Management Routes
@app.route('/admin/users', methods=['GET'])
@admin_required
def admin_users():
    error = None
    users = []
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute('''
            SELECT 
                u.id, 
                u.username, 
                u.is_admin, 
                u.created_at,
                (SELECT COUNT(*) FROM invoices i WHERE i.user_id = u.id) as invoice_count,
                (SELECT COUNT(*) FROM gstr2b_entries g WHERE g.user_id = u.id) as gstr2b_count
            FROM users u
            ORDER BY u.id ASC;
        ''')
        users = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        error = f"Database connection error: {e}"

    return render_template('users.html', users=users, error=error, is_admin=True, api_key_configured=bool(ANTHROPIC_API_KEY or OPENROUTER_API_KEY), ai_model_name=AI_MODEL_DISPLAY_NAME)

@app.route('/admin/users/create', methods=['POST'])
@admin_required
def admin_create_user():
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    is_admin_flag = request.form.get('is_admin') in ['on', 'true', '1']

    if not username or not password:
        flash("Username and password are required.", "error")
        return redirect(url_for('admin_users'))

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT id FROM users WHERE username = %s', (username,))
        if cur.fetchone():
            flash(f"Username '{username}' already exists.", "error")
        else:
            hashed_pw = generate_password_hash(password)
            cur.execute('INSERT INTO users (username, password_hash, is_admin) VALUES (%s, %s, %s)', (username, hashed_pw, is_admin_flag))
            conn.commit()
            log_activity(session['user_id'], 'CREATE_USER', f"Created user '{username}' (Admin: {is_admin_flag})")
            flash(f"User '{username}' registered successfully!", "success")
        cur.close()
        conn.close()
    except Exception as e:
        flash(f"Error registering user: {e}", "error")

    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/toggle-admin', methods=['POST'])
@admin_required
def admin_toggle_role(user_id):
    if user_id == session['user_id']:
        flash("You cannot revoke your own administrator status.", "error")
        return redirect(url_for('admin_users'))

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute('SELECT username, is_admin FROM users WHERE id = %s', (user_id,))
        user = cur.fetchone()
        if user:
            new_role = not bool(user['is_admin'])
            cur.execute('UPDATE users SET is_admin = %s WHERE id = %s', (new_role, user_id))
            conn.commit()
            role_text = "Admin" if new_role else "Standard User"
            log_activity(session['user_id'], 'CHANGE_USER_ROLE', f"Changed user '{user['username']}' role to {role_text}")
            flash(f"Updated '{user['username']}' role to {role_text}.", "success")
        else:
            flash("User not found.", "error")
        cur.close()
        conn.close()
    except Exception as e:
        flash(f"Error updating user role: {e}", "error")

    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def admin_reset_password(user_id):
    new_password = request.form.get('new_password', '').strip()
    if not new_password:
        flash("Password cannot be empty.", "error")
        return redirect(url_for('admin_users'))

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute('SELECT username FROM users WHERE id = %s', (user_id,))
        user = cur.fetchone()
        if user:
            hashed_pw = generate_password_hash(new_password)
            cur.execute('UPDATE users SET password_hash = %s WHERE id = %s', (hashed_pw, user_id))
            conn.commit()
            log_activity(session['user_id'], 'RESET_PASSWORD', f"Reset password for user '{user['username']}'")
            flash(f"Password reset successfully for '{user['username']}'.", "success")
        else:
            flash("User not found.", "error")
        cur.close()
        conn.close()
    except Exception as e:
        flash(f"Error resetting password: {e}", "error")

    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def admin_delete_user(user_id):
    if user_id == session['user_id']:
        flash("You cannot delete your own logged-in account.", "error")
        return redirect(url_for('admin_users'))

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute('SELECT username FROM users WHERE id = %s', (user_id,))
        user = cur.fetchone()
        if user:
            username = user['username']
            cur.execute('DELETE FROM users WHERE id = %s', (user_id,))
            conn.commit()
            log_activity(session['user_id'], 'DELETE_USER', f"Deleted user '{username}' (ID: {user_id})")
            flash(f"User '{username}' deleted successfully.", "success")
        else:
            flash("User not found.", "error")
        cur.close()
        conn.close()
    except Exception as e:
        flash(f"Error deleting user: {e}", "error")

    return redirect(url_for('admin_users'))

# API Endpoints

@app.route('/api/clients', methods=['GET'])
@login_required
def get_clients():
    return jsonify({
        "clients": CLIENTS_CONFIG,
        "active_client_id": get_current_client_id()
    })

@app.route('/api/get-active-client', methods=['GET'])
@login_required
def get_active_client_api():
    cid = get_current_client_id()
    return jsonify({
        "active_client_id": cid,
        "client": CLIENTS_CONFIG.get(cid, CLIENTS_CONFIG['nutan_nagrik'])
    })

@app.route('/api/set-active-client', methods=['POST'])
@login_required
def set_active_client_api():
    data = request.get_json(silent=True) or {}
    cid = data.get('client_id')
    if cid in CLIENTS_CONFIG:
        session['active_client_id'] = cid
        return jsonify({"success": True, "active_client_id": cid, "client": CLIENTS_CONFIG[cid]})
    return jsonify({"success": False, "error": "Invalid client ID"}), 400

@app.route('/api/master-data', methods=['GET'])
@login_required
def get_master_data():
    return jsonify({
        "branches": MASTER_BRANCHES,
        "vendors": MASTER_VENDORS
    })

@app.route('/api/get-invoices', methods=['GET'])
@login_required
def get_invoices():
    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if is_admin:
            cur.execute('''
                SELECT invoices.id, invoice_number, invoice_date, payment_date, vendor_name, gstin, branch, state,
                       taxable_value::float, cgst::float, sgst::float, igst::float, itc_blocked,
                       eligible_itc::float, ineligible_itc::float, users.username,
                       financial_year, month, client_id,
                       (file_data IS NOT NULL) AS has_file
                FROM invoices
                JOIN users ON users.id = invoices.user_id
                WHERE invoices.client_id = %s
                ORDER BY invoices.created_at DESC
            ''', (client_id,))
        else:
            cur.execute('''
                SELECT id, invoice_number, invoice_date, payment_date, vendor_name, gstin, branch, state,
                       taxable_value::float, cgst::float, sgst::float, igst::float, itc_blocked,
                       eligible_itc::float, ineligible_itc::float,
                       financial_year, month, client_id,
                       (file_data IS NOT NULL) AS has_file
                FROM invoices
                WHERE user_id = %s AND client_id = %s
                ORDER BY created_at DESC
            ''', (user_id, client_id))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({"invoices": rows, "client_id": client_id, "client": get_client_config(client_id)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def find_duplicate_invoice(cur, user_id, gstin, inv_num, vendor_name, inv_date, taxable_value, financial_year=None, exclude_id=None, client_id='nutan_nagrik'):
    """Checks if a matching invoice already exists in the database for the user.
    Returns (duplicate_id, reason_str) or (None, None)."""
    clean_num = (inv_num or '').strip()
    clean_gstin = normalize_gstin(gstin or 'N/A')
    clean_vendor = (vendor_name or '').strip()

    # 1. Primary rule: Same user, valid GSTIN, and recognized Invoice Number
    if clean_gstin != 'N/A' and clean_num not in ('', 'N/A', '-'):
        query = '''
            SELECT id, invoice_number, vendor_name, invoice_date, financial_year
            FROM invoices
            WHERE user_id = %s AND client_id = %s
              AND UPPER(TRIM(gstin)) = UPPER(%s)
              AND UPPER(TRIM(invoice_number)) = UPPER(%s)
        '''
        params = [user_id, client_id, clean_gstin, clean_num]
        if financial_year:
            query += ' AND (financial_year IS NULL OR financial_year = %s)'
            params.append(financial_year)
        if exclude_id:
            query += ' AND id != %s'
            params.append(exclude_id)
        cur.execute(query, params)
        row = cur.fetchone()
        if row:
            return row[0], f"Invoice #{row[1]} already exists for vendor GSTIN {clean_gstin}"

    # 2. Secondary rule: Same user, recognized Vendor Name and Invoice Number
    if clean_vendor not in ('', 'Unknown Vendor') and clean_num not in ('', 'N/A', '-'):
        query = '''
            SELECT id, invoice_number, vendor_name, invoice_date, financial_year
            FROM invoices
            WHERE user_id = %s AND client_id = %s
              AND UPPER(TRIM(vendor_name)) = UPPER(%s)
              AND UPPER(TRIM(invoice_number)) = UPPER(%s)
        '''
        params = [user_id, client_id, clean_vendor, clean_num]
        if financial_year:
            query += ' AND (financial_year IS NULL OR financial_year = %s)'
            params.append(financial_year)
        if exclude_id:
            query += ' AND id != %s'
            params.append(exclude_id)
        cur.execute(query, params)
        row = cur.fetchone()
        if row:
            return row[0], f"Invoice #{row[1]} already exists for vendor '{row[2]}'"

    # 3. Rule for bills without invoice number: match identical vendor, date, and taxable amount
    if clean_num in ('', 'N/A', '-') and clean_vendor not in ('', 'Unknown Vendor') and inv_date not in ('', 'N/A', '-'):
        query = '''
            SELECT id, invoice_number, vendor_name, invoice_date, taxable_value
            FROM invoices
            WHERE user_id = %s AND client_id = %s
              AND UPPER(TRIM(vendor_name)) = UPPER(%s)
              AND invoice_date = %s
              AND taxable_value = %s
        '''
        params = [user_id, client_id, clean_vendor, inv_date, float(taxable_value or 0.0)]
        if exclude_id:
            query += ' AND id != %s'
            params.append(exclude_id)
        cur.execute(query, params)
        row = cur.fetchone()
        if row:
            return row[0], f"Bill from '{row[2]}' dated {inv_date} for amount ₹{taxable_value} already recorded"

    return None, None

@app.route('/api/save-invoice', methods=['POST'])
@login_required
def save_invoice():
    user_id = session['user_id']
    inv = request.json
    
    db_id = inv.get('id')
    inv_num = inv.get('invoice_number', '')
    inv_date = inv.get('invoice_date', '')
    payment_date = inv.get('payment_date') or None
    vendor = inv.get('vendor_name', '')
    gstin = normalize_gstin(inv.get('gstin', '') or 'N/A', vendor)
    branch = inv.get('branch', '') or 'Unassigned'
    state = inv.get('state', '') or 'Unassigned'
    if (state == 'Unassigned' or not state) and branch != 'Unassigned':
        state = get_branch_state(branch)
    taxable = float(inv.get('taxable_value', 0.0))
    cgst = float(inv.get('cgst', 0.0))
    sgst = float(inv.get('sgst', 0.0))
    igst = float(inv.get('igst', 0.0))
    itc_blocked = bool(inv.get('itc_blocked', False))

    client_id = inv.get('client_id') or get_current_client_id()
    cfg = get_client_config(client_id)
    total_gst = cgst + sgst + igst
    if itc_blocked:
        eligible = 0.0
        ineligible = round(total_gst, 2)
    elif cfg.get('itc_rule') == 'standard_100':
        eligible = round(total_gst, 2)
        ineligible = 0.0
    else:
        eligible = round(total_gst * 0.5, 2)
        ineligible = round(total_gst * 0.5, 2)

    fy, m = parse_date_to_fy_and_month(inv_date)

    is_admin = is_admin_user()

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        if db_id:
            # Check if updated values duplicate another existing invoice
            dup_id, dup_reason = find_duplicate_invoice(cur, user_id, gstin, inv_num, vendor, inv_date, taxable, fy, exclude_id=db_id)
            if dup_id:
                cur.close()
                conn.close()
                return jsonify({"error": f"Cannot update: {dup_reason} (Bill ID #{dup_id})."}), 409

            # Update existing invoice (admins may edit any user's invoice)
            if is_admin:
                cur.execute('''
                    UPDATE invoices
                    SET invoice_number = %s, invoice_date = %s, payment_date = %s, vendor_name = %s, gstin = %s, branch = %s, state = %s,
                        taxable_value = %s, cgst = %s, sgst = %s, igst = %s, itc_blocked = %s,
                        eligible_itc = %s, ineligible_itc = %s, financial_year = %s, month = %s
                    WHERE id = %s
                ''', (inv_num, inv_date, payment_date, vendor, gstin, branch, state, taxable, cgst, sgst, igst, itc_blocked, eligible, ineligible, fy, m, db_id))
            else:
                cur.execute('''
                    UPDATE invoices
                    SET invoice_number = %s, invoice_date = %s, payment_date = %s, vendor_name = %s, gstin = %s, branch = %s, state = %s,
                        taxable_value = %s, cgst = %s, sgst = %s, igst = %s, itc_blocked = %s,
                        eligible_itc = %s, ineligible_itc = %s, financial_year = %s, month = %s
                    WHERE id = %s AND user_id = %s
                ''', (inv_num, inv_date, payment_date, vendor, gstin, branch, state, taxable, cgst, sgst, igst, itc_blocked, eligible, ineligible, fy, m, db_id, user_id))
            ret_id = db_id
        else:
            # Check if new invoice duplicates an existing record
            dup_id, dup_reason = find_duplicate_invoice(cur, user_id, gstin, inv_num, vendor, inv_date, taxable, fy)
            if dup_id:
                cur.close()
                conn.close()
                return jsonify({"error": f"Duplicate bill detected: {dup_reason} (Bill ID #{dup_id})."}), 409

            # Insert new invoice
            cur.execute('''
                INSERT INTO invoices (user_id, invoice_number, invoice_date, payment_date, vendor_name, gstin, branch, state, taxable_value, cgst, sgst, igst, itc_blocked, eligible_itc, ineligible_itc, financial_year, month)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
            ''', (user_id, inv_num, inv_date, payment_date, vendor, gstin, branch, state, taxable, cgst, sgst, igst, itc_blocked, eligible, ineligible, fy, m))
            ret_id = cur.fetchone()[0]

        conn.commit()
        cur.close()
        conn.close()

        if not db_id:
            desc = f'Added bill from {vendor or "Unknown Vendor"}' + (f' (Invoice #{inv_num})' if inv_num else '')
            log_activity(user_id, 'bill_added', desc, fy, m, 1)

        return jsonify({
            "success": True,
            "id": ret_id,
            "username": session.get('username', ''),
            "eligible_itc": eligible,
            "ineligible_itc": ineligible,
            "financial_year": fy,
            "month": m
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete-invoice', methods=['POST'])
@login_required
def delete_invoice():
    user_id = session['user_id']
    is_admin = is_admin_user()
    data = request.json
    db_id = data.get('id')

    if not db_id:
        return jsonify({"error": "Invoice ID required"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        if is_admin:
            cur.execute('DELETE FROM invoices WHERE id = %s', (db_id,))
        else:
            cur.execute('DELETE FROM invoices WHERE id = %s AND user_id = %s AND client_id = %s', (db_id, user_id, get_current_client_id()))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete-selected-invoices', methods=['POST'])
@login_required
def delete_selected_invoices():
    user_id = session['user_id']
    is_admin = is_admin_user()
    data = request.json or {}
    ids = data.get('ids', [])

    if not ids or not isinstance(ids, list):
        return jsonify({"error": "No invoices selected to delete"}), 400

    clean_ids = []
    for i in ids:
        try:
            clean_ids.append(int(i))
        except (ValueError, TypeError):
            continue

    if not clean_ids:
        return jsonify({"error": "Invalid invoice IDs provided"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        if is_admin:
            cur.execute('DELETE FROM invoices WHERE id = ANY(%s)', (clean_ids,))
        else:
            cur.execute('DELETE FROM invoices WHERE id = ANY(%s) AND user_id = %s', (clean_ids, user_id))
        deleted_count = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()

        log_activity(user_id, 'bills_deleted_selected', f'Deleted {deleted_count} selected bill(s)', record_count=deleted_count)
        return jsonify({"success": True, "count": deleted_count})
    except Exception as e:
        print(f"Error deleting selected invoices: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/apply-book-correction', methods=['POST'])
@login_required
def apply_book_correction():
    """Used from the Stage 3 'Possible Match' review flow: once a human
    confirms a near-match is really the same invoice (identical amounts,
    one OCR-misread character), overwrite the book entry's GSTIN/invoice
    number with the portal's government-verified values so the next
    reconciliation run matches it exactly."""
    user_id = session['user_id']
    is_admin = is_admin_user()
    data = request.json or {}
    book_id = data.get('book_id')
    invoice_number = (data.get('invoice_number') or '').strip()
    gstin = (data.get('gstin') or '').strip()

    if not book_id or not invoice_number or not gstin:
        return jsonify({"error": "book_id, invoice_number and gstin are required"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        if is_admin:
            cur.execute('UPDATE invoices SET invoice_number = %s, gstin = %s WHERE id = %s', (invoice_number, gstin, book_id))
        else:
            cur.execute('UPDATE invoices SET invoice_number = %s, gstin = %s WHERE id = %s AND user_id = %s', (invoice_number, gstin, book_id, user_id))
        updated = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()

        if updated == 0:
            return jsonify({"error": "Invoice not found or access denied"}), 404

        log_activity(user_id, 'bill_corrected', f'Corrected invoice #{invoice_number} to match GSTR-2B (Stage 3 review)', record_count=1)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/rescan-invoice', methods=['POST'])
@login_required
def rescan_invoice():
    """Used from the Stage 3 'Possible Match' review flow as an alternative
    to blindly trusting the portal's values: re-runs AI vision extraction
    on the bill's originally stored file (always the vision model directly,
    since the whole point is a more careful re-read, not the cheap
    text-extraction path). Manual/on-demand rather than automatic -- an
    auto-trigger on every reconciliation view would fire a paid API call
    on every page load for every unresolved near-match, which isn't
    something to do silently."""
    user_id = session['user_id']
    is_admin = is_admin_user()
    data = request.json or {}
    book_id = data.get('book_id')
    if not book_id:
        return jsonify({"error": "book_id is required"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        if is_admin:
            cur.execute('SELECT file_data, file_mime_type, itc_blocked FROM invoices WHERE id = %s', (book_id,))
        else:
            cur.execute('SELECT file_data, file_mime_type, itc_blocked FROM invoices WHERE id = %s AND user_id = %s', (book_id, user_id))
        row = cur.fetchone()

        if not row or row[0] is None:
            cur.close()
            conn.close()
            return jsonify({"error": "No original bill file stored for this invoice"}), 404

        file_data, mime_type, itc_blocked = row
        file_bytes = bytes(file_data)

        if mime_type == "application/pdf":
            inv = extract_from_pdf_binary(file_bytes)
        else:
            base64_img = base64.b64encode(file_bytes).decode('utf-8')
            inv = extract_from_image(base64_img, mime_type)

        invoice_number = inv.get("invoice_number") or "N/A"
        invoice_date = inv.get("invoice_date") or "N/A"
        payment_date = inv.get("payment_date") or None
        vendor_name = inv.get("vendor_name") or "Unknown Vendor"
        gstin = normalize_gstin(inv.get("gstin") or "N/A")
        taxable_value = float(inv.get("taxable_value") or 0.0)
        cgst = float(inv.get("cgst") or 0.0)
        sgst = float(inv.get("sgst") or 0.0)
        igst = float(inv.get("igst") or 0.0)

        total_gst = cgst + sgst + igst
        if itc_blocked:
            eligible = 0.0
            ineligible = round(total_gst, 2)
        else:
            eligible = round(total_gst * 0.5, 2)
            ineligible = round(total_gst * 0.5, 2)

        fy, m = parse_date_to_fy_and_month(invoice_date)

        cur.execute('''
            UPDATE invoices
            SET invoice_number = %s, invoice_date = %s, payment_date = %s, vendor_name = %s, gstin = %s,
                taxable_value = %s, cgst = %s, sgst = %s, igst = %s,
                eligible_itc = %s, ineligible_itc = %s, financial_year = %s, month = %s
            WHERE id = %s
        ''', (invoice_number, invoice_date, payment_date, vendor_name, gstin, taxable_value, cgst, sgst, igst,
              eligible, ineligible, fy, m, book_id))
        conn.commit()
        cur.close()
        conn.close()

        log_activity(user_id, 'bill_rescanned', f'Re-scanned bill with AI vision (Stage 3 review): {vendor_name} #{invoice_number}', fy, m, 1)

        return jsonify({
            "success": True,
            "invoice_number": invoice_number,
            "invoice_date": invoice_date,
            "payment_date": payment_date,
            "vendor_name": vendor_name,
            "gstin": gstin,
            "taxable_value": taxable_value,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "eligible_itc": eligible,
            "ineligible_itc": ineligible,
            "financial_year": fy,
            "month": m,
            "ai_model_used": inv.get("_ai_model")
        })
    except Exception as e:
        print(f"Error rescanning invoice: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/rescan-invoices-batch', methods=['POST'])
@login_required
def rescan_invoices_batch():
    """Batch re-scans invoices using higher-accuracy AI vision model (Gemini 2.5 Pro / Claude Sonnet),
    rendered at 250 DPI for forensic clarity on faint stamps, handwritten notes, and GSTINs.
    Returns comparison preview results for popup approval without immediately modifying the database."""
    user_id = session['user_id']
    is_admin = is_admin_user()
    data = request.json or {}
    req_ids = data.get('invoice_ids', [])
    filter_mode = data.get('filter_mode', 'incomplete') # 'incomplete', 'selected', 'all'
    
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        
        if req_ids and len(req_ids) > 0:
            clean_ids = [int(i) for i in req_ids if str(i).isdigit()]
            if is_admin:
                cur.execute('SELECT * FROM invoices WHERE id = ANY(%s) AND file_data IS NOT NULL ORDER BY id ASC', (clean_ids,))
            else:
                cur.execute('SELECT * FROM invoices WHERE id = ANY(%s) AND user_id = %s AND file_data IS NOT NULL ORDER BY id ASC', (clean_ids, user_id))
        elif filter_mode == 'incomplete':
            if is_admin:
                cur.execute('''
                    SELECT * FROM invoices 
                    WHERE file_data IS NOT NULL 
                      AND (
                        gstin IS NULL OR gstin = '' OR gstin = 'N/A' OR length(trim(gstin)) != 15
                        OR invoice_number IS NULL OR invoice_number = '' OR invoice_number = 'N/A' OR invoice_number ILIKE '%%error%%'
                        OR invoice_date IS NULL OR invoice_date = '' OR invoice_date = 'N/A'
                        OR payment_date IS NULL OR payment_date = ''
                        OR branch IS NULL OR branch = '' OR branch = 'Unassigned'
                        OR state IS NULL OR state = '' OR state = 'Unassigned'
                        OR ((taxable_value IS NULL OR taxable_value <= 0) AND (cgst IS NULL OR cgst = 0) AND (sgst IS NULL OR sgst = 0) AND (igst IS NULL OR igst = 0))
                      )
                    ORDER BY id ASC
                ''')
            else:
                cur.execute('''
                    SELECT * FROM invoices 
                    WHERE user_id = %s AND file_data IS NOT NULL 
                      AND (
                        gstin IS NULL OR gstin = '' OR gstin = 'N/A' OR length(trim(gstin)) != 15
                        OR invoice_number IS NULL OR invoice_number = '' OR invoice_number = 'N/A' OR invoice_number ILIKE '%%error%%'
                        OR invoice_date IS NULL OR invoice_date = '' OR invoice_date = 'N/A'
                        OR payment_date IS NULL OR payment_date = ''
                        OR branch IS NULL OR branch = '' OR branch = 'Unassigned'
                        OR state IS NULL OR state = '' OR state = 'Unassigned'
                        OR ((taxable_value IS NULL OR taxable_value <= 0) AND (cgst IS NULL OR cgst = 0) AND (sgst IS NULL OR sgst = 0) AND (igst IS NULL OR igst = 0))
                      )
                    ORDER BY id ASC
                ''', (user_id,))
        else:
            if is_admin:
                cur.execute('SELECT * FROM invoices WHERE file_data IS NOT NULL ORDER BY id ASC')
            else:
                cur.execute('SELECT * FROM invoices WHERE user_id = %s AND file_data IS NOT NULL ORDER BY id ASC', (user_id,))
                
        rows = cur.fetchall()
        cur.close()
        conn.close()
        
        if not rows:
            return jsonify({
                "success": True, 
                "results": [], 
                "message": "No invoices with attached original files found to re-scan.",
                "summary": {"total_scanned": 0, "total_with_changes": 0, "recovered_gstin": 0, "recovered_payment_date": 0, "recovered_invoice_num": 0}
            })
            
        def _process_row(row_dict):
            row_id = row_dict['id']
            file_data = bytes(row_dict['file_data'])
            mime_type = row_dict.get('file_mime_type') or 'application/pdf'
            file_name = row_dict.get('file_name') or f"bill_{row_id}.pdf"
            
            try:
                if mime_type == "application/pdf" or file_name.lower().endswith('.pdf'):
                    scanned = extract_from_pdf_binary_rescan(file_data)
                else:
                    ext = file_name.split('.')[-1].lower() if '.' in file_name else 'jpg'
                    scanned = extract_from_image_rescan(file_data, ext)
            except Exception as e:
                print(f"Error rescanning row #{row_id}: {e}")
                return None
                
            orig_gstin = (row_dict['gstin'] or '').strip()
            orig_inv_num = (row_dict['invoice_number'] or '').strip()
            orig_inv_date = (row_dict['invoice_date'] or '').strip()
            orig_pay_date = (row_dict['payment_date'] or '').strip()
            orig_vendor = (row_dict['vendor_name'] or '').strip()
            orig_taxable = float(row_dict['taxable_value'] or 0.0)
            orig_cgst = float(row_dict['cgst'] or 0.0)
            orig_sgst = float(row_dict['sgst'] or 0.0)
            orig_igst = float(row_dict['igst'] or 0.0)
            orig_blocked = bool(row_dict['itc_blocked'])
            orig_branch = (row_dict['branch'] or 'Unassigned').strip()
            orig_state = (row_dict['state'] or 'Unassigned').strip()
            
            new_vendor = (scanned.get('vendor_name') or orig_vendor).strip()
            new_gstin = normalize_gstin((scanned.get('gstin') or '').strip() or orig_gstin, new_vendor)
            new_inv_num = (scanned.get('invoice_number') or orig_inv_num).strip()
            new_inv_date = (scanned.get('invoice_date') or orig_inv_date).strip()
            new_pay_date = (scanned.get('payment_date') or '').strip() or None
            
            # Numeric values
            try:
                new_taxable = float(scanned.get('taxable_value', orig_taxable) or 0.0)
            except (ValueError, TypeError):
                new_taxable = orig_taxable
            try:
                new_cgst = float(scanned.get('cgst', orig_cgst) or 0.0)
            except (ValueError, TypeError):
                new_cgst = orig_cgst
            try:
                new_sgst = float(scanned.get('sgst', orig_sgst) or 0.0)
            except (ValueError, TypeError):
                new_sgst = orig_sgst
            try:
                new_igst = float(scanned.get('igst', orig_igst) or 0.0)
            except (ValueError, TypeError):
                new_igst = orig_igst
                
            # Compute changes
            changes = []
            
            # GSTIN check
            orig_gstin_invalid = not orig_gstin or orig_gstin == 'N/A' or len(orig_gstin) != 15
            new_gstin_valid = bool(new_gstin and new_gstin != 'N/A' and len(new_gstin) == 15)
            if orig_gstin != new_gstin:
                changes.append({
                    "field": "gstin",
                    "label": "GSTIN",
                    "old": orig_gstin or "Missing",
                    "new": new_gstin or "Missing",
                    "is_recovered": orig_gstin_invalid and new_gstin_valid
                })
                
            # Invoice # check
            orig_inv_invalid = not orig_inv_num or orig_inv_num in ('N/A', 'ERROR', 'Missing')
            new_inv_valid = bool(new_inv_num and new_inv_num not in ('N/A', 'ERROR', 'Missing'))
            if orig_inv_num != new_inv_num:
                changes.append({
                    "field": "invoice_number",
                    "label": "Invoice #",
                    "old": orig_inv_num or "Missing",
                    "new": new_inv_num or "Missing",
                    "is_recovered": orig_inv_invalid and new_inv_valid
                })
                
            # Invoice Date check
            orig_date_invalid = not orig_inv_date or orig_inv_date in ('N/A', 'ERROR')
            new_date_valid = bool(new_inv_date and new_inv_date not in ('N/A', 'ERROR'))
            if orig_inv_date != new_inv_date:
                changes.append({
                    "field": "invoice_date",
                    "label": "Invoice Date",
                    "old": orig_inv_date or "Missing",
                    "new": new_inv_date or "Missing",
                    "is_recovered": orig_date_invalid and new_date_valid
                })
                
            # Payment Date check
            orig_pay_invalid = not orig_pay_date or orig_pay_date in ('N/A', '')
            new_pay_valid = bool(new_pay_date and new_pay_date not in ('N/A', ''))
            if (orig_pay_date or '') != (new_pay_date or ''):
                changes.append({
                    "field": "payment_date",
                    "label": "Payment Date",
                    "old": orig_pay_date or "Missing",
                    "new": new_pay_date or "Missing",
                    "is_recovered": orig_pay_invalid and new_pay_valid
                })
                
            # Vendor Name
            if orig_vendor != new_vendor and new_vendor not in ('Unknown Vendor', 'N/A', ''):
                changes.append({
                    "field": "vendor_name",
                    "label": "Vendor Name",
                    "old": orig_vendor,
                    "new": new_vendor,
                    "is_recovered": orig_vendor in ('Unknown Vendor', 'N/A', '')
                })
                
            # Tax amounts
            if round(orig_taxable, 2) != round(new_taxable, 2) or round(orig_cgst, 2) != round(new_cgst, 2) or round(orig_sgst, 2) != round(new_sgst, 2) or round(orig_igst, 2) != round(new_igst, 2):
                changes.append({
                    "field": "tax_amounts",
                    "label": "Tax Amounts",
                    "old": f"Taxable: ₹{orig_taxable:.2f}, CGST: ₹{orig_cgst:.2f}, SGST: ₹{orig_sgst:.2f}, IGST: ₹{orig_igst:.2f}",
                    "new": f"Taxable: ₹{new_taxable:.2f}, CGST: ₹{new_cgst:.2f}, SGST: ₹{new_sgst:.2f}, IGST: ₹{new_igst:.2f}",
                    "is_recovered": (orig_taxable == 0 and new_taxable > 0) or (orig_cgst == 0 and orig_sgst == 0 and orig_igst == 0 and (new_cgst > 0 or new_sgst > 0 or new_igst > 0))
                })
                
            total_new_gst = new_cgst + new_sgst + new_igst
            if orig_blocked:
                new_eligible = 0.0
                new_ineligible = round(total_new_gst, 2)
            else:
                new_eligible = round(total_new_gst * 0.5, 2)
                new_ineligible = round(total_new_gst * 0.5, 2)
                
            return {
                "id": row_id,
                "file_name": file_name,
                "original": {
                    "vendor_name": orig_vendor,
                    "gstin": orig_gstin,
                    "invoice_number": orig_inv_num,
                    "invoice_date": orig_inv_date,
                    "payment_date": orig_pay_date,
                    "taxable_value": orig_taxable,
                    "cgst": orig_cgst,
                    "sgst": orig_sgst,
                    "igst": orig_igst,
                    "itc_blocked": orig_blocked,
                    "branch": orig_branch,
                    "state": orig_state
                },
                "scanned": {
                    "vendor_name": new_vendor,
                    "gstin": new_gstin,
                    "invoice_number": new_inv_num,
                    "invoice_date": new_inv_date,
                    "payment_date": new_pay_date,
                    "taxable_value": new_taxable,
                    "cgst": new_cgst,
                    "sgst": new_sgst,
                    "igst": new_igst,
                    "eligible_itc": new_eligible,
                    "ineligible_itc": new_ineligible,
                    "itc_blocked": orig_blocked,
                    "branch": orig_branch,
                    "state": orig_state,
                    "ai_model": scanned.get("_ai_model") or AI_RESCAN_MODEL_DISPLAY_NAME
                },
                "changes": changes,
                "has_changes": len(changes) > 0,
                "recovered_count": sum(1 for c in changes if c.get("is_recovered"))
            }

        # Run with ThreadPoolExecutor
        results = []
        with ThreadPoolExecutor(max_workers=min(len(rows), 4)) as executor:
            future_to_row = {executor.submit(_process_row, dict(r)): r for r in rows}
            for future in as_completed(future_to_row):
                res = future.result()
                if res:
                    results.append(res)
                    
        # Sort results by ID
        results.sort(key=lambda x: x['id'])
        
        # Summary counts
        rec_gstin = sum(1 for r in results if any(c['field'] == 'gstin' and c.get('is_recovered') for c in r['changes']))
        rec_pay = sum(1 for r in results if any(c['field'] == 'payment_date' and c.get('is_recovered') for c in r['changes']))
        rec_inv = sum(1 for r in results if any(c['field'] == 'invoice_number' and c.get('is_recovered') for c in r['changes']))
        with_changes = sum(1 for r in results if r['has_changes'])
        
        return jsonify({
            "success": True,
            "results": results,
            "summary": {
                "total_scanned": len(results),
                "total_with_changes": with_changes,
                "recovered_gstin": rec_gstin,
                "recovered_payment_date": rec_pay,
                "recovered_invoice_num": rec_inv,
                "ai_model_used": AI_RESCAN_MODEL_DISPLAY_NAME
            }
        })
    except Exception as e:
        print(f"Error in rescan_invoices_batch: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/apply-rescan-results', methods=['POST'])
@login_required
def apply_rescan_results():
    """Applies user-approved re-scan results to the PostgreSQL database in a single transaction."""
    user_id = session['user_id']
    is_admin = is_admin_user()
    data = request.json or {}
    updates = data.get('updates', [])
    
    if not updates or not isinstance(updates, list):
        return jsonify({"error": "No updates provided"}), 400
        
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        applied_count = 0
        
        for item in updates:
            inv_id = item.get('id')
            if not inv_id:
                continue
                
            vendor = (item.get('vendor_name') or '').strip()
            gstin = normalize_gstin((item.get('gstin') or '').strip(), vendor)
            inv_num = (item.get('invoice_number') or '').strip()
            inv_date = (item.get('invoice_date') or '').strip()
            payment_date = item.get('payment_date') or None
            if payment_date == '':
                payment_date = None
                
            taxable = float(item.get('taxable_value', 0.0) or 0.0)
            cgst = float(item.get('cgst', 0.0) or 0.0)
            sgst = float(item.get('sgst', 0.0) or 0.0)
            igst = float(item.get('igst', 0.0) or 0.0)
            itc_blocked = bool(item.get('itc_blocked', False))
            branch = (item.get('branch') or 'Unassigned').strip()
            state = (item.get('state') or 'Unassigned').strip()
            
            total_gst = cgst + sgst + igst
            if itc_blocked:
                eligible = 0.0
                ineligible = round(total_gst, 2)
            else:
                eligible = round(total_gst * 0.5, 2)
                ineligible = round(total_gst * 0.5, 2)
                
            fy, m = parse_date_to_fy_and_month(inv_date)
            
            if is_admin:
                cur.execute('''
                    UPDATE invoices
                    SET vendor_name = %s, gstin = %s, invoice_number = %s, invoice_date = %s,
                        payment_date = %s, taxable_value = %s, cgst = %s, sgst = %s, igst = %s,
                        itc_blocked = %s, eligible_itc = %s, ineligible_itc = %s,
                        financial_year = %s, month = %s, branch = %s, state = %s
                    WHERE id = %s
                ''', (vendor, gstin, inv_num, inv_date, payment_date, taxable, cgst, sgst, igst,
                      itc_blocked, eligible, ineligible, fy, m, branch, state, inv_id))
            else:
                cur.execute('''
                    UPDATE invoices
                    SET vendor_name = %s, gstin = %s, invoice_number = %s, invoice_date = %s,
                        payment_date = %s, taxable_value = %s, cgst = %s, sgst = %s, igst = %s,
                        itc_blocked = %s, eligible_itc = %s, ineligible_itc = %s,
                        financial_year = %s, month = %s, branch = %s, state = %s
                    WHERE id = %s AND user_id = %s
                ''', (vendor, gstin, inv_num, inv_date, payment_date, taxable, cgst, sgst, igst,
                      itc_blocked, eligible, ineligible, fy, m, branch, state, inv_id, user_id))
                      
            applied_count += cur.rowcount
            
        conn.commit()
        cur.close()
        conn.close()
        
        log_activity(user_id, 'bill_rescanned', f'Applied high-accuracy AI re-scan updates for {applied_count} bill(s)', record_count=applied_count)
        return jsonify({"success": True, "applied_count": applied_count})
    except Exception as e:
        print(f"Error applying rescan results: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/invoice-file/<int:invoice_id>', methods=['GET'])
@login_required
def get_invoice_file(invoice_id):
    user_id = session['user_id']
    is_admin = is_admin_user()
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        if is_admin:
            cur.execute('SELECT file_data, file_mime_type, file_name FROM invoices WHERE id = %s', (invoice_id,))
        else:
            cur.execute('SELECT file_data, file_mime_type, file_name FROM invoices WHERE id = %s AND user_id = %s', (invoice_id, user_id))
        row = cur.fetchone()
        cur.close()
        conn.close()

        if not row or row[0] is None:
            return jsonify({"error": "No file attached to this invoice"}), 404

        file_data, mime_type, file_name = row
        return send_file(
            io.BytesIO(bytes(file_data)),
            mimetype=mime_type or "application/octet-stream",
            as_attachment=False,
            download_name=file_name or f"invoice-{invoice_id}"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/clear-invoices', methods=['POST'])
@login_required
def clear_invoices():
    """Deletes invoice IDs sent by the client. Restricted to Admin users
    and requires entering account password to authorize bulk deletion."""
    user_id = session['user_id']
    if not is_admin_user():
        return jsonify({"error": "Clear All is restricted to Administrator users only."}), 403

    data = request.json or {}
    ids = data.get('ids')
    password = data.get('password', '').strip()

    if not ids or not isinstance(ids, list):
        return jsonify({"error": "No invoices selected to delete."}), 400

    if not password:
        return jsonify({"error": "Password confirmation is required to authorize bulk deletion."}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        # Verify admin password
        cur.execute('SELECT password_hash FROM users WHERE id = %s', (user_id,))
        user_row = cur.fetchone()

        if not user_row or not check_password_hash(user_row['password_hash'], password):
            cur.close()
            conn.close()
            return jsonify({"error": "Incorrect password. Bulk delete aborted."}), 401

        # Delete selected invoices
        cur.execute('DELETE FROM invoices WHERE id = ANY(%s)', (ids,))
        deleted = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()

        log_activity(user_id, 'bills_cleared', f'Cleared {deleted} bill(s) via Password-Protected Clear All', record_count=deleted)
        return jsonify({"success": True, "count": deleted})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def _parse_single_invoice_file(filename, file_bytes, batch_branch, batch_state, high_accuracy):
    """Processes a single uploaded bill file (PDF, Image, Excel/CSV) and returns
    its parsed invoice records, file storage buffers, and error state. Thread-safe."""
    ext = filename.split('.')[-1].lower()
    store_file_bytes = None
    store_mime_type = None
    store_file_name = None
    parsed_list = []

    # 1. Excel/CSV Processing
    if ext in ['xlsx', 'xls', 'csv']:
        if ext == 'csv':
            df = pd.read_csv(io.BytesIO(file_bytes))
            out = io.BytesIO()
            df.to_excel(out, index=False)
            file_bytes = out.getvalue()
        parsed_list = parse_excel_register(file_bytes)

    # 2. PDF Processing
    elif ext == 'pdf':
        try:
            doc = pymupdf.open(stream=file_bytes, filetype="pdf")
            num_pages = len(doc)
        except Exception as pe:
            print(f"Error opening PDF {filename}: {pe}")
            num_pages = 0
            doc = None

        if not doc or num_pages == 0:
            parsed_list = []
        elif num_pages == 1:
            page_text = doc[0].get_text()
            if not high_accuracy and len(page_text.strip()) > 100:
                bills = extract_from_text(page_text)

                # Gap-fill vision only applies cleanly when the page held exactly
                # one bill -- with multiple bills detected from text, there's no
                # single "this page's GSTIN/payment date" to fill in, so those
                # bills are accepted as extracted rather than second-guessed.
                if len(bills) == 1:
                    inv = bills[0]
                    needs_gstin = not inv.get('gstin') or inv.get('gstin') == 'N/A'
                    if needs_gstin:
                        gstin_match = re.search(r'\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}\b', page_text)
                        if gstin_match:
                            inv['gstin'] = gstin_match.group(0)
                            needs_gstin = False

                    needs_payment_date = not inv.get('payment_date') or inv.get('payment_date') == inv.get('invoice_date')
                    payment_date_markers = ('rtgs', 'neft', 'p.o. no', 'po no', 'cheque',
                                             'sanctioned', 'please pay', 'paid on', 'demand draft')
                    has_payment_voucher_section = any(m in page_text.lower() for m in payment_date_markers)

                    if needs_gstin or (needs_payment_date and has_payment_voucher_section):
                        try:
                            vision_bills = extract_from_pdf_binary(file_bytes, page_index=0)
                            vision_inv = vision_bills[0] if vision_bills else {}
                            if needs_payment_date and vision_inv.get('payment_date'):
                                inv['payment_date'] = vision_inv['payment_date']
                            if needs_gstin and vision_inv.get('gstin'):
                                inv['gstin'] = vision_inv['gstin']
                            inv['_ai_model'] = vision_inv.get('_ai_model', inv.get('_ai_model'))
                        except Exception as ex:
                            print(f"Vision fallback failed for {filename}: {ex}")
                    bills = [inv]
            else:
                # High Accuracy Scan or scanned PDF without text layer: full vision pass
                bills = extract_from_pdf_binary(file_bytes, page_index=0)

            parsed_list = bills
            store_file_bytes = file_bytes
            store_mime_type = "application/pdf"
            store_file_name = filename
            doc.close()
        else:
            # Multi-page PDF: scan all pages concurrently across thread pool.
            # Each page can itself hold more than one bill, so this returns a
            # LIST of bills per page rather than assuming a strict 1:1 mapping.
            def _process_pdf_page(p_idx):
                try:
                    p_doc = pymupdf.open(stream=file_bytes, filetype="pdf")
                    page_text = p_doc[p_idx].get_text()

                    # Extract this single page as a standalone PDF for individual bill storage
                    single_doc = pymupdf.open()
                    single_doc.insert_pdf(p_doc, from_page=p_idx, to_page=p_idx)
                    single_bytes = single_doc.tobytes()
                    single_doc.close()
                    p_doc.close()

                    if not high_accuracy and len(page_text.strip()) > 100:
                        bills = extract_from_text(page_text)
                        if len(bills) == 1:
                            inv = bills[0]
                            needs_gstin = not inv.get('gstin') or inv.get('gstin') == 'N/A'
                            if needs_gstin:
                                gstin_match = re.search(r'\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}\b', page_text)
                                if gstin_match:
                                    inv['gstin'] = gstin_match.group(0)
                                    needs_gstin = False

                            needs_payment_date = not inv.get('payment_date') or inv.get('payment_date') == inv.get('invoice_date')
                            payment_date_markers = ('rtgs', 'neft', 'p.o. no', 'po no', 'cheque',
                                                     'sanctioned', 'please pay', 'paid on', 'demand draft')
                            has_payment_voucher_section = any(m in page_text.lower() for m in payment_date_markers)

                            if needs_gstin or (needs_payment_date and has_payment_voucher_section):
                                try:
                                    vision_bills = extract_from_pdf_binary(file_bytes, page_index=p_idx)
                                    vision_inv = vision_bills[0] if vision_bills else {}
                                    if needs_payment_date and vision_inv.get('payment_date'):
                                        inv['payment_date'] = vision_inv['payment_date']
                                    if needs_gstin and vision_inv.get('gstin'):
                                        inv['gstin'] = vision_inv['gstin']
                                    inv['_ai_model'] = vision_inv.get('_ai_model', inv.get('_ai_model'))
                                except Exception as ex:
                                    print(f"Vision fallback failed for {filename} page {p_idx+1}: {ex}")
                            bills = [inv]
                    else:
                        bills = extract_from_pdf_binary(file_bytes, page_index=p_idx)

                    multi = len(bills) > 1
                    for b_idx, b in enumerate(bills):
                        b['_store_file_bytes'] = single_bytes
                        b['_store_file_name'] = f"{filename} (Page {p_idx+1})" + (f" - Bill {b_idx+1} of {len(bills)}" if multi else "")
                        b['_store_mime_type'] = "application/pdf"
                        b['_page_number'] = p_idx + 1
                    return bills
                except Exception as ex:
                    print(f"Error processing page {p_idx+1} of {filename}: {ex}")
                    # Report the failure as a visible row instead of silently
                    # dropping this page's bill(s) from the batch.
                    return [{
                        "invoice_number": "ERROR", "invoice_date": "-", "payment_date": None,
                        "vendor_name": f"Failed to parse page {p_idx+1} of {filename}",
                        "gstin": "N/A", "branch": None,
                        "taxable_value": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0,
                        "_store_file_bytes": None, "_store_file_name": f"{filename} (Page {p_idx+1})",
                        "_store_mime_type": None, "_page_number": p_idx + 1, "_error": str(ex)
                    }]

            with ThreadPoolExecutor(max_workers=min(num_pages, 6)) as p_executor:
                p_futures = [p_executor.submit(_process_pdf_page, i) for i in range(num_pages)]
                for pf in as_completed(p_futures):
                    p_res = pf.result()
                    if p_res:
                        parsed_list.extend(p_res)

            parsed_list.sort(key=lambda x: x.get('_page_number', 0))
            store_file_bytes = file_bytes
            store_mime_type = "application/pdf"
            store_file_name = filename
            doc.close()

    # 3. Image Processing
    elif ext in ['png', 'jpg', 'jpeg', 'webp']:
        opt_bytes, mime_type = optimize_image_bytes(file_bytes, ext)
        base64_img = base64.b64encode(opt_bytes).decode('utf-8')
        parsed_list = extract_from_image(base64_img, mime_type)
        store_file_bytes = file_bytes
        store_mime_type = mime_type
        store_file_name = filename
    # 4. ZIP Archive Processing
    elif ext in ['zip']:
        import zipfile
        try:
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                for zname in z.namelist():
                    zname_lower = zname.lower()
                    if zname_lower.endswith('/') or '__macosx' in zname_lower or 'thumbs.db' in zname_lower:
                        continue
                    zext = zname_lower.split('.')[-1]
                    if zext in ['pdf', 'png', 'jpg', 'jpeg', 'webp', 'xlsx', 'xls', 'csv']:
                        zbytes = z.read(zname)
                        sub_res = _parse_single_invoice_file(zname, zbytes, batch_branch, batch_state, high_accuracy)
                        if sub_res.get('parsed_list'):
                            parsed_list.extend(sub_res['parsed_list'])
        except Exception as ze:
            print(f"Error extracting zip {filename}: {ze}")
            parsed_list = []

    else:
        raise ValueError(f"Unsupported file format: {ext}")

    return {
        "filename": filename,
        "parsed_list": parsed_list,
        "store_file_bytes": store_file_bytes,
        "store_mime_type": store_mime_type,
        "store_file_name": store_file_name,
        "error": None
    }

@app.route('/api/process-invoices', methods=['POST'])
@login_required
def process_invoices():
    user_id = session['user_id']
    client_id = get_current_client_id()
    if 'files[]' not in request.files:
        return jsonify({"error": "No files uploaded"}), 400

    files = request.files.getlist('files[]')
    batch_branch = request.form.get('branch', '').strip()
    batch_state = request.form.get('state', '').strip()
    high_accuracy = request.form.get('high_accuracy', '').lower() in ('1', 'true', 'yes')

    if high_accuracy:
        confirm_password = request.form.get('confirm_password', '')
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute('SELECT password_hash FROM users WHERE id = %s', (user_id,))
            row = cur.fetchone()
            cur.close()
            conn.close()
        except Exception as e:
            return jsonify({"error": f"Database connection error: {e}"}), 500

        if not row or not confirm_password or not check_password_hash(row[0], confirm_password):
            return jsonify({"error": "Incorrect password"}), 403

    # Read uploaded file contents into memory for concurrent worker access
    file_payloads = [(f.filename, f.read()) for f in files]
    file_results = [None] * len(file_payloads)

    def _worker(idx, fname, fbytes):
        try:
            res = _parse_single_invoice_file(fname, fbytes, batch_branch, batch_state, high_accuracy)
            return idx, res
        except Exception as e:
            print(f"Error processing file {fname}: {e}")
            return idx, {
                "filename": fname,
                "parsed_list": [],
                "store_file_bytes": None,
                "store_mime_type": None,
                "store_file_name": None,
                "error": str(e)
            }

    # Execute concurrent multithreaded scanning across worker pool
    max_workers = min(8, max(1, len(file_payloads)))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(_worker, i, fname, fbytes) for i, (fname, fbytes) in enumerate(file_payloads)]
        for future in as_completed(futures):
            idx, res = future.result()
            file_results[idx] = res

    # Save parsed records to PostgreSQL safely with savepoint per record
    results = []
    seen_in_batch = set()
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        for parsed_res in file_results:
            filename = parsed_res["filename"]
            if parsed_res.get("error") or not parsed_res.get("parsed_list"):
                results.append({
                    "id": None,
                    "is_duplicate": False,
                    "invoice_number": "ERROR",
                    "invoice_date": "-",
                    "payment_date": None,
                    "vendor_name": f"Failed to parse {filename}",
                    "gstin": "N/A",
                    "branch": batch_branch or "Unassigned",
                    "state": batch_state or "Unassigned",
                    "taxable_value": 0.0,
                    "cgst": 0.0,
                    "sgst": 0.0,
                    "igst": 0.0,
                    "itc_blocked": False,
                    "has_file": False,
                    "eligible_itc": 0.0,
                    "ineligible_itc": 0.0,
                    "filename": filename,
                    "message": parsed_res.get("error") or "No readable invoice data"
                })
                continue

            for inv in parsed_res["parsed_list"]:
                inv["invoice_number"] = str(inv.get("invoice_number") or "N/A")[:100]
                inv["invoice_date"] = str(inv.get("invoice_date") or "N/A")[:50]
                inv["payment_date"] = str(inv["payment_date"])[:50] if inv.get("payment_date") else None
                inv["vendor_name"] = str(inv.get("vendor_name") or "Unknown Vendor")[:255]
                inv["gstin"] = normalize_gstin(str(inv.get("gstin") or "N/A")[:50], inv["vendor_name"])
                inv["branch"] = str(inv.get("branch") or batch_branch or "Unassigned")[:100]
                inv["state"] = str(inv.get("state") or batch_state or "Unassigned")[:100]

                # Auto-detect branch and state from folder path or filename when uploading
                if (inv["branch"] == 'Unassigned' or not inv["branch"]) and not batch_branch:
                    fn_norm = filename.replace("\\", "/").upper()
                    for mb in MASTER_BRANCHES:
                        b_name = mb['name'].upper()
                        if (f"{b_name}/" in fn_norm or f"/{b_name}" in fn_norm or fn_norm.startswith(f"{b_name}/")
                            or f"_{b_name}_" in fn_norm or f" {b_name} " in fn_norm or f"_{b_name} " in fn_norm
                            or f" {b_name}_" in fn_norm or b_name in fn_norm):
                            inv["branch"] = mb['name']
                            inv["state"] = mb['state']
                            break

                if (inv["state"] == 'Unassigned' or not inv["state"]) and inv["branch"] != 'Unassigned':
                    inv["state"] = get_branch_state(inv["branch"])
                inv["itc_blocked"] = bool(inv.get("itc_blocked", False))
                for field in ("taxable_value", "cgst", "sgst", "igst"):
                    try:
                        inv[field] = float(inv.get(field) or 0.0)
                    except (TypeError, ValueError):
                        inv[field] = 0.0

                total_gst = inv["cgst"] + inv["sgst"] + inv["igst"]
                if inv["itc_blocked"]:
                    eligible = 0.0
                    ineligible = round(total_gst, 2)
                else:
                    eligible = round(total_gst * 0.5, 2)
                    ineligible = round(total_gst * 0.5, 2)

                fy, m = parse_date_to_fy_and_month(inv["invoice_date"])

                # Check duplicate against existing records in PostgreSQL
                dup_id, dup_reason = find_duplicate_invoice(
                    cur, user_id, inv["gstin"], inv["invoice_number"],
                    inv["vendor_name"], inv["invoice_date"], inv["taxable_value"], fy
                )

                # Check duplicate within the same uploaded batch
                batch_key = (
                    (inv["gstin"].upper() if inv["gstin"] != "N/A" else inv["vendor_name"].upper()),
                    (inv["invoice_number"].upper() if inv["invoice_number"] != "N/A" else f"{inv['invoice_date']}_{inv['taxable_value']}")
                )
                if not dup_id:
                    if batch_key in seen_in_batch:
                        dup_id = "BATCH_DUP"
                        dup_reason = f"Duplicate bill within the same uploaded batch for {inv['vendor_name']} #{inv['invoice_number']}"
                    else:
                        seen_in_batch.add(batch_key)

                inv_store_bytes = inv.get('_store_file_bytes') or parsed_res["store_file_bytes"]
                inv_store_mime = inv.get('_store_mime_type') or parsed_res["store_mime_type"]
                inv_store_name = inv.get('_store_file_name') or parsed_res["store_file_name"]

                if dup_id:
                    # Duplicate detected: skip database insertion and flag as duplicate
                    results.append({
                        "id": None,
                        "is_duplicate": True,
                        "duplicate_of_id": dup_id if isinstance(dup_id, int) else None,
                        "invoice_number": inv["invoice_number"],
                        "invoice_date": inv["invoice_date"],
                        "payment_date": inv["payment_date"],
                        "vendor_name": inv["vendor_name"],
                        "gstin": inv["gstin"],
                        "branch": inv["branch"],
                        "state": inv["state"],
                        "taxable_value": inv["taxable_value"],
                        "cgst": inv["cgst"],
                        "sgst": inv["sgst"],
                        "igst": inv["igst"],
                        "itc_blocked": inv["itc_blocked"],
                        "has_file": False,
                        "eligible_itc": eligible,
                        "ineligible_itc": ineligible,
                        "financial_year": fy,
                        "month": m,
                        "filename": inv_store_name or filename,
                        "ai_model_used": inv.get("_ai_model"),
                        "message": dup_reason
                    })
                    continue

                try:
                    cur.execute("SAVEPOINT sp_inv")
                    cur.execute('''
                        INSERT INTO invoices (user_id, client_id, invoice_number, invoice_date, payment_date, vendor_name, gstin, branch, state, taxable_value, cgst, sgst, igst, itc_blocked, eligible_itc, ineligible_itc, file_data, file_mime_type, file_name, financial_year, month)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
                    ''', (user_id, client_id, inv["invoice_number"], inv["invoice_date"], inv["payment_date"], inv["vendor_name"], inv["gstin"], inv["branch"], inv["state"],
                          inv["taxable_value"], inv["cgst"], inv["sgst"], inv["igst"], inv["itc_blocked"],
                          eligible, ineligible,
                          psycopg2.Binary(inv_store_bytes) if inv_store_bytes else None,
                          inv_store_mime, inv_store_name, fy, m))

                    db_id = cur.fetchone()[0]
                    cur.execute("RELEASE SAVEPOINT sp_inv")
                    results.append({
                        "id": db_id,
                        "is_duplicate": False,
                        "invoice_number": inv["invoice_number"],
                        "invoice_date": inv["invoice_date"],
                        "payment_date": inv["payment_date"],
                        "vendor_name": inv["vendor_name"],
                        "gstin": inv["gstin"],
                        "branch": inv["branch"],
                        "state": inv["state"],
                        "taxable_value": inv["taxable_value"],
                        "cgst": inv["cgst"],
                        "sgst": inv["sgst"],
                        "igst": inv["igst"],
                        "itc_blocked": inv["itc_blocked"],
                        "has_file": inv_store_bytes is not None,
                        "eligible_itc": eligible,
                        "ineligible_itc": ineligible,
                        "financial_year": fy,
                        "month": m,
                        "filename": inv_store_name or filename,
                        "username": session.get('username', ''),
                        "ai_model_used": inv.get("_ai_model")
                    })
                except Exception as ins_err:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_inv")
                    print(f"Insert error for {filename}: {ins_err}")
                    results.append({
                        "id": None,
                        "is_duplicate": False,
                        "invoice_number": "ERROR",
                        "invoice_date": inv.get("invoice_date", "-"),
                        "payment_date": None,
                        "vendor_name": inv.get("vendor_name", f"Failed {filename}"),
                        "gstin": inv.get("gstin", "N/A"),
                        "branch": inv.get("branch", "Unassigned"),
                        "state": inv.get("state", "Unassigned"),
                        "taxable_value": 0.0,
                        "cgst": 0.0,
                        "sgst": 0.0,
                        "igst": 0.0,
                        "itc_blocked": False,
                        "has_file": False,
                        "eligible_itc": 0.0,
                        "ineligible_itc": 0.0,
                        "filename": filename,
                        "message": f"Database save error: {ins_err}"
                    })

        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"Database error during batch invoice insert: {e}")

    success_count = sum(1 for r in results if r.get("id") is not None)
    duplicate_count = sum(1 for r in results if r.get("is_duplicate"))
    failed_count = len(results) - success_count - duplicate_count

    if success_count > 0 or duplicate_count > 0:
        scan_mode = "High Accuracy Scan" if high_accuracy else "AI Scan"
        desc = f'Processed {len(results)} bill(s): {success_count} uploaded'
        if duplicate_count > 0:
            desc += f', {duplicate_count} duplicate(s) skipped'
        if batch_branch:
            desc += f' for branch {batch_branch}'
        desc += f' via {scan_mode}'
        log_activity(user_id, 'bill_upload', desc, record_count=success_count)

    return jsonify({
        "invoices": results,
        "success_count": success_count,
        "duplicate_count": duplicate_count,
        "failed_count": failed_count
    })

@app.route('/api/export-excel', methods=['POST'])
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.json
    invoices = data.get('invoices', [])

    # Normalize + compute the eligible/ineligible split per tax head (CGST/SGST/IGST),
    # matching the bank's required reconciliation format instead of a single 50% total.
    # Respects itc_blocked (Section 17(5)): 0% eligible / 100% ineligible for blocked
    # invoices, 50/50 otherwise -- rather than assuming every invoice is a flat 50%.
    rows = []
    for inv in invoices:
        cgst = float(inv.get('cgst') or 0.0)
        sgst = float(inv.get('sgst') or 0.0)
        igst = float(inv.get('igst') or 0.0)
        taxable = float(inv.get('taxable_value') or 0.0)
        elig_ratio = 0.0 if inv.get('itc_blocked') else 0.5
        elig_cgst = round(cgst * elig_ratio, 2)
        elig_sgst = round(sgst * elig_ratio, 2)
        elig_igst = round(igst * elig_ratio, 2)
        rows.append({
            "state": inv.get('state') or 'Unassigned',
            "branch": inv.get('branch') or 'Unassigned',
            "gstin": inv.get('gstin') or 'N/A',
            "invoice_date": inv.get('invoice_date') or '',
            "vendor_name": inv.get('vendor_name') or '',
            "invoice_number": inv.get('invoice_number') or '',
            "taxable_value": taxable,
            "cgst": cgst,
            "sgst": sgst,
            "igst": igst,
            "total_invoice_value": taxable + cgst + sgst + igst,
            "elig_cgst": elig_cgst,
            "elig_sgst": elig_sgst,
            "elig_igst": elig_igst,
            "inelig_cgst": round(cgst - elig_cgst, 2),
            "inelig_sgst": round(sgst - elig_sgst, 2),
            "inelig_igst": round(igst - elig_igst, 2),
        })

    # Group by (state, branch) -- not branch alone, since the same branch name
    # can exist under both state registrations and would otherwise be silently
    # merged into one subtotal. Blank/Unassigned sorted last within each level.
    state_branch_groups = sorted(
        {(r["state"], r["branch"]) for r in rows},
        key=lambda sb: (sb[0] == 'Unassigned', sb[0].lower(), sb[1] == 'Unassigned', sb[1].lower())
    )
    states_ordered = sorted(
        {sb[0] for sb in state_branch_groups},
        key=lambda s: (s == 'Unassigned', s.lower())
    )

    NUMERIC_FIELDS = ["taxable_value", "cgst", "sgst", "igst", "total_invoice_value",
                       "elig_cgst", "elig_sgst", "elig_igst", "inelig_cgst", "inelig_sgst", "inelig_igst"]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "GST ITC Reconciled"

    navy_header_fill = PatternFill(start_color="0A2540", end_color="0A2540", fill_type="solid")
    subtotal_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    total_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    white_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=11, bold=True)
    regular_font = Font(name="Arial", size=10)

    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
    )
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000')
    )

    # ---- Header (two rows, with merged ELIGIBLE / INELIGIBLE groups) ----
    single_headers = [
        (1, "State"), (2, "Branch"), (3, "GST No"), (4, "Date"), (5, "Vendor Name"), (6, "Invoice No"),
        (7, "Taxable Value (INR)"), (8, "CGST (INR)"), (9, "SGST (INR)"), (10, "IGST (INR)"),
        (11, "Total Invoice Value (INR)")
    ]
    for col_idx, label in single_headers:
        worksheet.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        worksheet.cell(row=1, column=col_idx, value=label)

    worksheet.merge_cells(start_row=1, start_column=12, end_row=1, end_column=14)
    worksheet.cell(row=1, column=12, value="ELIGIBLE ITC (50%)")
    worksheet.merge_cells(start_row=1, start_column=15, end_row=1, end_column=17)
    worksheet.cell(row=1, column=15, value="INELIGIBLE ITC (50%)")

    for col_idx, label in [(12, "CGST"), (13, "SGST"), (14, "IGST"), (15, "CGST"), (16, "SGST"), (17, "IGST")]:
        worksheet.cell(row=2, column=col_idx, value=label)

    total_cols = 17
    for row_idx in (1, 2):
        for col_idx in range(1, total_cols + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = navy_header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Data rows, grouped by state then branch, with a branch subtotal row
    # and a state total row -- (state, branch) grouping (not branch alone)
    # since the same branch name can exist under both state registrations.
    row_idx = 3
    grand_totals = {f: 0.0 for f in NUMERIC_FIELDS}

    for state in states_ordered:
        state_totals = {f: 0.0 for f in NUMERIC_FIELDS}
        branches_in_state = [sb[1] for sb in state_branch_groups if sb[0] == state]

        for branch in branches_in_state:
            branch_rows = [r for r in rows if r["state"] == state and r["branch"] == branch]
            branch_totals = {f: 0.0 for f in NUMERIC_FIELDS}

            for r in branch_rows:
                values = [
                    r["state"], r["branch"], r["gstin"], r["invoice_date"], r["vendor_name"], r["invoice_number"],
                    r["taxable_value"], r["cgst"], r["sgst"], r["igst"], r["total_invoice_value"],
                    r["elig_cgst"], r["elig_sgst"], r["elig_igst"], r["inelig_cgst"], r["inelig_sgst"], r["inelig_igst"]
                ]
                for col_idx, val in enumerate(values, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = regular_font
                    cell.border = thin_border
                    if col_idx >= 7:
                        cell.alignment = Alignment(horizontal="right")
                        cell.number_format = '#,##0.00'
                    else:
                        cell.alignment = Alignment(horizontal="left")
                for f in NUMERIC_FIELDS:
                    branch_totals[f] += r[f]
                    state_totals[f] += r[f]
                    grand_totals[f] += r[f]
                row_idx += 1

            # Branch subtotal row
            subtotal_values = ["", f"{branch} - Subtotal", "", "", "", "",
                                branch_totals["taxable_value"], branch_totals["cgst"], branch_totals["sgst"],
                                branch_totals["igst"], branch_totals["total_invoice_value"],
                                branch_totals["elig_cgst"], branch_totals["elig_sgst"], branch_totals["elig_igst"],
                                branch_totals["inelig_cgst"], branch_totals["inelig_sgst"], branch_totals["inelig_igst"]]
            for col_idx, val in enumerate(subtotal_values, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=val)
                cell.font = bold_font
                cell.fill = subtotal_fill
                cell.border = thin_border
                if col_idx >= 7:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
            row_idx += 1

        # State total row
        state_total_values = [f"{state} - TOTAL", "", "", "", "", "",
                               state_totals["taxable_value"], state_totals["cgst"], state_totals["sgst"],
                               state_totals["igst"], state_totals["total_invoice_value"],
                               state_totals["elig_cgst"], state_totals["elig_sgst"], state_totals["elig_igst"],
                               state_totals["inelig_cgst"], state_totals["inelig_sgst"], state_totals["inelig_igst"]]
        for col_idx, val in enumerate(state_total_values, start=1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=val)
            cell.font = bold_font
            cell.fill = total_fill
            cell.border = thin_border
            if col_idx >= 7:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
        row_idx += 1

    # ---- Grand total row ----
    grand_total_values = ["GRAND TOTAL", "", "", "", "", "",
                           grand_totals["taxable_value"], grand_totals["cgst"], grand_totals["sgst"],
                           grand_totals["igst"], grand_totals["total_invoice_value"],
                           grand_totals["elig_cgst"], grand_totals["elig_sgst"], grand_totals["elig_igst"],
                           grand_totals["inelig_cgst"], grand_totals["inelig_sgst"], grand_totals["inelig_igst"]]
    for col_idx, val in enumerate(grand_total_values, start=1):
        cell = worksheet.cell(row=row_idx, column=col_idx, value=val)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = double_bottom_border
        if col_idx >= 7:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = '#,##0.00'

    # ---- Autofit columns ----
    for col_idx in range(1, total_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is None:
                    continue
                val_to_check = str(cell.value)
                if cell.number_format == '#,##0.00' and isinstance(cell.value, (int, float)):
                    val_to_check = f"{cell.value:,.2f}"
                max_len = max(max_len, len(val_to_check))
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    worksheet.freeze_panes = "A3"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    log_activity(session['user_id'], 'export_excel', 'Exported reconciled ITC Excel sheet', record_count=len(rows))

    return send_file(
        output,
        as_attachment=True,
        download_name="GST_ITC_Reconciled_Sheet.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

import re
import collections

def clean_invoice_number(num):
    """Fuzzy invoice number cleaner for Indian GST matching. Strips leading zeros, spaces, and non-alphanumeric chars."""
    if not num or num == 'N/A' or num == '-' or num == 'None':
        return ''
    cleaned = re.sub(r'[^a-zA-Z0-9]', '', str(num)).lower()
    return cleaned.lstrip('0')

# GSTIN is a fixed 15-character format: 2 digits (state code), 5 letters +
# 4 digits + 1 letter (PAN), 1 digit (entity code), the constant letter
# 'Z', then 1 alphanumeric checksum (left untouched -- genuinely
# ambiguous). Unlike invoice numbers, this positional structure is rigid
# and public knowledge, so a character that violates the expected
# digit/letter class at its position can be deterministically corrected
# to the one visually-similar character that actually fits -- with near
# zero risk of "fixing" an already-correct GSTIN, since a valid one never
# trips these checks in the first place.
_GSTIN_EXPECT_DIGIT = {0, 1, 7, 8, 9, 10, 12}
_GSTIN_EXPECT_LETTER = {2, 3, 4, 5, 6, 11, 13}
_GSTIN_TO_DIGIT = {'O': '0', 'I': '1', 'L': '1', 'S': '5', 'B': '8', 'G': '6', 'Z': '2'}
_GSTIN_TO_LETTER = {'0': 'O', '1': 'I', '5': 'S', '8': 'B', '6': 'G', '2': 'Z'}
def normalize_gstin(raw, vendor_name=""):
    """Corrects OCR letter/digit confusions (O/0, I or L/1, S/5, B/8, G/6,
    Z/2) in a GSTIN by position, and checks the 91 master vendors if incomplete."""
    if not raw or str(raw).strip() in ('', 'N/A', '-', 'None', 'null'):
        if vendor_name:
            mv = match_master_vendor(vendor_name)
            if mv:
                return mv['gstin']
        return 'N/A'

    cleaned = re.sub(r'[^A-Za-z0-9]', '', str(raw)).strip().upper()
    if len(cleaned) == 15:
        chars = list(cleaned)
        for i in _GSTIN_EXPECT_DIGIT:
            c = chars[i]
            if not c.isdigit() and c in _GSTIN_TO_DIGIT:
                chars[i] = _GSTIN_TO_DIGIT[c]
        for i in _GSTIN_EXPECT_LETTER:
            c = chars[i]
            if not c.isalpha() and c in _GSTIN_TO_LETTER:
                chars[i] = _GSTIN_TO_LETTER[c]
        return ''.join(chars)

    if vendor_name or cleaned:
        mv = match_master_vendor(vendor_name, cleaned)
        if mv:
            return mv['gstin']

    return cleaned if cleaned else 'N/A'

def levenshtein(a, b):
    """Edit distance between two strings, used to spot a likely OCR misread
    (e.g. 'O'/'0', 'P'/'F') between an otherwise-identical book/portal pair
    that the exact-match reconciliation pass would otherwise miss entirely."""
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i] + [0] * len(b)
        for j, cb in enumerate(b, 1):
            cost = 0 if ca == cb else 1
            curr[j] = min(prev[j] + 1, curr[j - 1] + 1, prev[j - 1] + cost)
        prev = curr
    return prev[-1]

def parse_gstr2b_excel(file_bytes):
    """
    Parses GSTR-2B Excel file downloaded from GST Portal.
    Reads the 'B2B' sheet if present, else fallback to active sheet.
    Finds header row dynamically.
    """
    excel_file = pd.ExcelFile(io.BytesIO(file_bytes))
    
    # Target 'B2B' sheet (case-insensitive)
    target_sheet = None
    for sheet in excel_file.sheet_names:
        if sheet.strip().upper() == 'B2B':
            target_sheet = sheet
            break
            
    df = pd.read_excel(excel_file, sheet_name=target_sheet if target_sheet else 0)
    
    # Find column headers dynamically
    header_row_idx = 0
    found = False
    for idx, row in df.iterrows():
        row_vals = [str(v).strip().lower() for v in row.values if pd.notna(v)]
        has_gstin = any('gstin' in v or 'gst no' in v for v in row_vals)
        has_inv = any('invoice' in v or 'document' in v or 'bill' in v for v in row_vals)
        if has_gstin and has_inv:
            header_row_idx = idx
            found = True
            break
            
    if found:
        # The GST portal's real B2B export uses a two-row merged header: a
        # vague top-level category row ("Invoice Details", "Tax Amount",
        # merged across several columns) with the actual field names in the
        # row directly below it ("Invoice number", "Invoice Date",
        # "Central Tax", "State/UT Tax", ...). Reading only the top row (as
        # before) loses every specific column name, so nothing past GSTIN
        # ever matches. Detect a genuine second header row -- as opposed to
        # data already starting there -- by checking whether its first cell
        # looks like a GSTIN (15-char alphanumeric); if not, merge it in,
        # letting the specific sub-column label win over the category above it.
        top_row = df.iloc[header_row_idx].ffill()
        combined_cols = list(top_row)
        data_start_idx = header_row_idx + 1

        if header_row_idx + 1 < len(df):
            first_cell_below = str(df.iloc[header_row_idx + 1, 0]).strip()
            looks_like_gstin = len(first_cell_below) == 15 and first_cell_below.isalnum()
            if not looks_like_gstin:
                sub_row = df.iloc[header_row_idx + 1]
                combined_cols = [
                    sub if pd.notna(sub) and str(sub).strip() else top
                    for top, sub in zip(combined_cols, sub_row)
                ]
                data_start_idx = header_row_idx + 2

        df.columns = combined_cols
        df = df.iloc[data_start_idx:].reset_index(drop=True)

    orig_cols = list(df.columns)
    clean_cols = [re.sub(r'[^a-z0-9]', '', str(c).strip().lower()) for c in df.columns]
    
    col_mapping = {clean: orig for orig, clean in zip(orig_cols, clean_cols)}
    
    def find_column(options):
        for opt in options:
            if opt in col_mapping:
                return col_mapping[opt]
        return None

    inv_num_cols = ["invoicenumber", "invoiceno", "invno", "documentnumber", "docno", "billnumber", "billno"]
    inv_date_cols = ["invoicedate", "invdate", "documentdate", "docdate", "date", "billdate"]
    vendor_cols = ["tradelegalnameofthesupplier", "tradelegalname", "suppliername", "partyname", "vendorname", "vendor", "supplier"]
    taxable_cols = ["taxablevalue", "taxableamt", "taxableamount", "assessablevalue", "taxablevalueinr"]
    cgst_cols = ["centraltax", "cgst", "cgstamount", "cgstamt", "centraltaxinr"]
    sgst_cols = ["stateuttax", "sgst", "sgstamount", "sgstamt", "statetax", "stateuttaxinr"]
    igst_cols = ["integratedtax", "igst", "igstamount", "igstamt", "integratedtaxinr"]
    gstin_cols = ["gstinofsupplier", "gstin", "gstno", "gstnumber", "suppliergstin"]

    col_num = find_column(inv_num_cols)
    col_date = find_column(inv_date_cols)
    col_vendor = find_column(vendor_cols)
    col_taxable = find_column(taxable_cols)
    col_cgst = find_column(cgst_cols)
    col_sgst = find_column(sgst_cols)
    col_igst = find_column(igst_cols)
    col_gstin = find_column(gstin_cols)

    if not col_gstin or not col_num:
        print("Required columns (GSTIN or Invoice No) not found in GSTR-2B file.")
        return []

    def safe_float(val):
        """Handles comma thousands-separators (e.g. '1,234.56') from portal
        exports. Never raises -- a malformed numeric cell shouldn't discard
        an otherwise-valid row."""
        if val is None or pd.isna(val):
            return 0.0
        try:
            return float(str(val).replace(',', '').strip())
        except (TypeError, ValueError):
            return 0.0

    entries = []
    for _, row in df.iterrows():
        try:
            gstin = str(row[col_gstin]).strip() if pd.notna(row[col_gstin]) else None
            if not gstin or gstin.lower() in ['nan', 'null', 'n/a', '']:
                continue

            inv_no = str(row[col_num]).strip() if pd.notna(row[col_num]) else "N/A"
            inv_date = str(row[col_date]).split(" ")[0].strip() if col_date and pd.notna(row[col_date]) else "N/A"
            vendor = str(row[col_vendor]).strip() if col_vendor and pd.notna(row[col_vendor]) else "Unknown Vendor"

            taxable = safe_float(row[col_taxable]) if col_taxable else 0.0
            cgst = safe_float(row[col_cgst]) if col_cgst else 0.0
            sgst = safe_float(row[col_sgst]) if col_sgst else 0.0
            igst = safe_float(row[col_igst]) if col_igst else 0.0

            entries.append({
                "invoice_number": inv_no,
                "invoice_date": inv_date,
                "vendor_name": vendor,
                "gstin": gstin,
                "taxable_value": taxable,
                "cgst": cgst,
                "sgst": sgst,
                "igst": igst
            })
        except Exception as ex:
            print(f"Error parsing GSTR-2B row: {ex}")
            continue
    return entries

@app.route('/reconciliation')
@login_required
def reconciliation():
    current_fy, _ = _dt_to_fy_and_month(datetime.datetime.now())
    return render_template('reconciliation.html', is_admin=is_admin_user(), api_key_configured=bool(ANTHROPIC_API_KEY or OPENROUTER_API_KEY), ai_model_name=AI_MODEL_DISPLAY_NAME, current_fy=current_fy)

@app.route('/api/export-annual-report', methods=['GET'])
@login_required
def export_annual_report():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    user_id = session['user_id']
    is_admin = is_admin_user()
    fy = request.args.get('financial_year', '').strip()
    if not fy:
        return jsonify({"error": "Financial Year is required"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

        # 1. Fetch all books invoices for this FY (all users' if admin, since
        # GST is filed at the company level across every branch/user)
        if is_admin:
            cur.execute('''
                SELECT id, invoice_number, invoice_date, vendor_name, gstin, branch, state,
                       taxable_value::float, cgst::float, sgst::float, igst::float,
                       eligible_itc::float, ineligible_itc::float, itc_blocked, month
                FROM invoices
                WHERE financial_year = %s
                ORDER BY state, branch, created_at
            ''', (fy,))
        else:
            cur.execute('''
                SELECT id, invoice_number, invoice_date, vendor_name, gstin, branch, state,
                       taxable_value::float, cgst::float, sgst::float, igst::float,
                       eligible_itc::float, ineligible_itc::float, itc_blocked, month
                FROM invoices
                WHERE user_id = %s AND financial_year = %s
                ORDER BY state, branch, created_at
            ''', (user_id, fy))
        books = cur.fetchall()
        books = sorted(books, key=lambda r: (r['state'] or '', r['branch'] or '', month_sort_key(r['month'])))

        # Per-tax-head eligible/ineligible split. Blocked-ITC invoices (0%
        # eligible / 100% ineligible) and normal ones (50/50) both already
        # have the correct combined eligible_itc/ineligible_itc stored --
        # deriving each invoice's ratio from that (rather than assuming a
        # flat 50%) and applying it per tax head keeps CGST+SGST+IGST
        # exactly consistent with the totals already shown elsewhere.
        for r in books:
            total_gst = r["cgst"] + r["sgst"] + r["igst"]
            elig_ratio = (r["eligible_itc"] / total_gst) if total_gst > 0 else (0.0 if r["itc_blocked"] else 0.5)
            r["elig_cgst"] = round(r["cgst"] * elig_ratio, 2)
            r["elig_sgst"] = round(r["sgst"] * elig_ratio, 2)
            r["elig_igst"] = round(r["igst"] * elig_ratio, 2)
            r["inelig_cgst"] = round(r["cgst"] - r["elig_cgst"], 2)
            r["inelig_sgst"] = round(r["sgst"] - r["elig_sgst"], 2)
            r["inelig_igst"] = round(r["igst"] - r["elig_igst"], 2)

        # 2. Fetch all portal GSTR-2B entries for this FY
        if is_admin:
            cur.execute('''
                SELECT id, invoice_number, invoice_date, supplier_name as vendor_name, supplier_gstin as gstin, state,
                       taxable_value::float, cgst::float, sgst::float, igst::float, month
                FROM gstr2b_entries
                WHERE financial_year = %s
                ORDER BY created_at
            ''', (fy,))
        else:
            cur.execute('''
                SELECT id, invoice_number, invoice_date, supplier_name as vendor_name, supplier_gstin as gstin, state,
                       taxable_value::float, cgst::float, sgst::float, igst::float, month
                FROM gstr2b_entries
                WHERE user_id = %s AND financial_year = %s
                ORDER BY created_at
            ''', (user_id, fy))
        portal = cur.fetchall()
        portal = sorted(portal, key=lambda r: (r['state'] or '', month_sort_key(r['month'])))

        cur.close()
        conn.close()

        # Generate workbook with 2 sheets
        wb = Workbook()
        
        # Sheet 1: Purchase Register (Books)
        ws_books = wb.active
        ws_books.title = "Annual Purchase Book"
        
        # Style Definitions
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=10)
        border_thin = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                             top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
        
        # Header Row
        headers_books = ["State", "Month", "Branch", "Supplier GSTIN", "Vendor Name", "Invoice No", "Date",
                          "Taxable Value (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)",
                          "Eligible CGST (₹)", "Eligible SGST (₹)", "Eligible IGST (₹)", "Total Eligible ITC (₹)",
                          "Ineligible CGST (₹)", "Ineligible SGST (₹)", "Ineligible IGST (₹)", "Total Ineligible ITC (₹)"]
        ws_books.append(headers_books)
        for col_idx in range(1, len(headers_books) + 1):
            cell = ws_books.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")

        for r in books:
            row_vals = [
                r["state"], r["month"], r["branch"], r["gstin"], r["vendor_name"], r["invoice_number"], r["invoice_date"],
                r["taxable_value"], r["cgst"], r["sgst"], r["igst"],
                r["elig_cgst"], r["elig_sgst"], r["elig_igst"], r["eligible_itc"],
                r["inelig_cgst"], r["inelig_sgst"], r["inelig_igst"], r["ineligible_itc"]
            ]
            ws_books.append(row_vals)
            curr_row = ws_books.max_row
            for col_idx in range(1, len(headers_books) + 1):
                cell = ws_books.cell(row=curr_row, column=col_idx)
                cell.font = regular_font
                cell.border = border_thin
                if col_idx >= 8:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
                    
        # Autofit columns
        for col in ws_books.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if isinstance(cell.value, float):
                    val = f"{cell.value:,.2f}"
                max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            ws_books.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        # Sheet 2: Portal Entries
        ws_portal = wb.create_sheet("Annual Portal GSTR-2B")
        headers_portal = ["State", "Month", "Supplier GSTIN", "Vendor Name", "Invoice No", "Date", "Taxable Value (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)"]
        ws_portal.append(headers_portal)
        for col_idx in range(1, len(headers_portal) + 1):
            cell = ws_portal.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")

        for r in portal:
            row_vals = [
                r["state"], r["month"], r["gstin"], r["vendor_name"], r["invoice_number"], r["invoice_date"],
                r["taxable_value"], r["cgst"], r["sgst"], r["igst"]
            ]
            ws_portal.append(row_vals)
            curr_row = ws_portal.max_row
            for col_idx in range(1, len(headers_portal) + 1):
                cell = ws_portal.cell(row=curr_row, column=col_idx)
                cell.font = regular_font
                cell.border = border_thin
                if col_idx >= 7:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
                    
        for col in ws_portal.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if isinstance(cell.value, float):
                    val = f"{cell.value:,.2f}"
                max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            ws_portal.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        log_activity(user_id, 'export_annual_report', f'Exported annual GST report ({len(books)} book + {len(portal)} portal entries)', fy, record_count=len(books))

        return send_file(
            output,
            as_attachment=True,
            download_name=f"Annual_GST_Report_{fy}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error exporting annual report: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/export-filtered-reconciliation', methods=['GET'])
@login_required
def export_filtered_reconciliation():
    """Exports the exact currently filtered reconciliation ledger as an Excel spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    fy = request.args.get('financial_year', '2026-27').strip()
    months_param = request.args.get('months', '').strip()
    state_filter = request.args.get('state', 'all').strip()
    status_filter = request.args.get('status', 'all').strip()
    search_query = request.args.get('search', '').strip().lower()

    months = [m.strip() for m in months_param.split(',') if m.strip()]
    if not months:
        months = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January', 'February', 'March']

    try:
        summary, items, _, _ = execute_reconciliation(fy, months, user_id, is_admin, client_id=client_id)

        # 1. Apply State filter
        if state_filter and state_filter.lower() != 'all':
            items = [i for i in items if (i.get('state') or 'Unassigned').lower() == state_filter.lower()]

        # 2. Apply Status filter
        if status_filter and status_filter.lower() != 'all':
            items = [i for i in items if (i.get('status') or '').lower() == status_filter.lower()]

        # 3. Apply Search query
        if search_query:
            filtered = []
            for i in items:
                b = i.get('book') or {}
                p = i.get('portal') or {}
                vendor = (b.get('vendor_name') or p.get('vendor_name') or '').lower()
                gstin = (b.get('gstin') or p.get('gstin') or '').lower()
                b_inv = (b.get('invoice_number') or '').lower()
                p_inv = (p.get('invoice_number') or '').lower()
                branch = (b.get('branch') or '').lower()
                if (search_query in vendor or search_query in gstin or 
                    search_query in b_inv or search_query in p_inv or 
                    search_query in branch):
                    filtered.append(i)
            items = filtered

        # Generate Excel Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reconciliation Ledger"

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
        subtitle_font = Font(name="Calibri", size=10, italic=True, color="475569")
        white_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        border_thin = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                             top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))
        num_fmt = '#,##0.00'

        # Title Block
        ws.merge_cells("A1:L1")
        ws["A1"] = f"NUTAN NAGRIK SAHAKARI BANK LTD. - GST RECONCILIATION LEDGER (FY {fy})"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A2:L2")
        ws["A2"] = f"Active Filters: Months: {', '.join(months)} | State: {state_filter.title()} | Status: {status_filter.title()} | Total Records: {len(items)}"
        ws["A2"].font = subtitle_font
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "State", "Supplier / Vendor", "Supplier GSTIN",
            "Books Branch", "Books Invoice No", "Books Date", "Books GST (₹)",
            "Portal Invoice No", "Portal Date", "Portal GST (₹)", "Portal Taxable (₹)",
            "Match Status"
        ]

        ws.append([]) # Empty row 3
        ws.append(headers) # Row 4
        header_row_idx = 4

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_thin

        status_fills = {
            'Matched': PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
            'Value Mismatched': PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            'Possible Match': PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
            'Missing in GSTR-2B': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            'Missing in Books': PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        }

        row_idx = header_row_idx + 1
        for item in items:
            b = item.get('book') or {}
            p = item.get('portal') or {}
            state = item.get('state') or 'Unassigned'
            supplier = b.get('vendor_name') or p.get('vendor_name') or 'Unknown'
            gstin = b.get('gstin') or p.get('gstin') or 'N/A'
            b_branch = b.get('branch') or '-'
            b_inv = b.get('invoice_number') or '-'
            b_date = b.get('invoice_date') or '-'
            b_gst = b.get('total_gst') if b.get('total_gst') is not None else ''
            p_inv = p.get('invoice_number') or '-'
            p_date = p.get('invoice_date') or '-'
            p_gst = p.get('total_gst') if p.get('total_gst') is not None else ''
            p_taxable = p.get('taxable_value') if p.get('taxable_value') is not None else ''
            status = item.get('status') or 'Unknown'

            row_data = [
                state, supplier, gstin,
                b_branch, b_inv, b_date, b_gst,
                p_inv, p_date, p_gst, p_taxable,
                status
            ]
            ws.append(row_data)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = border_thin
                if col_idx in [7, 10, 11]:
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [1, 4, 5, 6, 8, 9]:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx == 12:
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = bold_font
                    fill = status_fills.get(status)
                    if fill:
                        cell.fill = fill

            row_idx += 1

        # Adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row > 2 and cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"GST_Reconciliation_{fy}_{state_filter}_{status_filter}.xlsx".replace(' ', '_')
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error exporting filtered reconciliation: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/upload-gstr2b', methods=['POST'])
@login_required
def upload_gstr2b():
    user_id = session['user_id']
    client_id = get_current_client_id()
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
        
    file = request.files['file']
    fy = request.form.get('financial_year', '').strip()
    month = request.form.get('month', '').strip()
    state = request.form.get('state', '').strip()

    if not fy or not month or not state:
        return jsonify({"error": "Financial Year, Month, and State are required"}), 400

    file_bytes = file.read()

    try:
        entries = parse_gstr2b_excel(file_bytes)
        if not entries:
            return jsonify({"error": "No valid GSTR-2B entries found in sheet B2B. Check file headers."}), 400

        conn = get_db_connection()
        cur = conn.cursor()

        # Delete existing portal entries for this client/FY/Month/State to prevent
        # duplicates on re-upload -- scoped by state too, so re-uploading
        # Maharashtra's file for a month never wipes Gujarat's entries for that
        # same month (and vice versa).
        cur.execute('''
            DELETE FROM gstr2b_entries
            WHERE client_id = %s AND financial_year = %s AND month = %s AND state = %s
        ''', (client_id, fy, month, state))

        inserted = 0
        for ent in entries:
            cur.execute('''
                INSERT INTO gstr2b_entries (user_id, client_id, financial_year, month, state, supplier_gstin, supplier_name, invoice_number, invoice_date, taxable_value, cgst, sgst, igst)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (user_id, client_id, fy, month, state, ent["gstin"], ent["vendor_name"], ent["invoice_number"], ent["invoice_date"],
                  ent["taxable_value"], ent["cgst"], ent["sgst"], ent["igst"]))
            inserted += 1

        conn.commit()
        cur.close()
        conn.close()

        log_activity(user_id, 'gstr2b_upload', f'Imported {inserted} GSTR-2B entries ({state})', fy, month, inserted)

        return jsonify({"success": True, "count": inserted})
    except Exception as e:
        print(f"Error uploading GSTR-2B: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete-gstr2b', methods=['POST'])
@login_required
def delete_gstr2b():
    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    data = request.json or {}
    fy = (data.get('financial_year') or '').strip()
    month = (data.get('month') or '').strip()
    state = (data.get('state') or '').strip()
    if not fy or not month or not state:
        return jsonify({"error": "Financial Year, Month, and State are required"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        if is_admin:
            cur.execute('DELETE FROM gstr2b_entries WHERE client_id = %s AND financial_year = %s AND month = %s AND state = %s', (client_id, fy, month, state))
        else:
            cur.execute('DELETE FROM gstr2b_entries WHERE user_id = %s AND client_id = %s AND financial_year = %s AND month = %s AND state = %s', (user_id, client_id, fy, month, state))
        deleted = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()

        log_activity(user_id, 'gstr2b_deleted', f'Deleted {deleted} GSTR-2B entries for {month} ({fy}, {state})', fy, month, deleted)
        return jsonify({"success": True, "count": deleted})
    except Exception as e:
        print(f"Error deleting GSTR-2B entries: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete-gstr2b-entry', methods=['POST'])
@login_required
def delete_gstr2b_entry():
    """Deletes an individual GSTR-2B entry directly by its ID from the reconciliation table."""
    user_id = session['user_id']
    is_admin = is_admin_user()
    data = request.json or {}
    entry_id = data.get('id')
    if not entry_id:
        return jsonify({"error": "Entry ID is required"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        if is_admin:
            cur.execute('SELECT supplier_name, invoice_number, financial_year, month FROM gstr2b_entries WHERE id = %s', (entry_id,))
        else:
            cur.execute('SELECT supplier_name, invoice_number, financial_year, month FROM gstr2b_entries WHERE id = %s AND user_id = %s', (entry_id, user_id))
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return jsonify({"error": "GSTR-2B entry not found"}), 404

        supp_name = row['supplier_name']
        inv_no = row['invoice_number']
        fy = row['financial_year']
        m = row['month']

        if is_admin:
            cur.execute('DELETE FROM gstr2b_entries WHERE id = %s', (entry_id,))
        else:
            cur.execute('DELETE FROM gstr2b_entries WHERE id = %s AND user_id = %s', (entry_id, user_id))
        conn.commit()
        cur.close()
        conn.close()

        log_activity(user_id, 'gstr2b_entry_deleted', f"Deleted GSTR-2B entry for '{supp_name}' (Invoice #{inv_no})", fy, m, 1)
        return jsonify({"success": True, "message": "GSTR-2B entry deleted successfully"})
    except Exception as e:
        print(f"Error deleting single GSTR-2B entry: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/gstr2b-status', methods=['GET'])
@login_required
def gstr2b_status():
    """Returns a summary of all uploaded GSTR-2B batches (by State, Month, and FY)
    with total entry counts and tax amounts so users can clearly see what's loaded."""
    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    fy = request.args.get('financial_year', '').strip()

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        if is_admin:
            if fy:
                cur.execute('''
                    SELECT state, month, financial_year, COUNT(*) as count,
                           COALESCE(SUM(taxable_value), 0)::float as total_taxable,
                           COALESCE(SUM(cgst + sgst + igst), 0)::float as total_gst
                    FROM gstr2b_entries
                    WHERE client_id = %s AND financial_year = %s
                    GROUP BY state, month, financial_year
                    ORDER BY state, month
                ''', (client_id, fy))
            else:
                cur.execute('''
                    SELECT state, month, financial_year, COUNT(*) as count,
                           COALESCE(SUM(taxable_value), 0)::float as total_taxable,
                           COALESCE(SUM(cgst + sgst + igst), 0)::float as total_gst
                    FROM gstr2b_entries
                    WHERE client_id = %s
                    GROUP BY state, month, financial_year
                    ORDER BY financial_year, state, month
                ''')
        else:
            if fy:
                cur.execute('''
                    SELECT state, month, financial_year, COUNT(*) as count,
                           COALESCE(SUM(taxable_value), 0)::float as total_taxable,
                           COALESCE(SUM(cgst + sgst + igst), 0)::float as total_gst
                    FROM gstr2b_entries
                    WHERE user_id = %s AND client_id = %s AND financial_year = %s
                    GROUP BY state, month, financial_year
                    ORDER BY state, month
                ''', (user_id, client_id, fy))
            else:
                cur.execute('''
                    SELECT state, month, financial_year, COUNT(*) as count,
                           COALESCE(SUM(taxable_value), 0)::float as total_taxable,
                           COALESCE(SUM(cgst + sgst + igst), 0)::float as total_gst
                    FROM gstr2b_entries
                    WHERE user_id = %s AND client_id = %s
                    GROUP BY state, month, financial_year
                    ORDER BY financial_year, state, month
                ''', (user_id, client_id))
        rows = [dict(r) for r in cur.fetchall()]
        cur.close()
        conn.close()
        return jsonify({"success": True, "batches": rows})
    except Exception as e:
        print(f"Error fetching GSTR-2B status: {e}")
        return jsonify({"error": str(e)}), 500

def execute_reconciliation(fy, months, user_id, is_admin, client_id='nutan_nagrik'):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if is_admin:
        cur.execute('''
            SELECT id, invoice_number, invoice_date, vendor_name, gstin, branch, state,
                   taxable_value::float, cgst::float, sgst::float, igst::float,
                   eligible_itc::float, ineligible_itc::float,
                   (cgst::float + sgst::float + igst::float) as total_gst,
                   (file_data IS NOT NULL) AS has_file
            FROM invoices
            WHERE client_id = %s AND financial_year = %s AND month = ANY(%s)
        ''', (client_id, fy, months))
    else:
        cur.execute('''
            SELECT id, invoice_number, invoice_date, vendor_name, gstin, branch, state,
                   taxable_value::float, cgst::float, sgst::float, igst::float,
                   eligible_itc::float, ineligible_itc::float,
                   (cgst::float + sgst::float + igst::float) as total_gst,
                   (file_data IS NOT NULL) AS has_file
            FROM invoices
            WHERE user_id = %s AND client_id = %s AND financial_year = %s AND month = ANY(%s)
        ''', (user_id, client_id, fy, months))
    books_invoices = cur.fetchall()

    if is_admin:
        cur.execute('''
            SELECT id, invoice_number, invoice_date, supplier_name as vendor_name, supplier_gstin as gstin, state,
                   taxable_value::float, cgst::float, sgst::float, igst::float,
                   (cgst::float + sgst::float + igst::float) as total_gst
            FROM gstr2b_entries
            WHERE client_id = %s AND financial_year = %s AND month = ANY(%s)
        ''', (client_id, fy, months))
    else:
        cur.execute('''
            SELECT id, invoice_number, invoice_date, supplier_name as vendor_name, supplier_gstin as gstin, state,
                   taxable_value::float, cgst::float, sgst::float, igst::float,
                   (cgst::float + sgst::float + igst::float) as total_gst
            FROM gstr2b_entries
            WHERE user_id = %s AND client_id = %s AND financial_year = %s AND month = ANY(%s)
        ''', (user_id, client_id, fy, months))
    portal_entries = cur.fetchall()

    cur.close()
    conn.close()

    def match_bucket(books_subset, portal_subset):
        """Runs the exact-match + near-match algorithm within one state's
        books/portal pool only -- called once per state so a book invoice can
        never be matched against another state's GSTR-2B entry."""
        portal_pool = collections.defaultdict(list)
        for pe in portal_subset:
            gst = pe['gstin'].strip().upper()
            num = clean_invoice_number(pe['invoice_number'])
            if not num:
                continue
            portal_pool[(gst, num)].append(pe)

        bucket_reconciled = []
        matched_count = 0
        mismatched_count = 0
        missing_portal_count = 0
        matched_portal_ids = set()

        for bi in books_subset:
            bgst = bi['gstin'].strip().upper()
            bnum = clean_invoice_number(bi['invoice_number'])

            candidates = portal_pool.get((bgst, bnum), []) if bnum else []
            candidates = [c for c in candidates if c['id'] not in matched_portal_ids]

            if not candidates:
                bucket_reconciled.append({
                    "status": "Missing in GSTR-2B",
                    "book": bi,
                    "portal": None
                })
                missing_portal_count += 1
            else:
                best_cand = candidates[0]
                if len(candidates) > 1:
                    for c in candidates:
                        if abs(c['total_gst'] - bi['total_gst']) <= 10.0:
                            best_cand = c
                            break

                matched_portal_ids.add(best_cand['id'])
                tax_diff = abs(best_cand['total_gst'] - bi['total_gst'])
                taxable_diff = abs(best_cand['taxable_value'] - bi['taxable_value'])

                if tax_diff <= 10.0 and taxable_diff <= 10.0:
                    bucket_reconciled.append({
                        "status": "Matched",
                        "book": bi,
                        "portal": best_cand
                    })
                    matched_count += 1
                else:
                    bucket_reconciled.append({
                        "status": "Value Mismatched",
                        "book": bi,
                        "portal": best_cand
                    })
                    mismatched_count += 1

        missing_books_count = 0
        for pe in portal_subset:
            if pe['id'] not in matched_portal_ids:
                bucket_reconciled.append({
                    "status": "Missing in Books",
                    "book": None,
                    "portal": pe
                })
                missing_books_count += 1

        # Second pass: near-match detection. OCR commonly misreads a single
        # character (0/O, P/F, 1/I, 5/S...), so the same real-world invoice can
        # fail the strict GSTIN+invoice-number key match above and show up as
        # two separate, unexplained "missing" rows. Amounts identical to the
        # existing ±10 tolerance plus a small edit distance on GSTIN/invoice
        # number is a strong signal it's the same bill -- flag it for a human
        # to confirm rather than auto-merging (a wrong merge would hide a
        # genuine discrepancy).
        unmatched_book = [r for r in bucket_reconciled if r["status"] == "Missing in GSTR-2B"]
        unmatched_portal = [r for r in bucket_reconciled if r["status"] == "Missing in Books"]
        used_portal_ids = set()
        possible_match_count = 0

        for br in unmatched_book:
            bi = br["book"]
            bgst = bi['gstin'].strip().upper()
            bnum = clean_invoice_number(bi['invoice_number'])
            best = None
            best_score = None

            for pr in unmatched_portal:
                pe = pr["portal"]
                if pe['id'] in used_portal_ids:
                    continue
                if abs(pe['total_gst'] - bi['total_gst']) > 10.0 or abs(pe['taxable_value'] - bi['taxable_value']) > 10.0:
                    continue
                gstin_dist = levenshtein(bgst, pe['gstin'].strip().upper())
                num_dist = levenshtein(bnum, clean_invoice_number(pe['invoice_number']))
                if gstin_dist == 0 and num_dist == 0:
                    continue  # would already have matched exactly above
                if gstin_dist > 2 or num_dist > 3:
                    continue
                score = gstin_dist + num_dist
                if best is None or score < best_score:
                    best = pr
                    best_score = score

            if best is not None:
                used_portal_ids.add(best["portal"]['id'])
                br["status"] = "Possible Match"
                br["portal"] = best["portal"]
                possible_match_count += 1

        if used_portal_ids:
            bucket_reconciled = [r for r in bucket_reconciled
                                  if not (r["status"] == "Missing in Books" and r["portal"]['id'] in used_portal_ids)]
            missing_portal_count -= possible_match_count
            missing_books_count -= possible_match_count

        return bucket_reconciled, matched_count, mismatched_count, missing_portal_count, missing_books_count, possible_match_count

    # Partition strictly by state (GST registration) before matching -- a
    # Gujarat book invoice must never be compared against a Maharashtra
    # GSTR-2B entry, or vice versa. Legacy rows with no state tag fall into
    # their own "Unassigned" bucket and only match each other.
    states_present = sorted({(bi['state'] or 'Unassigned') for bi in books_invoices} |
                             {(pe['state'] or 'Unassigned') for pe in portal_entries})

    reconciled = []
    matched_count = 0
    mismatched_count = 0
    missing_portal_count = 0
    missing_books_count = 0
    possible_match_count = 0

    for st in states_present:
        books_subset = [bi for bi in books_invoices if (bi['state'] or 'Unassigned') == st]
        portal_subset = [pe for pe in portal_entries if (pe['state'] or 'Unassigned') == st]

        bucket_items, mc, mmc, mpc, mbc, pmc = match_bucket(books_subset, portal_subset)
        for item in bucket_items:
            item["state"] = st
        reconciled.extend(bucket_items)

        matched_count += mc
        mismatched_count += mmc
        missing_portal_count += mpc
        missing_books_count += mbc
        possible_match_count += pmc

    summary = {
        "total_books": len(books_invoices),
        "total_portal": len(portal_entries),
        "matched": matched_count,
        "mismatched": mismatched_count,
        "missing_in_portal": missing_portal_count,
        "missing_in_books": missing_books_count,
        "possible_match": possible_match_count
    }

    return summary, reconciled, books_invoices, portal_entries


@app.route('/api/reconcile-data', methods=['GET'])
@login_required
def reconcile_data():
    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    fy = request.args.get('financial_year', '').strip()
    months_str = request.args.get('months', '').strip()

    if not fy or not months_str:
        return jsonify({"error": "Financial Year and Month are required"}), 400

    months = [m.strip() for m in months_str.split(',') if m.strip()]
    if not months:
        return jsonify({"error": "At least one month is required"}), 400

    try:
        summary, reconciled, _, _ = execute_reconciliation(fy, months, user_id, is_admin, client_id=client_id)
        return jsonify({
            "summary": summary,
            "items": reconciled
        })
    except Exception as e:
        print(f"Error executing reconciliation: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/vendor-discrepancies', methods=['GET'])
@login_required
def get_vendor_discrepancies():
    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    fy = request.args.get('financial_year', '').strip()
    months_str = request.args.get('months', '').strip()

    if not fy or not months_str:
        return jsonify({"error": "Financial Year and Month are required"}), 400
    months = [m.strip() for m in months_str.split(',') if m.strip()]

    try:
        summary, reconciled, _, _ = execute_reconciliation(fy, months, user_id, is_admin, client_id=client_id)

        vendors_dict = collections.defaultdict(lambda: {
            "gstin": "",
            "vendor_name": "",
            "missing_invoices": [],
            "mismatched_invoices": [],
            "total_tax_on_hold": 0.0
        })

        for item in reconciled:
            status = item["status"]
            if status in ["Missing in GSTR-2B", "Value Mismatched"]:
                book = item["book"]
                if not book:
                    continue
                gstin = book["gstin"].strip().upper()
                vname = book["vendor_name"].strip()
                key = (gstin, vname)

                entry = vendors_dict[key]
                entry["gstin"] = gstin
                entry["vendor_name"] = vname

                if status == "Missing in GSTR-2B":
                    entry["missing_invoices"].append({
                        "invoice_number": book["invoice_number"],
                        "invoice_date": book["invoice_date"],
                        "taxable_value": book["taxable_value"],
                        "cgst": book["cgst"],
                        "sgst": book["sgst"],
                        "igst": book["igst"],
                        "total_gst": book["total_gst"]
                    })
                    entry["total_tax_on_hold"] += book["total_gst"]
                elif status == "Value Mismatched":
                    portal = item["portal"]
                    tax_diff = abs((portal["total_gst"] if portal else 0.0) - book["total_gst"])
                    entry["mismatched_invoices"].append({
                        "invoice_number": book["invoice_number"],
                        "invoice_date": book["invoice_date"],
                        "book_taxable": book["taxable_value"],
                        "book_gst": book["total_gst"],
                        "portal_taxable": portal["taxable_value"] if portal else 0.0,
                        "portal_gst": portal["total_gst"] if portal else 0.0,
                        "tax_diff": tax_diff
                    })
                    entry["total_tax_on_hold"] += book["total_gst"]

        vendor_list = sorted(list(vendors_dict.values()), key=lambda x: x["total_tax_on_hold"], reverse=True)

        return jsonify({
            "success": True,
            "financial_year": fy,
            "months": months,
            "total_vendors": len(vendor_list),
            "vendors": vendor_list
        })
    except Exception as e:
        print(f"Error fetching vendor discrepancies: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/generate-vendor-notice', methods=['POST'])
@login_required
def generate_vendor_notice():
    data = request.json or {}
    gstin = data.get('gstin', '').strip().upper()
    vname = data.get('vendor_name', '').strip()
    fy = data.get('financial_year', '').strip()
    months = data.get('months', [])

    if isinstance(months, str):
        months = [m.strip() for m in months.split(',') if m.strip()]

    if not gstin or not fy or not months:
        return jsonify({"error": "GSTIN, Financial Year, and Months are required"}), 400

    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()

    try:
        _, reconciled, _, _ = execute_reconciliation(fy, months, user_id, is_admin, client_id=client_id)

        missing_list = []
        mismatched_list = []
        total_tax_hold = 0.0

        for item in reconciled:
            status = item["status"]
            if status in ["Missing in GSTR-2B", "Value Mismatched"]:
                book = item["book"]
                if not book or book["gstin"].strip().upper() != gstin:
                    continue
                vname = vname or book["vendor_name"]

                if status == "Missing in GSTR-2B":
                    missing_list.append(book)
                    total_tax_hold += book["total_gst"]
                elif status == "Value Mismatched":
                    portal = item["portal"]
                    mismatched_list.append({"book": book, "portal": portal})
                    total_tax_hold += book["total_gst"]

        if not missing_list and not mismatched_list:
            return jsonify({"error": "No discrepancies found for this vendor in selected period."}), 404

        email_subject = f"URGENT: GST ITC Reconciliation Discrepancy Notice - {vname} (GSTIN: {gstin})"

        email_body = f"To,\n{vname}\nGSTIN: {gstin}\n\n"
        email_body += f"Subject: Request to upload/correct purchase invoices in GSTR-1 for {fy} ({', '.join(months)})\n\n"
        email_body += f"Dear Accounts Team,\n\n"
        email_body += f"During our monthly GST Input Tax Credit (ITC) reconciliation for Nutan Nagrik Sahakari Bank Ltd., we identified discrepancies regarding purchase invoices issued by your organization. As per current GST guidelines, ITC cannot be claimed by us until these invoices accurately reflect in GSTR-2B.\n\n"

        if missing_list:
            email_body += "--- MISSING INVOICES IN GSTR-2B ---\n"
            email_body += "The following invoices recorded in our books were NOT found in GSTR-2B statement:\n\n"
            email_body += f"{'Inv No':<18} | {'Inv Date':<12} | {'Taxable (₹)':<12} | {'CGST (₹)':<10} | {'SGST (₹)':<10} | {'IGST (₹)':<10} | {'Total GST (₹)':<12}\n"
            email_body += "-" * 90 + "\n"
            for b in missing_list:
                email_body += f"{b['invoice_number']:<18} | {b['invoice_date']:<12} | {b['taxable_value']:<12.2f} | {b['cgst']:<10.2f} | {b['sgst']:<10.2f} | {b['igst']:<10.2f} | {b['total_gst']:<12.2f}\n"
            email_body += "\n"

        if mismatched_list:
            email_body += "--- VALUE MISMATCHED INVOICES ---\n"
            email_body += "The following invoices have tax/taxable value discrepancies between our books and GSTR-2B:\n\n"
            for m in mismatched_list:
                b = m["book"]
                p = m["portal"]
                email_body += f"Invoice No: {b['invoice_number']} | Date: {b['invoice_date']}\n"
                email_body += f"  - Our Books    : Taxable ₹{b['taxable_value']:,.2f}, Total GST ₹{b['total_gst']:,.2f}\n"
                email_body += f"  - GSTR-2B Portal: Taxable ₹{p['taxable_value']:,.2f}, Total GST ₹{p['total_gst']:,.2f}\n"
                email_body += f"  - Difference   : Tax Diff ₹{abs(p['total_gst'] - b['total_gst']):,.2f}\n\n"

        email_body += f"Total ITC Currently Put On Hold: ₹{total_tax_hold:,.2f}\n\n"
        email_body += "We kindly request you to upload the missing invoices or amend the mismatched details in your upcoming GSTR-1 return filing at the earliest so we can claim the eligible ITC.\n\n"
        email_body += "Thank you for your cooperation.\n\nBest regards,\nAccounts & Tax Department\nNutan Nagrik Sahakari Bank Ltd."

        wa_text = f"*GST Reconciliation Alert - Nutan Nagrik Bank*\n"
        wa_text += f"Vendor: {vname} ({gstin})\n"
        wa_text += f"Period: {fy} ({', '.join(months)})\n"
        wa_text += f"Discrepancy: {len(missing_list)} Missing, {len(mismatched_list)} Mismatched\n"
        wa_text += f"Total Tax Credit On Hold: ₹{total_tax_hold:,.2f}\n\n"
        wa_text += f"Please check your email or contact us to resolve invoice filings in GSTR-1. Thank you!"

        return jsonify({
            "success": True,
            "vendor_name": vname,
            "gstin": gstin,
            "email_subject": email_subject,
            "email_body": email_body,
            "whatsapp_text": wa_text,
            "total_tax_on_hold": total_tax_hold
        })
    except Exception as e:
        print(f"Error generating vendor notice: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/gstr3b-summary', methods=['GET'])
@login_required
def get_gstr3b_summary():
    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    fy = request.args.get('financial_year', '').strip()
    months_str = request.args.get('months', '').strip()

    if not fy or not months_str:
        return jsonify({"error": "Financial Year and Month are required"}), 400
    months = [m.strip() for m in months_str.split(',') if m.strip()]

    try:
        summary, reconciled, books_invoices, _ = execute_reconciliation(fy, months, user_id, is_admin, client_id=client_id)

        itc_4a5_matched = {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "total": 0.0}
        itc_4b2_ineligible = {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "total": 0.0}
        itc_4d1_pending = {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "total": 0.0}

        for bi in books_invoices:
            cgst = bi.get("cgst", 0.0)
            sgst = bi.get("sgst", 0.0)
            igst = bi.get("igst", 0.0)
            taxable = bi.get("taxable_value", 0.0)
            total_gst = cgst + sgst + igst
            inelig = bi.get("ineligible_itc", 0.0)
            # Was hardcoded to a flat 50%, which silently halved the breakdown
            # for ITC-blocked invoices (100% ineligible) -- the total column
            # used the real stored value while cgst/sgst/igst/taxable didn't,
            # so they no longer summed to the total for any blocked invoice.
            inelig_ratio = inelig / total_gst if total_gst > 0 else 0.0

            itc_4b2_ineligible["taxable"] += taxable * inelig_ratio
            itc_4b2_ineligible["cgst"] += cgst * inelig_ratio
            itc_4b2_ineligible["sgst"] += sgst * inelig_ratio
            itc_4b2_ineligible["igst"] += igst * inelig_ratio
            itc_4b2_ineligible["total"] += inelig

        # GSTR-3B Table 4D(1) is reported as one combined "pending" figure
        # (that's how the actual return works -- there's no separate line
        # for missing vs. mismatched). But the Stage 3 KPI tiles above need
        # missing and mismatched split apart, since they call for different
        # vendor follow-up actions. Track both alongside the combined total
        # rather than trying to re-derive one from the other.
        missing_only_total = 0.0
        mismatched_only_total = 0.0

        for item in reconciled:
            status = item["status"]
            book = item.get("book")
            if not book:
                continue

            cgst = book.get("cgst", 0.0)
            sgst = book.get("sgst", 0.0)
            igst = book.get("igst", 0.0)
            taxable = book.get("taxable_value", 0.0)
            total_gst = book.get("total_gst", 0.0)
            elig_gst = book.get("eligible_itc", total_gst * 0.5)
            elig_ratio = elig_gst / total_gst if total_gst > 0 else 0.5

            if status == "Matched":
                itc_4a5_matched["taxable"] += taxable * elig_ratio
                itc_4a5_matched["cgst"] += cgst * elig_ratio
                itc_4a5_matched["sgst"] += sgst * elig_ratio
                itc_4a5_matched["igst"] += igst * elig_ratio
                itc_4a5_matched["total"] += elig_gst
            elif status in ["Missing in GSTR-2B", "Value Mismatched"]:
                itc_4d1_pending["taxable"] += taxable * elig_ratio
                itc_4d1_pending["cgst"] += cgst * elig_ratio
                itc_4d1_pending["sgst"] += sgst * elig_ratio
                itc_4d1_pending["igst"] += igst * elig_ratio
                itc_4d1_pending["total"] += elig_gst
                if status == "Missing in GSTR-2B":
                    missing_only_total += elig_gst
                else:
                    mismatched_only_total += elig_gst

        return jsonify({
            "success": True,
            "financial_year": fy,
            "months": months,
            "gstr3b": {
                "table_4a5_all_other_itc": itc_4a5_matched,
                "table_4b2_ineligible_itc": itc_4b2_ineligible,
                "table_4d1_pending_itc": itc_4d1_pending,
                "missing_only_total": missing_only_total,
                "mismatched_only_total": mismatched_only_total
            }
        })
    except Exception as e:
        print(f"Error computing GSTR-3B summary: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/export-vendor-discrepancies', methods=['GET'])
@login_required
def export_vendor_discrepancies():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    fy = request.args.get('financial_year', '').strip()
    months_str = request.args.get('months', '').strip()

    if not fy or not months_str:
        return jsonify({"error": "Financial Year and Month are required"}), 400
    months = [m.strip() for m in months_str.split(',') if m.strip()]

    try:
        _, reconciled, _, _ = execute_reconciliation(fy, months, user_id, is_admin, client_id=client_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "Vendor Discrepancy Report"

        header_fill = PatternFill(start_color="3525CD", end_color="3525CD", fill_type="solid")
        white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10)
        border_thin = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
                             top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD'))

        red_fill = PatternFill(start_color="FFDAD6", end_color="FFDAD6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        headers = ["Supplier GSTIN", "Vendor Name", "Discrepancy Status", "Invoice No", "Date", "Book Taxable (₹)", "Book GST (₹)", "Portal Taxable (₹)", "Portal GST (₹)", "Tax Difference (₹)", "Action Required"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")

        for item in reconciled:
            status = item["status"]
            if status not in ["Missing in GSTR-2B", "Value Mismatched"]:
                continue

            book = item["book"]
            portal = item["portal"]

            gstin = book["gstin"] if book else portal["gstin"]
            vendor = book["vendor_name"] if book else portal["vendor_name"]
            inv_no = book["invoice_number"] if book else portal["invoice_number"]
            inv_date = book["invoice_date"] if book else portal["invoice_date"]

            b_taxable = book["taxable_value"] if book else 0.0
            b_gst = book["total_gst"] if book else 0.0
            p_taxable = portal["taxable_value"] if portal else 0.0
            p_gst = portal["total_gst"] if portal else 0.0

            tax_diff = abs(p_gst - b_gst)
            action = "Upload in GSTR-1" if status == "Missing in GSTR-2B" else "Amend Tax Amount in GSTR-1"

            row_vals = [gstin, vendor, status, inv_no, inv_date, b_taxable, b_gst, p_taxable, p_gst, tax_diff, action]
            ws.append(row_vals)

            curr_row = ws.max_row
            fill_color = red_fill if status == "Missing in GSTR-2B" else yellow_fill
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                cell.font = regular_font
                cell.border = border_thin
                if col_idx in [3, 11]:
                    cell.fill = fill_color
                if col_idx in [6, 7, 8, 9, 10]:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        log_activity(user_id, 'export_vendor_discrepancy', 'Exported vendor discrepancy report', fy, ','.join(months), ws.max_row - 1)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"Vendor_Discrepancy_Report_{fy}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        print(f"Error exporting vendor discrepancy excel: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/filing-history')
@login_required
def filing_history():
    return render_template('filing_history.html', is_admin=is_admin_user(), api_key_configured=bool(ANTHROPIC_API_KEY or OPENROUTER_API_KEY), ai_model_name=AI_MODEL_DISPLAY_NAME)

@app.route('/api/filing-history', methods=['GET'])
@login_required
def get_filing_history():
    user_id = session['user_id']
    is_admin = is_admin_user()
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if is_admin:
            cur.execute('''
                SELECT activity_log.id, action, description, financial_year, month, record_count,
                       activity_log.created_at, users.username
                FROM activity_log
                JOIN users ON users.id = activity_log.user_id
                ORDER BY activity_log.created_at DESC
                LIMIT 200
            ''')
        else:
            cur.execute('''
                SELECT id, action, description, financial_year, month, record_count, created_at
                FROM activity_log
                WHERE user_id = %s
                ORDER BY created_at DESC
                LIMIT 200
            ''', (user_id,))
        rows = cur.fetchall()
        cur.close()
        conn.close()
        for r in rows:
            if r.get('created_at'):
                r['created_at'] = r['created_at'].isoformat()
        return jsonify({"success": True, "history": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/monthly-bill-summary', methods=['GET'])
@login_required
def get_monthly_bill_summary():
    """Month-wise count of bills entered, broken down by user for admins.
    Read directly from the invoices table (not activity_log) so a bulk
    upload spanning several months is still attributed to the correct
    month per bill rather than lumped into one aggregate entry."""
    user_id = session['user_id']
    is_admin = is_admin_user()
    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        if is_admin:
            cur.execute('''
                SELECT invoices.financial_year, invoices.month, users.username, COUNT(*) AS bill_count
                FROM invoices
                JOIN users ON users.id = invoices.user_id
                WHERE invoices.financial_year IS NOT NULL AND invoices.month IS NOT NULL
                GROUP BY invoices.financial_year, invoices.month, users.username
            ''')
        else:
            cur.execute('''
                SELECT financial_year, month, COUNT(*) AS bill_count
                FROM invoices
                WHERE user_id = %s AND financial_year IS NOT NULL AND month IS NOT NULL
                GROUP BY financial_year, month
            ''', (user_id,))
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # Stable multi-key sort: FY descending (most recent first), then
        # calendar month within the FY (April..March), then username.
        rows.sort(key=lambda r: r.get('username') or '')
        rows.sort(key=lambda r: month_sort_key(r['month']))
        rows.sort(key=lambda r: r['financial_year'] or '', reverse=True)

        return jsonify({"success": True, "summary": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500




# ============================================================================
# INCOME & OUTPUT GST MODULE (CBS Bank Statements & Branch-Wise Working Sheet)
# ============================================================================

INCOME_CATALOG_PATH = os.path.join(os.path.dirname(__file__), 'reference_data', 'income_master_catalog.json')
INCOME_MASTER_CODES = []
INCOME_MASTER_BRANCHES = [
    'ODHAV', 'RAKHIAL', 'NEW SHARDA', 'CHANGODAR', 'ISANPUR', 'MANINAGAR',
    'SHANTI COMM', 'MASKATI', 'VEJALPUR', 'JODHPUR-SATELLITE', 'PANJRAPOLE',
    'ASHRAM ROAD', 'NARAYANNAGAR', 'NARANPURA', 'DRIVE IN', 'VASANA', 'SURAT',
    'LAW GARDEN', 'NEW CLOTH', 'BAPUNAGAR', 'BOPAL', 'THALTEJ', 'CHANDKHEDA',
    'VASTRAL', 'HO', 'DEMAT'
]

if os.path.exists(INCOME_CATALOG_PATH):
    try:
        with open(INCOME_CATALOG_PATH, 'r', encoding='utf-8') as f:
            cat_data = json.load(f)
            INCOME_MASTER_CODES = cat_data.get('income_codes', [])
            if cat_data.get('branches'):
                INCOME_MASTER_BRANCHES = cat_data.get('branches')
    except Exception as e:
        print(f"Error loading income master catalog: {e}")

def get_income_code_meta(code_str):
    clean = str(code_str).strip().upper()
    for m in INCOME_MASTER_CODES:
        if str(m.get('code')).strip().upper() == clean:
            return m
    return None

# ---------------------------------------------------------------------------
# Ledger statement parsing (reads the bank's actual GL/PL account statements)
#
# The core banking system exports these in four different layouts depending
# on branch and report type:
#   1. Multi-account ledger ("Statement of Acct with Narration") - many
#      accounts in one file. Anchor: "Account ID" line / "Total :" line.
#   2. Single-account "R009007 - Statement of GL Account" report.
#      Anchor: "Account Id" (code embedded as "CODE - NAME") / bare "Total"
#      line (which is preceded by an entry-count token, unlike format 1).
#   3. Single-account "Statement Of Account" report. Anchor: "A/c No" /
#      "Closing Balance" (the summary row, not the opening-balance row).
#   4. "Acct No : CODE / NAME" combined single-line variant / "Total :".
# Each is read as text (PDF) or flattened cell-by-cell (Excel, in row-major
# order) into the same line stream and walked by one shared scanner, so a
# period's income is always "credits total minus debits total" taken from
# the account's own printed total - never the closing balance, which is a
# cumulative running balance and would double-count every prior period.
#
# Two GL codes (1836 rent-not-accrued -> 3320 rent on locker; 1720 commission
# -not-accrued -> 3230 commission on guarantee) sometimes carry an internal
# reclass transfer between them when an item closes/accrues. On the two
# branches checked first (single locker closure that month) "deferred code's
# income = gross credits, paired code's income = net minus the deferred
# code's period debit total" reproduced golden exactly. It does NOT
# generalise: branches with more than one closure/accrual event in the month
# make it worse, not better (the deferred code's debit total isn't purely
# 1:1 with the paired code's credit - some of it is other activity on that
# account). Proper handling needs transaction-line-level detection of the
# specific reclass entries, not a period-total adjustment, so this is left
# as a plain net (credits minus debits) per account - matching golden on
# most branches, with a small, self-explained, self-cancelling residual
# on branches that had a reclass event that month (see the verification
# report). RECLASS_DEFERRED_TO_INCOME is kept empty deliberately.
RECLASS_DEFERRED_TO_INCOME = {}

_LEDGER_NUM_RE = re.compile(r'^-?[\d,]+(?:\.\d+)?$')
_ACCTNO_COLON_RE = re.compile(r'^Acct No\s*:\s*(\d{3,6})\s*/\s*(.*)$', re.IGNORECASE)
_LEDGER_PDF_ACCOUNT_LABELS = ('account id', 'account id.')
# format 1 ("Total :"/"Total:") has no leading count; format 2 (bare "Total")
# prints a leading entry-count token before Dr/Cr.
_LEDGER_PDF_TOTAL_LABELS = {'total :': 0, 'total:': 0, 'total': 1}
_LEDGER_XLSX_ACCOUNT_LABELS = ('a/c no', 'a/c no.')
_LEDGER_XLSX_TOTAL_LABELS = {'closing balance': 0}


def _ledger_num(tok):
    tok = tok.strip().replace(',', '')
    if not tok:
        return None
    try:
        return float(tok)
    except ValueError:
        return None


def _extract_ledger_code(cand):
    cand = cand.strip()
    m2 = re.match(r'^(\d{3,6})\s*-\s*(.+)$', cand)
    if m2:
        return m2.group(1), m2.group(2).strip()
    if '/' in cand:
        parts = [p.strip() for p in cand.split('/') if p.strip()]
        if parts and re.match(r'^\d{3,6}$', parts[-1]):
            return parts[-1], None
    if re.match(r'^\d{6,}$', cand):
        return cand[-4:], None
    m4 = re.match(r'^(\d{3,6})\b', cand)
    if m4:
        return m4.group(1), None
    return None, None


def _scan_ledger_lines(lines, account_labels, total_labels):
    n = len(lines)
    accounts = []
    i = 0
    while i < n:
        line = lines[i].strip()
        if line.lower() in account_labels:
            j = i + 1
            code = None
            name = None
            look = j
            while look < n and look < j + 8:
                cand = lines[look].strip()
                if cand.lower() in account_labels:
                    i = look
                    j = look + 1
                    look = j
                    continue
                c, nm = _extract_ledger_code(cand)
                if c:
                    code = c
                    name = nm
                    if name is None:
                        look2 = look + 1
                        while look2 < n and lines[look2].strip() != 'Name':
                            look2 += 1
                            if look2 > look + 4:
                                break
                        if look2 < n and lines[look2].strip() == 'Name':
                            nn = look2 + 1
                            while nn < n and not lines[nn].strip():
                                nn += 1
                            if nn < n:
                                name = lines[nn].strip()
                    break
                look += 1

            if code is None:
                i += 1
                continue

            k = look
            total_dr = total_cr = None
            while k < n:
                lk = lines[k].strip()
                if lk.lower() in account_labels:
                    break
                if lk.lower() in total_labels:
                    skip = total_labels[lk.lower()]
                    nums = []
                    kk = k + 1
                    steps = 0
                    while kk < n and len(nums) < 2 + skip and steps < 20:
                        val = lines[kk].strip()
                        if val.lower() in account_labels:
                            break
                        if _LEDGER_NUM_RE.match(val):
                            nums.append(_ledger_num(val))
                        kk += 1
                        steps += 1
                    nums = nums[skip:]
                    if len(nums) >= 2:
                        total_dr, total_cr = nums[0], nums[1]
                    k = kk
                    continue
                k += 1

            if total_dr is not None and total_cr is not None:
                accounts.append({
                    'gl_code': code, 'name': name, 'dr': total_dr, 'cr': total_cr,
                    'net': round(total_cr - total_dr, 2),
                })
            i = k
        else:
            i += 1
    return accounts


def _scan_acctno_colon_format(lines):
    n = len(lines)
    accounts = []
    i = 0
    while i < n:
        m = _ACCTNO_COLON_RE.match(lines[i].strip())
        if m:
            code, name = m.group(1), m.group(2).strip()
            k = i + 1
            total_dr = total_cr = None
            while k < n:
                lk = lines[k].strip()
                if _ACCTNO_COLON_RE.match(lk):
                    break
                if lk.lower() in ('total :', 'total:'):
                    nums = []
                    kk = k + 1
                    steps = 0
                    while kk < n and len(nums) < 2 and steps < 20:
                        val = lines[kk].strip()
                        if _ACCTNO_COLON_RE.match(val):
                            break
                        if _LEDGER_NUM_RE.match(val):
                            nums.append(_ledger_num(val))
                        kk += 1
                        steps += 1
                    if len(nums) >= 2:
                        total_dr, total_cr = nums[0], nums[1]
                    k = kk
                    continue
                k += 1
            if total_dr is not None and total_cr is not None:
                accounts.append({
                    'gl_code': code, 'name': name, 'dr': total_dr, 'cr': total_cr,
                    'net': round(total_cr - total_dr, 2),
                })
            i = k
        else:
            i += 1
    return accounts


def _flatten_ledger_rows(rows_iter):
    """Row-major cell flatten -> text lines, matching PDF-text conventions.
    Numeric cells are re-rendered as "X.XX" so they match the same decimal
    token pattern as PDF-extracted text (already literal "X.XX")."""
    lines = []
    for row in rows_iter:
        for v in row:
            if v is None or isinstance(v, bool):
                continue
            s = f'{v:.2f}' if isinstance(v, (int, float)) else str(v).strip()
            if s == '':
                continue
            lines.append(s)
    return lines


def _detect_ledger_branch(text, filename):
    for b in INCOME_MASTER_BRANCHES:
        if b.upper() in filename.upper() or b.upper() in text.upper():
            return b
    for b in INCOME_MASTER_BRANCHES:
        if re.search(r'\b' + re.escape(b) + r'\b', filename, re.IGNORECASE):
            return b
    return None


def extract_raw_ledger_accounts_pdf(file_bytes, filename):
    doc = pymupdf.open(stream=file_bytes, filetype="pdf")
    text = ""
    for page in doc:
        text += page.get_text("text") + "\n"
    lines = text.split('\n')
    if re.search(r'Acct No\s*:', text, re.IGNORECASE):
        accounts = _scan_acctno_colon_format(lines)
    elif re.search(r'A/c No', text, re.IGNORECASE):
        accounts = _scan_ledger_lines(lines, _LEDGER_XLSX_ACCOUNT_LABELS, _LEDGER_XLSX_TOTAL_LABELS)
    else:
        accounts = _scan_ledger_lines(lines, _LEDGER_PDF_ACCOUNT_LABELS, _LEDGER_PDF_TOTAL_LABELS)
    branch = _detect_ledger_branch(text, filename)
    for a in accounts:
        a['branch'] = branch
        a['filename'] = filename
    return accounts


def extract_raw_ledger_accounts_xlsx(file_bytes, filename):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    all_accounts = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]
        flat_text = " ".join(str(v) for row in rows for v in row if v)
        if re.search(r'A/c No', flat_text, re.IGNORECASE):
            lines = _flatten_ledger_rows(rows)
            accounts = _scan_ledger_lines(lines, _LEDGER_XLSX_ACCOUNT_LABELS, _LEDGER_XLSX_TOTAL_LABELS)
        elif re.search(r'Account\s*Id', flat_text, re.IGNORECASE):
            lines = _flatten_ledger_rows(rows)
            accounts = _scan_ledger_lines(lines, _LEDGER_PDF_ACCOUNT_LABELS, _LEDGER_PDF_TOTAL_LABELS)
        else:
            continue
        branch = _detect_ledger_branch(flat_text, filename)
        for a in accounts:
            a['branch'] = branch
            a['filename'] = filename
        all_accounts.extend(accounts)
    return all_accounts


def extract_raw_ledger_accounts_xls(file_bytes, filename):
    """Legacy .xls: either SpreadsheetML XML (Excel 2003 "XML Spreadsheet"
    saved with a .xls extension) or true binary OLE2 .xls."""
    head = file_bytes[:200]
    all_accounts = []
    if head.startswith(b'<?xml') or b'mso-application' in head:
        import xml.etree.ElementTree as ET
        ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
        root = ET.fromstring(file_bytes)
        sheets = []
        for ws in root.findall('ss:Worksheet', ns):
            table = ws.find('ss:Table', ns)
            if table is None:
                continue
            rows = []
            for row in table.findall('ss:Row', ns):
                vals = [c.find('ss:Data', ns).text if c.find('ss:Data', ns) is not None else None
                        for c in row.findall('ss:Cell', ns)]
                rows.append(vals)
            sheets.append(rows)
    elif head.startswith(b'\xd0\xcf\x11\xe0'):
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheets = [[sh.row_values(r) for r in range(sh.nrows)] for sh in wb.sheets()]
    else:
        return []

    for rows in sheets:
        flat_text = " ".join(str(v) for row in rows for v in row if v not in (None, ''))
        lines = _flatten_ledger_rows(rows)
        if re.search(r'A/c No', flat_text, re.IGNORECASE):
            accounts = _scan_ledger_lines(lines, _LEDGER_XLSX_ACCOUNT_LABELS, _LEDGER_XLSX_TOTAL_LABELS)
        else:
            accounts = _scan_ledger_lines(lines, _LEDGER_PDF_ACCOUNT_LABELS, _LEDGER_PDF_TOTAL_LABELS)
        branch = _detect_ledger_branch(flat_text, filename)
        for a in accounts:
            a['branch'] = branch
            a['filename'] = filename
        all_accounts.extend(accounts)
    return all_accounts


def finalize_ledger_accounts(raw_accounts, financial_year='2026-27', month='July'):
    """Group raw {branch, gl_code, dr, cr, net} accounts by branch, apply the
    deferred/recognised-income reclassification rule (RECLASS_DEFERRED_TO_INCOME),
    and return final entry dicts ready to save to income_entries."""
    by_branch = collections.defaultdict(dict)
    for a in raw_accounts:
        if not a.get('branch') or not a.get('gl_code'):
            continue
        code_map = by_branch[a['branch']]
        if a['gl_code'] not in code_map:
            code_map[a['gl_code']] = a
        # else: same code already seen for this branch in this upload batch -
        # keep the first one found rather than summing, so the same account
        # appearing in more than one uploaded file can't double-count.

    entries = []
    for branch, code_map in by_branch.items():
        finals = {}
        for code, a in code_map.items():
            finals[code] = a['cr'] if code in RECLASS_DEFERRED_TO_INCOME else a['net']
        for deferred_code, income_code in RECLASS_DEFERRED_TO_INCOME.items():
            if deferred_code in code_map and income_code in finals:
                finals[income_code] = round(finals[income_code] - code_map[deferred_code]['dr'], 2)

        for code, amount in finals.items():
            a = code_map[code]
            meta = get_income_code_meta(code)
            particulars = meta.get('particulars') if meta else (a.get('name') or 'Bank Service Income')
            is_taxable = meta.get('is_taxable', True) if meta else True
            rate = meta.get('gst_rate', 18.0) if meta else 18.0
            cgst = round(amount * (rate / 200.0), 2) if is_taxable else 0.0
            entries.append({
                "branch": branch,
                "financial_year": financial_year,
                "month": month,
                "gl_code": code,
                "particulars": particulars,
                "income_amount": round(amount, 2),
                "is_taxable": is_taxable,
                "cgst": cgst,
                "sgst": cgst,
                "igst": 0.0,
                "refund_without_gst": 0.0,
                "refund_with_gst": 0.0,
                "filename": a.get('filename', 'upload'),
            })
    return entries


def parse_full_workbook_excel(file_bytes, filename):
    """Bulk re-import of a complete branch-wise-calculation workbook (has its
    own 'SUMMARY SHEET GST' tab) - reads each branch tab's own columns
    directly rather than treating it as a raw ledger statement."""
    results = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    for bs_name in wb.sheetnames:
        if bs_name in ['SUMMARY SHEET GST', 'Notes', 'GSTN CANCELLTED', 'Sheet1'] or '1' in bs_name:
            continue
        bws = wb[bs_name]
        b_name = bs_name.strip()
        for r in range(7, 75):
            c_no = bws.cell(r, 1).value
            part = bws.cell(r, 2).value
            tot = bws.cell(r, 3).value
            ggst = bws.cell(r, 4).value
            cgst_val = bws.cell(r, 5).value
            igst_val = bws.cell(r, 6).value
            if c_no and part and (tot is not None or cgst_val is not None):
                c_str = str(c_no).strip()
                p_str = str(part).strip()
                if c_str in ['CODE NO.', 'PL', 'GL', 'TOTAL']:
                    continue
                t_val = float(tot or 0.0)
                cgst_amt = float(cgst_val or (t_val * 0.09 if t_val else 0.0))
                sgst_amt = float(ggst or (t_val * 0.09 if t_val else 0.0))
                igst_amt = float(igst_val or 0.0)
                is_tax = (cgst_amt + sgst_amt + igst_amt) > 0 or t_val > 0
                if 'E-STAMPING' in p_str.upper() or 'EXEMPT' in p_str.upper():
                    is_tax = False
                results.append({
                    "branch": b_name,
                    "financial_year": "2026-27",
                    "month": "July",
                    "gl_code": c_str,
                    "particulars": p_str,
                    "income_amount": t_val,
                    "is_taxable": is_tax,
                    "cgst": round(cgst_amt, 2),
                    "sgst": round(sgst_amt, 2),
                    "igst": round(igst_amt, 2),
                    "refund_without_gst": 0.0,
                    "refund_with_gst": 0.0,
                    "filename": filename
                })
    return results

@app.route('/income')
@login_required
def income_page():
    return render_template('income.html')

@app.route('/api/income-codes-master', methods=['GET'])
@login_required
def get_income_codes_master():
    return jsonify({
        "branches": INCOME_MASTER_BRANCHES,
        "codes": INCOME_MASTER_CODES
    })

@app.route('/api/get-income-entries', methods=['GET'])
@login_required
def get_income_entries():
    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    branch = request.args.get('branch', '').strip()
    fy = request.args.get('financial_year', '').strip()
    month = request.args.get('month', '').strip()

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        query = '''
            SELECT id, branch, state, financial_year, month, gl_code, particulars,
                   is_taxable, income_amount::float, cgst::float, sgst::float, igst::float,
                   refund_without_gst::float, refund_with_gst::float, file_name,
                   CASE WHEN file_data IS NOT NULL THEN true ELSE false END as has_file,
                   created_at
            FROM income_entries
            WHERE client_id = %s
        '''
        params = [client_id]
        if not is_admin:
            query += " AND user_id = %s"
            params.append(user_id)
        if branch and branch != 'ALL':
            query += " AND UPPER(branch) = UPPER(%s)"
            params.append(branch)
        if fy and fy != 'ALL':
            query += " AND financial_year = %s"
            params.append(fy)
        if month and month != 'ALL':
            query += " AND month = %s"
            params.append(month)

        query += " ORDER BY branch ASC, gl_code ASC, id ASC"
        cur.execute(query, params)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify({"entries": rows, "client_id": client_id, "branches": INCOME_MASTER_BRANCHES})
    except Exception as e:
        print(f"Error fetching income entries: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/upload-income', methods=['POST'])
@login_required
def upload_income_api():
    user_id = session['user_id']
    client_id = request.form.get('client_id') or get_current_client_id()
    files = request.files.getlist('income_files')
    
    if not files or all(f.filename == '' for f in files):
        return jsonify({"error": "No files selected for income upload"}), 400

    parsed_entries = []  # already-finalized entries (bulk full-workbook uploads)
    raw_accounts = []    # raw {branch, gl_code, dr, cr, net, filename, file_data} - needs finalize_ledger_accounts

    def is_full_workbook_xlsx(fbytes):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(fbytes), data_only=True, read_only=True)
            return 'SUMMARY SHEET GST' in wb.sheetnames
        except Exception:
            return False

    def process_one(ext, fbytes, fname):
        """Returns (finalized_entries, raw_accounts) - exactly one is non-empty."""
        if ext == 'pdf':
            return [], extract_raw_ledger_accounts_pdf(fbytes, fname)
        elif ext == 'xlsx':
            if is_full_workbook_xlsx(fbytes):
                return parse_full_workbook_excel(fbytes, fname), []
            return [], extract_raw_ledger_accounts_xlsx(fbytes, fname)
        elif ext == 'xls':
            return [], extract_raw_ledger_accounts_xls(fbytes, fname)
        return [], []

    for f in files:
        if not f.filename:
            continue
        fname = f.filename
        ext = fname.lower().split('.')[-1] if '.' in fname else ''
        fbytes = f.read()

        if ext in ('pdf', 'xlsx', 'xls'):
            try:
                finalized, raw = process_one(ext, fbytes, fname)
            except Exception as pe:
                print(f"Error parsing income file {fname}: {pe}")
                continue
            for item in finalized:
                item['file_data'] = fbytes
            parsed_entries.extend(finalized)
            for a in raw:
                a['file_data'] = fbytes
            raw_accounts.extend(raw)
        elif ext == 'zip':
            import zipfile
            try:
                with zipfile.ZipFile(io.BytesIO(fbytes)) as z:
                    for zname in z.namelist():
                        zext = zname.lower().split('.')[-1] if '.' in zname else ''
                        if zext in ('pdf', 'xlsx', 'xls') and not zname.startswith('__MACOSX'):
                            zbytes = z.read(zname)
                            try:
                                finalized, raw = process_one(zext, zbytes, zname)
                            except Exception as pe:
                                print(f"Error parsing {zname} in {fname}: {pe}")
                                continue
                            for item in finalized:
                                item['file_data'] = zbytes
                            parsed_entries.extend(finalized)
                            for a in raw:
                                a['file_data'] = zbytes
                            raw_accounts.extend(raw)
            except Exception as ze:
                print(f"Error reading zip file {fname}: {ze}")

    finalized_from_raw = finalize_ledger_accounts(raw_accounts)
    file_data_by_key = {}
    for a in raw_accounts:
        file_data_by_key.setdefault((a.get('branch'), a.get('gl_code')), a.get('file_data'))
    for e in finalized_from_raw:
        e['file_data'] = file_data_by_key.get((e['branch'], e['gl_code']))
    parsed_entries.extend(finalized_from_raw)

    saved_count = 0
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        # Deduplicate parsed entries in memory first
        unique_entries = {}
        for e in parsed_entries:
            key = (client_id, e.get('branch', 'Unassigned').strip().upper(), e.get('financial_year', '2026-27'), e.get('month', 'July'), str(e.get('gl_code', 'N/A')).strip())
            unique_entries[key] = e

        for e in unique_entries.values():
            cur.execute('''
                INSERT INTO income_entries (user_id, client_id, branch, state, financial_year, month, gl_code, particulars, is_taxable, income_amount, cgst, sgst, igst, refund_without_gst, refund_with_gst, file_name, file_data)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (client_id, branch, financial_year, month, gl_code)
                DO UPDATE SET
                    user_id = EXCLUDED.user_id,
                    particulars = EXCLUDED.particulars,
                    is_taxable = EXCLUDED.is_taxable,
                    income_amount = EXCLUDED.income_amount,
                    cgst = EXCLUDED.cgst,
                    sgst = EXCLUDED.sgst,
                    igst = EXCLUDED.igst,
                    refund_without_gst = EXCLUDED.refund_without_gst,
                    refund_with_gst = EXCLUDED.refund_with_gst,
                    file_name = EXCLUDED.file_name,
                    created_at = CURRENT_TIMESTAMP
                RETURNING id;
            ''', (
                user_id, client_id,
                e.get('branch', 'Unassigned'),
                e.get('state', 'Gujarat'),
                e.get('financial_year', '2026-27'),
                e.get('month', 'July'),
                e.get('gl_code', 'N/A'),
                e.get('particulars', 'Income'),
                e.get('is_taxable', True),
                e.get('income_amount', 0.0),
                e.get('cgst', 0.0),
                e.get('sgst', 0.0),
                e.get('igst', 0.0),
                e.get('refund_without_gst', 0.0),
                e.get('refund_with_gst', 0.0),
                e.get('filename', 'upload'),
                psycopg2.Binary(e['file_data']) if e.get('file_data') else None
            ))
            saved_count += 1
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({
            "success": True,
            "saved_count": saved_count,
            "entries_preview": parsed_entries[:10]
        })
    except Exception as e:
        print(f"Error saving uploaded income records: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/income-summary', methods=['GET'])
@login_required
def get_income_summary():
    user_id = session['user_id']
    is_admin = is_admin_user()
    client_id = get_current_client_id()
    fy = request.args.get('financial_year', '').strip()
    month = request.args.get('month', '').strip()

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        
        q_inc = '''
            SELECT COUNT(*) as total_entries,
                   COALESCE(SUM(income_amount), 0)::float as total_income,
                   COALESCE(SUM(CASE WHEN is_taxable THEN income_amount ELSE 0 END), 0)::float as taxable_income,
                   COALESCE(SUM(CASE WHEN NOT is_taxable THEN income_amount ELSE 0 END), 0)::float as exempt_income,
                   COALESCE(SUM(cgst), 0)::float as total_cgst,
                   COALESCE(SUM(sgst), 0)::float as total_sgst,
                   COALESCE(SUM(igst), 0)::float as total_igst,
                   COALESCE(SUM(cgst + sgst + igst), 0)::float as total_output_gst
            FROM income_entries
            WHERE client_id = %s
        '''
        params = [client_id]
        if not is_admin:
            q_inc += " AND user_id = %s"
            params.append(user_id)
        if fy and fy != 'ALL':
            q_inc += " AND financial_year = %s"
            params.append(fy)
        if month and month != 'ALL':
            q_inc += " AND month = %s"
            params.append(month)

        cur.execute(q_inc, params)
        inc_stat = cur.fetchone()

        q_itc = '''
            SELECT COALESCE(SUM(eligible_itc), 0)::float as eligible_itc,
                   COALESCE(SUM(ineligible_itc), 0)::float as ineligible_itc,
                   COALESCE(SUM(cgst), 0)::float as itc_cgst,
                   COALESCE(SUM(sgst), 0)::float as itc_sgst,
                   COALESCE(SUM(igst), 0)::float as itc_igst
            FROM invoices
            WHERE client_id = %s
        '''
        params_itc = [client_id]
        if not is_admin:
            q_itc += " AND user_id = %s"
            params_itc.append(user_id)
        if fy and fy != 'ALL':
            q_itc += " AND financial_year = %s"
            params_itc.append(fy)
        if month and month != 'ALL':
            q_itc += " AND month = %s"
            params_itc.append(month)

        cur.execute(q_itc, params_itc)
        itc_stat = cur.fetchone()

        q_branch = '''
            SELECT branch,
                   COUNT(*) as record_count,
                   COALESCE(SUM(income_amount), 0)::float as branch_income,
                   COALESCE(SUM(CASE WHEN is_taxable THEN income_amount ELSE 0 END), 0)::float as branch_taxable,
                   COALESCE(SUM(CASE WHEN NOT is_taxable THEN income_amount ELSE 0 END), 0)::float as branch_exempt,
                   COALESCE(SUM(cgst), 0)::float as branch_cgst,
                   COALESCE(SUM(sgst), 0)::float as branch_sgst,
                   COALESCE(SUM(igst), 0)::float as branch_igst,
                   COALESCE(SUM(cgst + sgst + igst), 0)::float as branch_gst
            FROM income_entries
            WHERE client_id = %s
        '''
        params_br = [client_id]
        if not is_admin:
            q_branch += " AND user_id = %s"
            params_br.append(user_id)
        if fy and fy != 'ALL':
            q_branch += " AND financial_year = %s"
            params_br.append(fy)
        if month and month != 'ALL':
            q_branch += " AND month = %s"
            params_br.append(month)
        q_branch += " GROUP BY branch ORDER BY branch ASC"

        cur.execute(q_branch, params_br)
        branch_stats = cur.fetchall()

        cur.close()
        conn.close()

        output_gst = inc_stat['total_output_gst'] if inc_stat else 0.0
        eligible_itc = itc_stat['eligible_itc'] if itc_stat else 0.0
        net_gst_payable = max(0.0, round(output_gst - eligible_itc, 2))

        return jsonify({
            "income": inc_stat,
            "itc": itc_stat,
            "net_gst_payable": net_gst_payable,
            "branches": branch_stats
        })
    except Exception as e:
        print(f"Error generating income summary: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/income-file/<int:entry_id>', methods=['GET'])
@login_required
def get_income_file(entry_id):
    client_id = get_current_client_id()
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('SELECT file_data, file_name FROM income_entries WHERE id = %s AND client_id = %s', (entry_id, client_id))
        row = cur.fetchone()
        cur.close()
        conn.close()

        if not row or row[0] is None:
            return jsonify({"error": "No statement file attached to this income entry"}), 404

        file_data, file_name = row
        fname = file_name or f"income-statement-{entry_id}.pdf"
        ext = fname.lower().split('.')[-1] if '.' in fname else 'pdf'
        mimetypes = {
            'pdf': 'application/pdf',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xls': 'application/vnd.ms-excel',
            'csv': 'text/csv',
            'png': 'image/png',
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg'
        }
        mime = mimetypes.get(ext, 'application/octet-stream')

        return send_file(
            io.BytesIO(bytes(file_data)),
            mimetype=mime,
            as_attachment=False,
            download_name=fname
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete-income-entry', methods=['POST'])
@login_required
def delete_income_entry():
    client_id = get_current_client_id()
    data = request.json or {}
    entry_id = data.get('id')
    if not entry_id:
        return jsonify({"error": "Entry ID required"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM income_entries WHERE id = %s AND client_id = %s', (entry_id, client_id))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/delete-income-batch', methods=['POST'])
@login_required
def delete_income_batch():
    client_id = get_current_client_id()
    data = request.json or {}
    ids = data.get('ids')
    if not ids or not isinstance(ids, list):
        return jsonify({"error": "No entries selected"}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute('DELETE FROM income_entries WHERE id = ANY(%s) AND client_id = %s', (ids, client_id))
        deleted_count = cur.rowcount
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True, "deleted_count": deleted_count})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/export-income-working-sheet', methods=['POST', 'GET'])
@login_required
def export_income_working_sheet():
    import collections, openpyxl, io, re
    from openpyxl.cell.cell import MergedCell

    def _set(ws, r, c, value):
        """Write a cell, silently skipping merged-range member cells (only
        the top-left cell of a merge is writable in openpyxl)."""
        cell = ws.cell(r, c)
        if isinstance(cell, MergedCell):
            return
        cell.value = value

    month = request.args.get('month', 'July')
    fy = request.args.get('financial_year', '2026-27')
    client_id = get_current_client_id()
    template_path = os.path.join(os.path.dirname(__file__), 'reference_data', 'MASTER_BRANCH_WISE_CALCULATION_TEMPLATE.xlsx')

    try:
        conn = get_db_connection()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute('''
            SELECT branch, gl_code, particulars, is_taxable,
                   income_amount::float, cgst::float, sgst::float, igst::float,
                   refund_without_gst::float, refund_with_gst::float
            FROM income_entries
            WHERE client_id = %s
            ORDER BY branch ASC, gl_code ASC
        ''', (client_id,))
        entries = cur.fetchall()
        cur.close()
        conn.close()

        # Group entries by normalized branch and GL code
        branch_map = collections.defaultdict(dict)
        for e in entries:
            b_name = re.sub(r'[^A-Z0-9]', '', e['branch'].upper().strip())
            c_code = str(e['gl_code']).strip()
            branch_map[b_name][c_code] = e

        if os.path.exists(template_path):
            try:
                # Load with data_only=False to preserve layout & formulas
                wb = openpyxl.load_workbook(template_path)
                wb.calculation.fullCalcOnLoad = True

                # Remove duplicate trailing tabs (ODHAV1, SURAT1, etc.)
                for sname in list(wb.sheetnames):
                    if sname.endswith('1') and sname not in ['Sheet1'] and sname[:-1].strip().upper() in INCOME_MASTER_BRANCHES:
                        try:
                            wb.remove(wb[sname])
                        except Exception:
                            pass

                # Update branch sheets with uploaded statement numbers
                grand_income = grand_ggst = grand_cgst = grand_igst = 0.0
                grand_ref_wo = grand_ref_w = 0.0
                code_row_re = re.compile(r'^(?:PL|GL)?\s*(\d{3,6})$', re.IGNORECASE)
                code_level = collections.defaultdict(lambda: {
                    'inc': 0.0, 'sgst': 0.0, 'cgst': 0.0, 'igst': 0.0, 'refwo': 0.0, 'refw': 0.0,
                })
                # Sheet1 is titled "ALL BRANCHES EXCEPT HO AND DEMAT" - keep a
                # separate rollup that excludes those two for it specifically.
                s1_code_level = collections.defaultdict(lambda: {
                    'inc': 0.0, 'sgst': 0.0, 'cgst': 0.0, 'igst': 0.0, 'refwo': 0.0, 'refw': 0.0,
                })
                s1_grand = {'inc': 0.0, 'sgst': 0.0, 'cgst': 0.0, 'igst': 0.0, 'refwo': 0.0}

                for sname in wb.sheetnames:
                    if sname in ['SUMMARY SHEET GST', 'Notes', 'GSTN CANCELLTED', 'Sheet1'] or sname.endswith('1'):
                        continue

                    ws_b = wb[sname]
                    b_norm = re.sub(r'[^A-Z0-9]', '', sname.upper().strip())

                    matched_dict = None
                    for b_key in branch_map:
                        if b_key == b_norm or b_key in b_norm or b_norm in b_key:
                            matched_dict = branch_map[b_key]
                            break
                    matched_dict = matched_dict or {}
                    total_row = {'HO': 57, 'DEMAT': 31}.get(sname.strip().upper(), 62)

                    b_tot_inc = 0.0
                    b_tot_ggst = 0.0
                    b_tot_cgst = 0.0
                    b_tot_igst = 0.0
                    b_tot_ref_wo = 0.0
                    b_tot_ref_w = 0.0

                    for r in range(7, total_row):
                        c_val = ws_b.cell(r, 1).value
                        if c_val is None:
                            continue
                        c_str = str(c_val).strip()
                        m_code = code_row_re.match(c_str)
                        if not m_code:
                            continue  # section-header / label row, not a GL-code data row
                        norm_code = m_code.group(1)

                        item = matched_dict.get(c_str) or matched_dict.get(norm_code)
                        if item:
                            inc_amt = float(item.get('income_amount') or 0.0)
                            sgst_amt = float(item.get('sgst') or 0.0)
                            cgst_amt = float(item.get('cgst') or 0.0)
                            igst_amt = float(item.get('igst') or 0.0)
                            ref_wo = float(item.get('refund_without_gst') or 0.0)
                            ref_w = float(item.get('refund_with_gst') or 0.0)
                        else:
                            # No ingested data for this code - write zero rather than
                            # leaving whatever the template happened to already hold.
                            inc_amt = sgst_amt = cgst_amt = igst_amt = ref_wo = ref_w = 0.0

                        _set(ws_b, r, 3, inc_amt)
                        _set(ws_b, r, 4, sgst_amt)
                        _set(ws_b, r, 5, cgst_amt)
                        _set(ws_b, r, 6, igst_amt if igst_amt else None)
                        _set(ws_b, r, 7, ref_wo if ref_wo else None)
                        _set(ws_b, r, 8, ref_w if ref_w else None)

                        b_tot_inc += inc_amt
                        b_tot_ggst += sgst_amt
                        b_tot_cgst += cgst_amt
                        b_tot_igst += igst_amt
                        b_tot_ref_wo += ref_wo
                        b_tot_ref_w += ref_w

                        cl = code_level[norm_code]
                        cl['inc'] += inc_amt
                        cl['sgst'] += sgst_amt
                        cl['cgst'] += cgst_amt
                        cl['igst'] += igst_amt
                        cl['refwo'] += ref_wo
                        cl['refw'] += ref_w

                        if sname.strip().upper() not in ('HO', 'DEMAT'):
                            s1cl = s1_code_level[norm_code]
                            s1cl['inc'] += inc_amt
                            s1cl['sgst'] += sgst_amt
                            s1cl['cgst'] += cgst_amt
                            s1cl['igst'] += igst_amt
                            s1cl['refwo'] += ref_wo
                            s1cl['refw'] += ref_w

                    _set(ws_b, total_row, 3, b_tot_inc)
                    _set(ws_b, total_row, 4, b_tot_ggst)
                    _set(ws_b, total_row, 5, b_tot_cgst)
                    _set(ws_b, total_row, 6, b_tot_igst if b_tot_igst else None)
                    _set(ws_b, total_row, 7, b_tot_ref_wo if b_tot_ref_wo else None)
                    _set(ws_b, total_row, 8, b_tot_ref_w if b_tot_ref_w else None)

                    grand_income += b_tot_inc
                    grand_ggst += b_tot_ggst
                    grand_cgst += b_tot_cgst
                    grand_igst += b_tot_igst
                    grand_ref_wo += b_tot_ref_wo
                    grand_ref_w += b_tot_ref_w

                    if sname.strip().upper() not in ('HO', 'DEMAT'):
                        s1_grand['inc'] += b_tot_inc
                        s1_grand['sgst'] += b_tot_ggst
                        s1_grand['cgst'] += b_tot_cgst
                        s1_grand['igst'] += b_tot_igst
                        s1_grand['refwo'] += b_tot_ref_wo

                if 'Sheet1' in wb.sheetnames:
                    # Sheet1 is a bank-wide (all branches except HO & DEMAT), GL-code
                    # level roll-up - the same code_level totals accumulated above,
                    # written the same zero-if-absent way as each branch sheet so no
                    # template-contaminated figure can survive.
                    ws_s1 = wb['Sheet1']
                    s1_total = {'inc': 0.0, 'sgst': 0.0, 'cgst': 0.0, 'igst': 0.0, 'refwo': 0.0}
                    for r in range(9, 68):
                        c_val = ws_s1.cell(r, 1).value
                        if c_val is None:
                            continue
                        m_code = code_row_re.match(str(c_val).strip())
                        if not m_code:
                            continue
                        cl = s1_code_level.get(m_code.group(1))
                        inc_amt = cl['inc'] if cl else 0.0
                        sgst_amt = cl['sgst'] if cl else 0.0
                        cgst_amt = cl['cgst'] if cl else 0.0
                        igst_amt = cl['igst'] if cl else 0.0
                        refwo_amt = cl['refwo'] if cl else 0.0
                        _set(ws_s1, r, 3, inc_amt)
                        _set(ws_s1, r, 4, sgst_amt)
                        _set(ws_s1, r, 5, cgst_amt)
                        _set(ws_s1, r, 6, igst_amt if igst_amt else None)
                        _set(ws_s1, r, 7, refwo_amt if refwo_amt else None)
                        s1_total['inc'] += inc_amt
                        s1_total['sgst'] += sgst_amt
                        s1_total['cgst'] += cgst_amt
                        s1_total['igst'] += igst_amt
                        s1_total['refwo'] += refwo_amt

                    _set(ws_s1, 64, 3, s1_total['inc'])
                    _set(ws_s1, 64, 4, s1_total['sgst'])
                    _set(ws_s1, 64, 5, s1_total['cgst'])
                    _set(ws_s1, 64, 6, s1_total['igst'] if s1_total['igst'] else None)
                    _set(ws_s1, 64, 7, s1_total['refwo'] if s1_total['refwo'] else None)

                    for rr in range(73, 76):
                        for cc in range(3, 8):
                            _set(ws_s1, rr, cc, None)
                    _set(ws_s1, 73, 3, s1_grand['inc'])
                    _set(ws_s1, 73, 4, s1_grand['sgst'])
                    _set(ws_s1, 73, 5, s1_grand['cgst'])
                    _set(ws_s1, 73, 6, s1_grand['igst'])
                    ref_gst_s1 = round(s1_grand['refwo'] * 0.09, 2)
                    _set(ws_s1, 74, 3, s1_grand['refwo'])
                    _set(ws_s1, 74, 4, ref_gst_s1)
                    _set(ws_s1, 74, 5, ref_gst_s1)
                    _set(ws_s1, 75, 3, s1_grand['inc'] + s1_grand['refwo'])
                    _set(ws_s1, 75, 4, s1_grand['sgst'] + ref_gst_s1)
                    _set(ws_s1, 75, 5, s1_grand['cgst'] + ref_gst_s1)

                # SUMMARY SHEET GST - computed from real ingested data, not hardcoded.
                if 'SUMMARY SHEET GST' in wb.sheetnames:
                    ws_sum = wb['SUMMARY SHEET GST']

                    # The master template's SUMMARY SHEET GST cells were previously
                    # pre-filled with the CA's golden figures as literal numbers, not
                    # computed. Clear all of them first so nothing stale survives
                    # under the genuinely-computed values written below.
                    for rr in range(7, 33):
                        for cc in range(3, 10):
                            _set(ws_sum, rr, cc, None)

                    # Section 1: Non-Taxable / Exempt Income (rows 7-21, "as on"
                    # columns D/E/F). These are bank-wide ledger/trial-balance items
                    # (interest income exempted, dividend, profit on sale of
                    # investments, provisions written back, etc.) - this app has no
                    # source document for them (they aren't branch commission GL
                    # codes), so only the total (row 21) is summed from any
                    # income_entries explicitly marked is_taxable=False, rather than
                    # left as stale copied figures. Reads 0 until such entries exist.
                    exempt_total = 0.0
                    for e in entries:
                        if not e.get('is_taxable'):
                            exempt_total += float(e.get('income_amount') or 0.0)
                    _set(ws_sum, 21, 4, exempt_total)

                    # Section 2: Taxable income, rolled up from the branch totals
                    # actually computed above - not copied from the golden file.
                    _set(ws_sum, 27, 3, grand_income)
                    _set(ws_sum, 27, 4, grand_ggst)
                    _set(ws_sum, 27, 5, grand_cgst)
                    _set(ws_sum, 27, 6, grand_igst)
                    _set(ws_sum, 27, 7, grand_ggst + grand_cgst + grand_igst)

                    _set(ws_sum, 29, 3, grand_ref_wo)
                    ref_gst = round(grand_ref_wo * 0.09, 2)
                    _set(ws_sum, 29, 4, ref_gst)
                    _set(ws_sum, 29, 5, ref_gst)
                    _set(ws_sum, 29, 7, ref_gst * 2)

                    tot_inc = grand_income + grand_ref_wo
                    tot_ggst = grand_ggst + ref_gst
                    tot_cgst = grand_cgst + ref_gst
                    _set(ws_sum, 30, 3, tot_inc)
                    _set(ws_sum, 30, 4, tot_ggst)
                    _set(ws_sum, 30, 5, tot_cgst)
                    _set(ws_sum, 30, 6, grand_igst)
                    _set(ws_sum, 30, 7, tot_ggst + tot_cgst + grand_igst)

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                out_filename = f"1_BRANCH_WISE_CALCULATION_{month.upper()}_{fy}_WORKING_SHEET.xlsx"
                return send_file(
                    buf,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    as_attachment=True,
                    download_name=out_filename
                )
            except Exception as te:
                print(f"Master template loading error, falling back to dynamic builder: {te}")

        # 2. Dynamic High-Fidelity 27-Tab Builder (Fallback)
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "SUMMARY SHEET GST"
        
        title_font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
        section_font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        num_font = Font(name="Calibri", size=10)
        bold_num_font = Font(name="Calibri", size=10, bold=True)

        ws_summary['B1'] = f"{client_cfg.get('name', 'Nutan Nagrik Sahakari Bank Ltd.')}"
        ws_summary['B1'].font = title_font
        ws_summary['B2'] = f"GST CALCULATION SUMMARY FOR THE MONTH OF {month.upper()} {fy}"
        ws_summary['B2'].font = section_font

        ws_summary['B4'] = "(1) NON TAXABLE INCOME / EXEMPT INCOME"
        ws_summary['B4'].font = section_font
        headers_exempt = ["PARTICULARS", "INCOME AMOUNT (₹)", "TAX RATE", "EXEMPTION STATUS"]
        for c_idx, h in enumerate(headers_exempt, start=2):
            cell = ws_summary.cell(row=5, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        exempt_total = 0.0
        taxable_total = 0.0
        cgst_total = 0.0
        sgst_total = 0.0
        igst_total = 0.0

        r_ptr = 6
        for e in entries:
            if not e['is_taxable']:
                ws_summary.cell(row=r_ptr, column=2, value=e['particulars']).font = num_font
                ws_summary.cell(row=r_ptr, column=3, value=e['income_amount']).font = num_font
                ws_summary.cell(row=r_ptr, column=4, value="0%").font = num_font
                ws_summary.cell(row=r_ptr, column=5, value="Exempt (Notification 12/2017)").font = num_font
                exempt_total += e['income_amount']
                r_ptr += 1

        if r_ptr == 6:
            ws_summary.cell(row=6, column=2, value="No direct exempt items recorded").font = num_font
            r_ptr = 7

        ws_summary.cell(row=r_ptr, column=2, value="TOTAL EXEMPT INCOME").font = bold_num_font
        ws_summary.cell(row=r_ptr, column=3, value=exempt_total).font = bold_num_font
        r_ptr += 2

        ws_summary.cell(row=r_ptr, column=2, value="(2) TAXABLE INCOME & OUTPUT GST").font = section_font
        r_ptr += 1
        headers_tax = ["BRANCH", "TAXABLE INCOME (₹)", "CGST 9% (₹)", "SGST 9% (₹)", "IGST 18% (₹)", "TOTAL GST (₹)"]
        for c_idx, h in enumerate(headers_tax, start=2):
            cell = ws_summary.cell(row=r_ptr, column=c_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        r_ptr += 1
        for b_name in INCOME_MASTER_BRANCHES:
            b_entries = branch_entries_list.get(b_name, [])
            b_tax = sum(e['income_amount'] for e in b_entries if e['is_taxable'])
            b_cgst = sum(e['cgst'] for e in b_entries if e['is_taxable'])
            b_sgst = sum(e['sgst'] for e in b_entries if e['is_taxable'])
            b_igst = sum(e['igst'] for e in b_entries if e['is_taxable'])
            b_tot_gst = b_cgst + b_sgst + b_igst

            ws_summary.cell(row=r_ptr, column=2, value=b_name).font = num_font
            ws_summary.cell(row=r_ptr, column=3, value=b_tax).font = num_font
            ws_summary.cell(row=r_ptr, column=4, value=b_cgst).font = num_font
            ws_summary.cell(row=r_ptr, column=5, value=b_sgst).font = num_font
            ws_summary.cell(row=r_ptr, column=6, value=b_igst).font = num_font
            ws_summary.cell(row=r_ptr, column=7, value=b_tot_gst).font = num_font

            taxable_total += b_tax
            cgst_total += b_cgst
            sgst_total += b_sgst
            igst_total += b_igst
            r_ptr += 1

        ws_summary.cell(row=r_ptr, column=2, value="CONSOLIDATED TOTAL").font = bold_num_font
        ws_summary.cell(row=r_ptr, column=3, value=taxable_total).font = bold_num_font
        ws_summary.cell(row=r_ptr, column=4, value=cgst_total).font = bold_num_font
        ws_summary.cell(row=r_ptr, column=5, value=sgst_total).font = bold_num_font
        ws_summary.cell(row=r_ptr, column=6, value=igst_total).font = bold_num_font
        ws_summary.cell(row=r_ptr, column=7, value=cgst_total + sgst_total + igst_total).font = bold_num_font

        for b_name in INCOME_MASTER_BRANCHES:
            ws_b = wb.create_sheet(title=b_name[:31])
            ws_b['A1'] = f"{client_cfg.get('name', 'Nutan Nagrik Sahakari Bank Ltd.')}"
            ws_b['A1'].font = title_font
            ws_b['A3'] = f"BRANCH NAME : {b_name} BRANCH"
            ws_b['A3'].font = section_font
            ws_b['A5'] = f"SUMMARY OF INCOME FOR THE MONTH OF {month.upper()} {fy}"
            ws_b['A5'].font = num_font

            headers_b = ["CODE NO.", "PARTICULARS", "TOTAL INCOME", "GGST (9%)", "CGST (9%)", "IGST (18%)", "REFUND WITHOUT GST", "REFUND WITH GST"]
            for col_i, h_text in enumerate(headers_b, start=1):
                c = ws_b.cell(row=7, column=col_i, value=h_text)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")

            b_list = branch_entries_list.get(b_name, [])
            e_by_code = {e['gl_code'].strip(): e for e in b_list}

            b_row = 8
            for m_item in INCOME_MASTER_CODES:
                c_code = m_item['code']
                c_part = m_item['particulars']
                rec = e_by_code.get(c_code)
                
                t_amt = rec['income_amount'] if rec else 0.0
                c_cgst = rec['cgst'] if rec else (t_amt * 0.09 if m_item.get('is_taxable') else 0.0)
                c_sgst = rec['sgst'] if rec else (t_amt * 0.09 if m_item.get('is_taxable') else 0.0)
                c_igst = rec['igst'] if rec else 0.0

                ws_b.cell(row=b_row, column=1, value=c_code).font = num_font
                ws_b.cell(row=b_row, column=2, value=c_part).font = num_font
                ws_b.cell(row=b_row, column=3, value=t_amt if t_amt > 0 else "").font = num_font
                ws_b.cell(row=b_row, column=4, value=c_sgst if c_sgst > 0 else 0).font = num_font
                ws_b.cell(row=b_row, column=5, value=c_cgst if c_cgst > 0 else 0).font = num_font
                ws_b.cell(row=b_row, column=6, value=c_igst if c_igst > 0 else "").font = num_font
                ws_b.cell(row=b_row, column=7, value="").font = num_font
                ws_b.cell(row=b_row, column=8, value="").font = num_font
                b_row += 1

            ws_b.cell(row=b_row, column=2, value="TOTAL").font = bold_num_font
            ws_b.cell(row=b_row, column=3, value=f"=SUM(C8:C{b_row-1})").font = bold_num_font
            ws_b.cell(row=b_row, column=4, value=f"=SUM(D8:D{b_row-1})").font = bold_num_font
            ws_b.cell(row=b_row, column=5, value=f"=SUM(E8:E{b_row-1})").font = bold_num_font
            ws_b.cell(row=b_row, column=6, value=f"=SUM(F8:F{b_row-1})").font = bold_num_font

        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        out_filename = f"1_BRANCH_WISE_CALCULATION_{month.upper()}_{fy}_WORKING_SHEET.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=out_filename
        )
    except Exception as e:
        print(f"Error exporting working sheet: {e}")
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    # Ensure static and template folders exist
    os.makedirs(app.static_folder, exist_ok=True)
    os.makedirs(app.template_folder, exist_ok=True)
    
    # Initialize DB Tables
    init_db()
    
    # Run server
    port = int(os.getenv("PORT", 5588))
    print(f"Starting GST Calculation Server on port {port}...")
    app.run(host='0.0.0.0', port=port, debug=True)
